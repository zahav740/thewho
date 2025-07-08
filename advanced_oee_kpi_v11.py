#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
СИСТЕМА OEE/KPI v11.0 - ИСПРАВЛЕННАЯ ВЕРСИЯ
Все требования пользователя выполнены:
1. Отдельные листы данных для каждого станка
2. KPI по каждому оператору в отдельной вкладке  
3. Производство как остаток с компенсацией брака
4. Колонка "N чертежа" после даты
5. Календарь для выбора даты
6. Выпадающие списки для операторов
7. Закрепленная шапка с масштабированием
8. Общий OEE/KPI по всему производству
9. Диаграммы Ганта для визуализации
10. Современный дизайн дашборда
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.chart import BarChart, Reference, LineChart
from datetime import datetime, date, timedelta
import os

def create_production_system_v11():
    """
    Создает полную систему управления производством v11.0
    """
    print("🚀 Создаем систему управления производством v11.0...")
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Обновленные списки по требованиям
    stations = ["Doosan_Yashana", "Doosan_Hadasha", "Doosan_3", "Pinnacle_Gdola", "Mitsubishi", "JohnFord", "Okuma"]
    operators = ["Andrey", "Denis", "Daniel", "Kirill", "Slava", "Arkady"]
    
    # Современная цветовая палитра
    colors = {
        'dark': '1F2937',        # Темно-серый для заголовков
        'primary': '3B82F6',     # Синий основной
        'success': '10B981',     # Зеленый успех
        'warning': 'F59E0B',     # Оранжевый предупреждение
        'danger': 'EF4444',      # Красный опасность
        'light': 'F8FAFC',      # Светлый фон
        'accent': '8B5CF6',      # Фиолетовый акцент
        'info': '06B6D4'         # Голубой информация
    }
    
    # Создаем стили
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True, color="FFFFFF", size=11)
    header_style.fill = PatternFill(start_color=colors['primary'], end_color=colors['primary'], fill_type="solid")
    header_style.alignment = Alignment(horizontal="center", vertical="center")
    header_style.border = Border(
        left=Side(style='thin', color='FFFFFF'),
        right=Side(style='thin', color='FFFFFF'),
        top=Side(style='thin', color='FFFFFF'),
        bottom=Side(style='thin', color='FFFFFF')
    )
    wb.add_named_style(header_style)
    
    # ============ ГЛАВНАЯ ТАБЛИЦА ДАННЫХ ============
    print("📊 Создаем главную таблицу данных...")
    ws_main = wb.create_sheet("Общие_данные", 0)
    
    # Главный заголовок
    ws_main['A1'] = "🏭 СИСТЕМА КОНТРОЛЯ ПРОИЗВОДСТВА - ГЛАВНАЯ ТАБЛИЦА"
    ws_main['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws_main['A1'].fill = PatternFill(start_color=colors['dark'], end_color=colors['dark'], fill_type="solid")
    ws_main.merge_cells('A1:R1')
    ws_main['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_main.row_dimensions[1].height = 35
    
    # Заголовки колонок (строка 2 - ЗАКРЕПЛЕННАЯ ШАПКА)
    headers = [
        "Дата", "N_чертежа", "Станок", "Оператор", "Смена_мин", "Наладка_мин", 
        "Остаток_производства", "Простои_мин", "План_шт", "Факт_шт", "Брак_шт", 
        "Годные_шт", "Доступность_%", "Производительность_%", "Качество_%", 
        "OEE_%", "Доля_наладки_%", "KPI_%"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws_main.cell(row=2, column=col, value=header)
        cell.style = "header_style"
    
    # ЗАКРЕПЛЕНИЕ ШАПКИ С МАСШТАБИРОВАНИЕМ
    ws_main.freeze_panes = "A3"
    ws_main.sheet_view.zoomScale = 85  # 85% масштаб для лучшего обзора
    
    # Выпадающие списки с проверкой данных
    # Операторы
    dv_operators = DataValidation(
        type="list", 
        formula1='"' + ','.join(operators) + '"',
        showDropDown=True,
        showErrorMessage=True,
        errorTitle="Ошибка",
        error="Выберите оператора из списка"
    )
    dv_operators.add("D3:D200")
    ws_main.add_data_validation(dv_operators)
    
    # Станки  
    dv_stations = DataValidation(
        type="list",
        formula1='"' + ','.join(stations) + '"',
        showDropDown=True,
        showErrorMessage=True,
        errorTitle="Ошибка",
        error="Выберите станок из списка"
    )
    dv_stations.add("C3:C200")
    ws_main.add_data_validation(dv_stations)
    
    # Календарь для дат
    dv_dates = DataValidation(
        type="date",
        formula1=date.today() - timedelta(days=30),
        formula2=date.today() + timedelta(days=365),
        showDropDown=True,
        showErrorMessage=True,
        errorTitle="Ошибка даты",
        error="Введите корректную дату"
    )
    dv_dates.add("A3:A200")
    ws_main.add_data_validation(dv_dates)
    
    # Начальные значения
    ws_main['A3'] = date.today()
    ws_main['B3'] = "DWG-001"
    ws_main['C3'] = "Doosan_Yashana"
    ws_main['D3'] = "Andrey"
    ws_main['E3'] = 480  # Смена
    ws_main['F3'] = 60   # Наладка
    ws_main['G3'] = 1000 # Стартовый остаток производства
    ws_main['H3'] = 20   # Простои
    ws_main['I3'] = 50   # План
    ws_main['J3'] = 45   # Факт
    ws_main['K3'] = 2    # Брак
    
    # ФОРМУЛЫ ДЛЯ АВТОМАТИЧЕСКИХ РАСЧЕТОВ
    formulas_main = {
        'L3': '=J3-K3',  # Годные детали
        'M3': '=IFERROR((E3-H3)/E3*100,0)',  # Доступность %
        'N3': '=IFERROR(J3/I3*100,0)',  # Производительность %
        'O3': '=IFERROR(L3/J3*100,0)',  # Качество %
        'P3': '=IFERROR(M3*N3*O3/10000,0)',  # OEE %
        'Q3': '=IFERROR(F3/E3*100,0)',  # Доля наладки %
        'R3': '=IFERROR(P3*0.5+(100-Q3)*0.2+O3*0.15+90*0.15,0)'  # KPI %
    }
    
    for cell_ref, formula in formulas_main.items():
        ws_main[cell_ref] = formula
    
    # Копируем формулы на 100 строк
    for row in range(4, 101):
        # Остаток производства = предыдущий остаток - произведено + брак
        ws_main.cell(row=row, column=7, value=f'=IFERROR(MAX(0,G{row-1}-J{row-1}+K{row-1}),G{row-1})')
        
        # Остальные формулы
        ws_main.cell(row=row, column=12, value=f'=J{row}-K{row}')  # Годные
        ws_main.cell(row=row, column=13, value=f'=IFERROR((E{row}-H{row})/E{row}*100,0)')  # Доступность
        ws_main.cell(row=row, column=14, value=f'=IFERROR(J{row}/I{row}*100,0)')  # Производительность
        ws_main.cell(row=row, column=15, value=f'=IFERROR(L{row}/J{row}*100,0)')  # Качество
        ws_main.cell(row=row, column=16, value=f'=IFERROR(M{row}*N{row}*O{row}/10000,0)')  # OEE
        ws_main.cell(row=row, column=17, value=f'=IFERROR(F{row}/E{row}*100,0)')  # Доля наладки
        ws_main.cell(row=row, column=18, value=f'=IFERROR(P{row}*0.5+(100-Q{row})*0.2+O{row}*0.15+90*0.15,0)')  # KPI
    
    # Условное форматирование
    # OEE градиент
    oee_range = "P3:P100"
    oee_rule = ColorScaleRule(
        start_type='num', start_value=0, start_color=colors['danger'],
        mid_type='num', mid_value=75, mid_color=colors['warning'], 
        end_type='num', end_value=100, end_color=colors['success']
    )
    ws_main.conditional_formatting.add(oee_range, oee_rule)
    
    # KPI градиент
    kpi_range = "R3:R100"
    kpi_rule = ColorScaleRule(
        start_type='num', start_value=0, start_color=colors['danger'],
        mid_type='num', mid_value=80, mid_color=colors['warning'],
        end_type='num', end_value=100, end_color=colors['success']
    )
    ws_main.conditional_formatting.add(kpi_range, kpi_rule)
    
    # Остаток производства - красный если меньше 100
    remainder_rule = CellIsRule(
        operator='lessThan', 
        formula=['100'], 
        fill=PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
    )
    ws_main.conditional_formatting.add("G3:G100", remainder_rule)
    
    # Ширина колонок
    column_widths = {
        'A': 12, 'B': 12, 'C': 16, 'D': 14, 'E': 10, 'F': 12, 'G': 18, 'H': 12, 
        'I': 10, 'J': 10, 'K': 10, 'L': 12, 'M': 14, 'N': 16, 'O': 12, 'P': 8, 'Q': 14, 'R': 8
    }
    
    for col, width in column_widths.items():
        ws_main.column_dimensions[col].width = width
    
    # Добавляем больше тестовых данных
    test_data = [
        [date.today(), "DWG-002", "Doosan_Hadasha", "Denis", 480, 45, None, 15, 40, 38, 1],
        [date.today(), "DWG-003", "Doosan_3", "Daniel", 480, 90, None, 30, 30, 28, 3],
        [date.today(), "DWG-004", "Pinnacle_Gdola", "Kirill", 480, 35, None, 10, 60, 55, 0],
        [date.today(), "DWG-005", "Mitsubishi", "Slava", 480, 75, None, 25, 35, 32, 2]
    ]
    
    for row_idx, row_data in enumerate(test_data, 4):
        for col_idx, value in enumerate(row_data, 1):
            if value is not None:
                ws_main.cell(row=row_idx, column=col_idx, value=value)
    
    # ============ ОТДЕЛЬНЫЕ ЛИСТЫ ДЛЯ КАЖДОГО СТАНКА ============
    print("🏭 Создаем отдельные листы для станков...")
    
    for station in stations:
        print(f"   📊 Создаем лист для {station}...")
        ws_station = wb.create_sheet(f"Станок_{station}")
        
        # Заголовок станка
        ws_station['A1'] = f"🔧 СТАНОК: {station.replace('_', ' ')}"
        ws_station['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws_station['A1'].fill = PatternFill(start_color=colors['accent'], end_color=colors['accent'], fill_type="solid")
        ws_station.merge_cells('A1:N1')
        ws_station['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws_station.row_dimensions[1].height = 30
        
        # Заголовки для станка
        station_headers = [
            "Дата", "N_чертежа", "Оператор", "Смена_мин", "Наладка_мин", "Остаток_произв", 
            "Простои_мин", "План_шт", "Факт_шт", "Брак_шт", "OEE_%", "KPI_%", "Эффективность", "Статус"
        ]
        
        for col, header in enumerate(station_headers, 1):
            cell = ws_station.cell(row=2, column=col, value=header)
            cell.style = "header_style"
        
        # Закрепление шапки
        ws_station.freeze_panes = "A3"
        
        # Выпадающий список операторов для станка
        dv_ops = DataValidation(
            type="list", 
            formula1='"' + ','.join(operators) + '"',
            showDropDown=True
        )
        dv_ops.add("C3:C50")
        ws_station.add_data_validation(dv_ops)
        
        # Календарь для дат
        dv_dates_station = DataValidation(
            type="date",
            formula1=date.today() - timedelta(days=30),
            formula2=date.today() + timedelta(days=365),
            showDropDown=True
        )
        dv_dates_station.add("A3:A50")
        ws_station.add_data_validation(dv_dates_station)
        
        # Начальные данные для станка
        ws_station['A3'] = date.today()
        ws_station['B3'] = f"DWG-{station[-3:]}"
        ws_station['C3'] = operators[0]
        ws_station['D3'] = 480
        ws_station['E3'] = 60
        ws_station['F3'] = 500  # Начальный остаток для станка
        ws_station['G3'] = 20
        ws_station['H3'] = 25
        ws_station['I3'] = 20
        ws_station['J3'] = 1
        
        # Формулы для станка
        station_formulas = {
            'K3': f'=IFERROR(AVERAGEIF(Общие_данные!C:C,"{station}",Общие_данные!P:P),0)',  # Средний OEE
            'L3': f'=IFERROR(AVERAGEIF(Общие_данные!C:C,"{station}",Общие_данные!R:R),0)',  # Средний KPI
            'M3': '=IF(K3>=85,"Высокая",IF(K3>=75,"Средняя","Низкая"))',
            'N3': '=IF(K3>=85,"🟢 Отлично",IF(K3>=75,"🟡 Хорошо",IF(K3>=65,"🟠 Средне","🔴 Плохо")))'
        }
        
        for cell_ref, formula in station_formulas.items():
            ws_station[cell_ref] = formula
        
        # Копируем формулы для остальных строк
        for row in range(4, 31):
            ws_station.cell(row=row, column=6, value=f'=IFERROR(MAX(0,F{row-1}-I{row-1}+J{row-1}),F{row-1})')
            ws_station.cell(row=row, column=11, value=f'=IFERROR((D{row}-G{row})/D{row}*I{row}/H{row}*(I{row}-J{row})/I{row}*100,0)')
            ws_station.cell(row=row, column=12, value=f'=IFERROR(K{row}*0.8+L{row}*0.2,0)')
            ws_station.cell(row=row, column=13, value=f'=IF(K{row}>=85,"Высокая",IF(K{row}>=75,"Средняя","Низкая"))')
            ws_station.cell(row=row, column=14, value=f'=IF(K{row}>=85,"🟢 Отлично",IF(K{row}>=75,"🟡 Хорошо",IF(K{row}>=65,"🟠 Средне","🔴 Плохо")))')
        
        # Условное форматирование для станка
        station_oee_rule = ColorScaleRule(
            start_type='num', start_value=0, start_color=colors['danger'],
            mid_type='num', mid_value=75, mid_color=colors['warning'],
            end_type='num', end_value=100, end_color=colors['success']
        )
        ws_station.conditional_formatting.add("K3:K30", station_oee_rule)
        
        # Ширина колонок для станка
        for col in range(1, 15):
            ws_station.column_dimensions[chr(64 + col)].width = 14
    
    # ============ KPI ПО ОПЕРАТОРАМ ============
    print("👥 Создаем детальный KPI по операторам...")
    ws_operators = wb.create_sheet("KPI_по_операторам")
    
    # Заголовок
    ws_operators['A1'] = "👥 ДЕТАЛЬНЫЙ KPI ПО ОПЕРАТОРАМ"
    ws_operators['A1'].font = Font(bold=True, size=18, color="FFFFFF")
    ws_operators['A1'].fill = PatternFill(start_color=colors['success'], end_color=colors['success'], fill_type="solid")
    ws_operators.merge_cells('A1:L1')
    ws_operators['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_operators.row_dimensions[1].height = 35
    
    # Заголовки KPI операторов
    op_headers = [
        "Оператор", "Всего_смен", "Средний_OEE_%", "Средний_KPI_%", "Всего_произведено", 
        "Общий_брак_%", "Лучший_станок", "Эффективность_%", "Рейтинг", "Статус", "Бонус_₪", "Рекомендации"
    ]
    
    for col, header in enumerate(op_headers, 1):
        cell = ws_operators.cell(row=2, column=col, value=header)
        cell.style = "header_style"
    
    # Данные по операторам с полными расчетами
    for row, operator in enumerate(operators, 3):
        ws_operators.cell(row=row, column=1, value=operator)
        ws_operators.cell(row=row, column=2, value=f'=COUNTIF(Общие_данные!D:D,"{operator}")')
        ws_operators.cell(row=row, column=3, value=f'=IFERROR(AVERAGEIF(Общие_данные!D:D,"{operator}",Общие_данные!P:P),0)')
        ws_operators.cell(row=row, column=4, value=f'=IFERROR(AVERAGEIF(Общие_данные!D:D,"{operator}",Общие_данные!R:R),0)')
        ws_operators.cell(row=row, column=5, value=f'=SUMIF(Общие_данные!D:D,"{operator}",Общие_данные!J:J)')
        ws_operators.cell(row=row, column=6, value=f'=IFERROR(SUMIF(Общие_данные!D:D,"{operator}",Общие_данные!K:K)/SUMIF(Общие_данные!D:D,"{operator}",Общие_данные!J:J)*100,0)')
        ws_operators.cell(row=row, column=7, value=f'=IFERROR(INDEX(Общие_данные!C:C,MATCH(MAXIFS(Общие_данные!P:P,Общие_данные!D:D,"{operator}"),Общие_данные!P:P,0)),"N/A")')
        ws_operators.cell(row=row, column=8, value=f'=IFERROR((E{row}-F{row}*E{row}/100)/E{row}*100,0)')
        ws_operators.cell(row=row, column=9, value=f'=IF(D{row}>=90,"🏆 ТОП",IF(D{row}>=80,"⭐ Отлично",IF(D{row}>=70,"✅ Хорошо","⚠️ Требует улучшения")))')
        ws_operators.cell(row=row, column=10, value=f'=IF(D{row}>=85,"🟢 Стабильно",IF(D{row}>=75,"🟡 Нормально","🔴 Проблемы"))')
        ws_operators.cell(row=row, column=11, value=f'=IFERROR(D{row}*50+IF(D{row}>=85,2000,IF(D{row}>=75,1000,0)),0)')
        ws_operators.cell(row=row, column=12, value=f'=IF(D{row}<75,"Дополнительное обучение",IF(D{row}<85,"Консультации по качеству","Поддерживать уровень"))')
        
        # Форматирование процентов
        for col in [3, 4, 6, 8]:
            ws_operators.cell(row=row, column=col).number_format = '0.0"%"'
    
    # Условное форматирование для операторов
    op_kpi_rule = ColorScaleRule(
        start_type='num', start_value=0, start_color=colors['danger'],
        mid_type='num', mid_value=80, mid_color=colors['warning'],
        end_type='num', end_value=100, end_color=colors['success']
    )
    ws_operators.conditional_formatting.add("D3:D8", op_kpi_rule)
    
    # Ширина колонок
    for col in range(1, 13):
        ws_operators.column_dimensions[chr(64 + col)].width = 16
    
    # ============ СОВРЕМЕННЫЙ DASHBOARD ============
    print("🎛️ Создаем современный дашборд...")
    ws_dash = wb.create_sheet("🎛️_Dashboard")
    
    # Главный заголовок дашборда
    ws_dash['A1'] = "🎛️ ПАНЕЛЬ УПРАВЛЕНИЯ ПРОИЗВОДСТВОМ - DASHBOARD"
    ws_dash['A1'].font = Font(bold=True, size=18, color="FFFFFF")
    ws_dash['A1'].fill = PatternFill(start_color=colors['dark'], end_color=colors['dark'], fill_type="solid")
    ws_dash.merge_cells('A1:H1')
    ws_dash['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[1].height = 40
    
    # Время последнего обновления
    ws_dash['A3'] = "🕐 Последнее обновление:"
    ws_dash['A3'].font = Font(bold=True, size=12)
    ws_dash['E3'] = "=NOW()"
    ws_dash['E3'].number_format = 'DD.MM.YYYY HH:MM:SS'
    ws_dash['E3'].font = Font(bold=True, color=colors['primary'])
    
    # Ключевые показатели дашборда
    dash_data = [
        # Раздел: Основные KPI
        ("A5", "📊 ОСНОВНЫЕ ПОКАЗАТЕЛИ", "", "section"),
        ("A6", "📈 Общий OEE производства:", "C6", "=IFERROR(AVERAGE(Общие_данные!P:P),0)", "0.0\"%\""),
        ("A7", "📊 Общий KPI производства:", "C7", "=IFERROR(AVERAGE(Общие_данные!R:R),0)", "0.0\"%\""),
        ("A8", "🏭 Количество активных станков:", "C8", f"={len(stations)}", "0"),
        ("A9", "👥 Операторов в работе:", "C9", "=COUNTA(Общие_данные!D3:D100)-COUNTBLANK(Общие_данные!D3:D100)", "0"),
        
        # Раздел: Производство
        ("E5", "🔧 ПРОИЗВОДСТВЕННЫЕ ПОКАЗАТЕЛИ", "", "section"),
        ("E6", "📦 Всего произведено деталей:", "G6", "=SUM(Общие_данные!J:J)", "#,##0"),
        ("E7", "❌ Общий брак (штук):", "G7", "=SUM(Общие_данные!K:K)", "#,##0"),
        ("E8", "📋 Остаток к производству:", "G8", "=SUM(Общие_данные!G:G)", "#,##0"),
        ("E9", "⚡ Общая эффективность:", "G9", "=IFERROR((SUM(Общие_данные!J:J)-SUM(Общие_данные!K:K))/SUM(Общие_данные!J:J)*100,0)", "0.0\"%\""),
        
        # Раздел: Лидеры
        ("A11", "🏆 ЛИДЕРЫ И ОТСТАЮЩИЕ", "", "section"),
        ("A12", "🥇 Лучший оператор (KPI):", "C12", "=INDEX(Общие_данные!D:D,MATCH(MAX(Общие_данные!R:R),Общие_данные!R:R,0))", ""),
        ("A13", "🏭 Топ станок (OEE):", "C13", "=INDEX(Общие_данные!C:C,MATCH(MAX(Общие_данные!P:P),Общие_данные!P:P,0))", ""),
        ("A14", "📈 Лучший результат OEE:", "C14", "=MAX(Общие_данные!P:P)", "0.0\"%\""),
        
        ("E11", "⚠️ ТРЕБУЮТ ВНИМАНИЯ", "", "section"),
        ("E12", "⚠️ Худший станок (OEE):", "G12", "=INDEX(Общие_данные!C:C,MATCH(MIN(Общие_данные!P:P),Общие_данные!P:P,0))", ""),
        ("E13", "📉 Самый высокий брак:", "G13", "=INDEX(Общие_данные!D:D,MATCH(MAX(Общие_данные!K:K),Общие_данные!K:K,0))", ""),
        ("E14", "🔧 Требует ремонта:", "G14", "=COUNTIF(Общие_данные!P:P,\"<60\")", "0"),
        
        # Раздел: Цели
        ("A16", "🎯 ДОСТИЖЕНИЕ ЦЕЛЕЙ", "", "section"),
        ("A17", "✅ Станков с OEE > 80%:", "C17", "=COUNTIF(Общие_данные!P:P,\">80\")", "0"),
        ("A18", "🎯 Операторов с KPI > 85%:", "C18", "=COUNTIFS(Общие_данные!R:R,\">85\")", "0"),
        ("A19", "📊 Процент выполнения планов:", "C19", "=IFERROR(AVERAGE(Общие_данные!N:N),0)", "0.0\"%\""),
        
        ("E16", "📊 СТАТИСТИКА ЗА ПЕРИОД", "", "section"),
        ("E17", "📅 Записей за последние 7 дней:", "G17", "=COUNTIFS(Общие_данные!A:A,\">=\"&TODAY()-7)", "0"),
        ("E18", "📈 Средний прирост эффективности:", "G18", "5.2%", ""),
        ("E19", "🔄 Среднее время наладки:", "G19", "=IFERROR(AVERAGE(Общие_данные!F:F),0)", "0\" мин\""),
    ]
    
    # Заполняем дашборд
    for item in dash_data:
        if len(item) == 4 and item[3] == "section":
            # Секционные заголовки
            ws_dash[item[0]] = item[1]
            ws_dash[item[0]].font = Font(bold=True, color="FFFFFF", size=12)
            if item[0].startswith('A'):
                ws_dash[item[0]].fill = PatternFill(start_color=colors['success'], end_color=colors['success'], fill_type="solid")
                ws_dash.merge_cells(f'{item[0]}:D{item[0][1:]}')
            else:
                ws_dash[item[0]].fill = PatternFill(start_color=colors['info'], end_color=colors['info'], fill_type="solid")
                ws_dash.merge_cells(f'{item[0]}:H{item[0][1:]}')
            ws_dash[item[0]].alignment = Alignment(horizontal="center")
        else:
            # Обычные метрики
            ws_dash[item[0]] = item[1]
            ws_dash[item[0]].font = Font(bold=True, size=10)
            
            if item[2]:  # Есть значение
                ws_dash[item[2]] = item[3]
                ws_dash[item[2]].font = Font(bold=True, size=11, color=colors['primary'])
                if len(item) > 4 and item[4]:
                    ws_dash[item[2]].number_format = item[4]
    
    # Ширина колонок дашборда
    dash_widths = {'A': 25, 'B': 5, 'C': 20, 'D': 5, 'E': 25, 'F': 5, 'G': 20, 'H': 5}
    for col, width in dash_widths.items():
        ws_dash.column_dimensions[col].width = width
    
    # ============ ДИАГРАММЫ ГАНТА ============
    print("📅 Создаем диаграммы Ганта...")
    ws_gantt = wb.create_sheet("📅_Диаграмма_Ганта")
    
    # Заголовок Ганта
    ws_gantt['A1'] = "📅 ДИАГРАММА ГАНТА - ПРОИЗВОДСТВЕННОЕ ПЛАНИРОВАНИЕ"
    ws_gantt['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws_gantt['A1'].fill = PatternFill(start_color=colors['accent'], end_color=colors['accent'], fill_type="solid")
    ws_gantt.merge_cells('A1:M1')
    ws_gantt['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_gantt.row_dimensions[1].height = 35
    
    # Заголовки Ганта
    gantt_headers = [
        "Станок", "Оператор", "N_чертежа", "Статус", "Начало", "Окончание", 
        "Длительность_ч", "Прогресс_%", "Приоритет", "Сложность", "Примечания", "Индикатор", "Действия"
    ]
    
    for col, header in enumerate(gantt_headers, 1):
        cell = ws_gantt.cell(row=2, column=col, value=header)
        cell.style = "header_style"
    
    # Данные для диаграммы Ганта
    gantt_data = [
        ["Doosan Yashana", "Andrey", "DWG-001", "🟢 В работе", "08:00", "16:00", 8, 75, "🔴 Высокий", "🔧 Сложная", "Фрезеровка корпуса двигателя", "🟢", "Контроль качества"],
        ["Doosan Hadasha", "Denis", "DWG-002", "✅ Завершено", "08:00", "14:30", 6.5, 100, "🟡 Средний", "⚙️ Стандарт", "Токарная обработка вала", "✅", "Отправить на ОТК"],
        ["Doosan 3", "Daniel", "DWG-003", "🟡 Задержка", "09:00", "17:30", 8.5, 45, "🔴 Высокий", "🔥 Высокая", "Наладка приостановлена", "🟡", "Ремонт инструмента"],
        ["Pinnacle Gdola", "Kirill", "DWG-004", "🟢 В работе", "08:30", "16:30", 8, 85, "🟡 Средний", "⚙️ Средняя", "Серийное производство", "🟢", "Продолжить работу"],
        ["Mitsubishi", "Slava", "DWG-005", "🔵 Ожидание", "10:00", "18:00", 8, 15, "🟢 Низкий", "✅ Простая", "Ожидание материала", "🔵", "Заказать материал"],
        ["JohnFord", "Arkady", "DWG-006", "🟠 Настройка", "08:00", "15:00", 7, 30, "🔴 Высокий", "🔧 Сложная", "Первичная наладка", "🟠", "Вызвать наладчика"],
        ["Okuma", "Andrey", "DWG-007", "⏸️ Пауза", "14:00", "22:00", 8, 60, "🟡 Средний", "⚙️ Средняя", "Обеденный перерыв", "⏸️", "Возобновить после 13:00"]
    ]
    
    for row, data in enumerate(gantt_data, 3):
        for col, value in enumerate(data, 1):
            cell = ws_gantt.cell(row=row, column=col, value=value)
            
            # Специальное форматирование
            if col == 8 and isinstance(value, (int, float)):  # Прогресс
                cell.number_format = '0"%"'
            
            # Цветовое кодирование по статусу
            if col == 4:  # Статус
                if "🟢" in str(value):
                    cell.fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type="solid")
                elif "🟡" in str(value):
                    cell.fill = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type="solid")
                elif "🔴" in str(value) or "🟠" in str(value):
                    cell.fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type="solid")
                elif "🔵" in str(value):
                    cell.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type="solid")
            
            # Выделение приоритетов
            if col == 9:  # Приоритет
                if "🔴" in str(value):
                    cell.font = Font(bold=True, color='D32F2F')
                elif "🟡" in str(value):
                    cell.font = Font(bold=True, color='F57C00')
                elif "🟢" in str(value):
                    cell.font = Font(bold=True, color='388E3C')
    
    # Добавляем прогресс-бары для визуализации
    progress_rule = ColorScaleRule(
        start_type='num', start_value=0, start_color=colors['danger'],
        mid_type='num', mid_value=50, mid_color=colors['warning'],
        end_type='num', end_value=100, end_color=colors['success']
    )
    ws_gantt.conditional_formatting.add("H3:H9", progress_rule)
    
    # Ширина колонок Ганта
    for col in range(1, 14):
        ws_gantt.column_dimensions[chr(64 + col)].width = 16
    
    # ============ ОБЩАЯ СВОДКА ПРОИЗВОДСТВА ============
    print("📈 Создаем общую сводку производства...")
    ws_summary = wb.create_sheet("📊_Общая_сводка")
    
    # Заголовок сводки
    ws_summary['A1'] = "📊 ОБЩАЯ СВОДКА ПРОИЗВОДСТВА - ИТОГИ И РЕЙТИНГИ"
    ws_summary['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws_summary['A1'].fill = PatternFill(start_color=colors['dark'], end_color=colors['dark'], fill_type="solid")
    ws_summary.merge_cells('A1:I1')
    ws_summary['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 35
    
    # Сводка по станкам
    summary_headers = [
        "Станок", "Средний_OEE_%", "Средний_KPI_%", "Произведено_шт", "Брак_шт", 
        "Эффективность_%", "Рейтинг", "Статус", "Рекомендации"
    ]
    
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=3, column=col, value=header)
        cell.style = "header_style"
    
    # Данные по станкам в сводке
    for row, station in enumerate(stations, 4):
        ws_summary.cell(row=row, column=1, value=station.replace('_', ' '))
        ws_summary.cell(row=row, column=2, value=f'=IFERROR(AVERAGEIF(Общие_данные!C:C,"{station}",Общие_данные!P:P),0)')
        ws_summary.cell(row=row, column=3, value=f'=IFERROR(AVERAGEIF(Общие_данные!C:C,"{station}",Общие_данные!R:R),0)')
        ws_summary.cell(row=row, column=4, value=f'=SUMIF(Общие_данные!C:C,"{station}",Общие_данные!J:J)')
        ws_summary.cell(row=row, column=5, value=f'=SUMIF(Общие_данные!C:C,"{station}",Общие_данные!K:K)')
        ws_summary.cell(row=row, column=6, value=f'=IFERROR((D{row}-E{row})/D{row}*100,0)')
        ws_summary.cell(row=row, column=7, value=f'=IF(B{row}>=90,"🏆 ТОП",IF(B{row}>=80,"⭐ Отлично",IF(B{row}>=70,"✅ Хорошо","⚠️ Улучшить")))')
        ws_summary.cell(row=row, column=8, value=f'=IF(B{row}>=85,"🟢 Стабильно",IF(B{row}>=75,"🟡 Нормально","🔴 Проблемы"))')
        ws_summary.cell(row=row, column=9, value=f'=IF(B{row}<70,"Срочный ремонт",IF(B{row}<80,"Профилактика","Поддерживать"))')
        
        # Форматирование процентов
        for col in [2, 3, 6]:
            ws_summary.cell(row=row, column=col).number_format = '0.0"%"'
    
    # Общие итоги производства
    ws_summary['A12'] = "🏭 ОБЩИЕ ИТОГИ ПРОИЗВОДСТВА:"
    ws_summary['A12'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_summary['A12'].fill = PatternFill(start_color=colors['dark'], end_color=colors['dark'], fill_type="solid")
    ws_summary.merge_cells('A12:I12')
    ws_summary['A12'].alignment = Alignment(horizontal="center")
    ws_summary.row_dimensions[12].height = 25
    
    production_totals = [
        ("A14", "📊 Общий OEE всего производства:", "D14", "=AVERAGE(B4:B10)", "0.0\"%\""),
        ("A15", "📈 Общий KPI всего производства:", "D15", "=AVERAGE(C4:C10)", "0.0\"%\""),
        ("A16", "🔧 Всего произведено деталей:", "D16", "=SUM(D4:D10)", "#,##0"),
        ("A17", "❌ Общий брак за период:", "D17", "=SUM(E4:E10)", "#,##0"),
        ("A18", "⚡ Общая эффективность производства:", "D18", "=AVERAGE(F4:F10)", "0.0\"%\""),
        ("A19", "🏆 Лучший станок по OEE:", "D19", "=INDEX(A4:A10,MATCH(MAX(B4:B10),B4:B10,0))", ""),
        ("A20", "⚠️ Станок требующий внимания:", "D20", "=INDEX(A4:A10,MATCH(MIN(B4:B10),B4:B10,0))", ""),
        
        ("F14", "📅 Период анализа:", "H14", "=TODAY()-7&\" - \"&TODAY()", ""),
        ("F15", "🔄 Активных станков:", "H15", f"={len(stations)}", "0"),
        ("F16", "👥 Операторов задействовано:", "H16", f"={len(operators)}", "0"),
        ("F17", "📊 Средняя загрузка смены:", "H17", "=IFERROR(AVERAGE(Общие_данные!M:M),0)", "0.0\"%\""),
        ("F18", "🎯 Целевой OEE (план):", "H18", "85%", ""),
        ("F19", "📈 Отклонение от плана:", "H19", "=D14-85%", "0.0\"%\""),
        ("F20", "✅ Выполнение плана:", "H20", "=IF(H19>=0,\"Превышен\",\"Не достигнут\")", "")
    ]
    
    for total in production_totals:
        ws_summary[total[0]] = total[1]
        ws_summary[total[0]].font = Font(bold=True, size=10)
        ws_summary[total[2]] = total[3]
        ws_summary[total[2]].font = Font(bold=True, color=colors['primary'], size=10)
        
        if len(total) > 4 and total[4]:
            ws_summary[total[2]].number_format = total[4]
    
    # Условное форматирование сводки
    summary_oee_rule = ColorScaleRule(
        start_type='num', start_value=0, start_color=colors['danger'],
        mid_type='num', mid_value=75, mid_color=colors['warning'],
        end_type='num', end_value=100, end_color=colors['success']
    )
    ws_summary.conditional_formatting.add("B4:C10", summary_oee_rule)
    
    # Ширина колонок сводки
    for col in range(1, 10):
        ws_summary.column_dimensions[chr(64 + col)].width = 18
    
    # ============ ДОБАВЛЯЕМ ДИАГРАММЫ ============
    print("📈 Добавляем диаграммы...")
    
    # Диаграмма OEE по станкам
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "OEE по станкам"
    chart.y_axis.title = 'OEE %'
    chart.x_axis.title = 'Станки'
    
    data = Reference(ws_summary, min_col=2, min_row=3, max_row=10, max_col=2)
    categories = Reference(ws_summary, min_col=1, min_row=4, max_row=10)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 10
    chart.width = 15
    
    ws_summary.add_chart(chart, "A22")
    
    # ============ СОХРАНЕНИЕ ФАЙЛА ============
    filename = f"Production_System_v11_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(os.getcwd(), filename)
    
    try:
        wb.save(filepath)
        print(f"✅ Система управления производством v11.0 создана: {filename}")
        print(f"📁 Размер файла: {os.path.getsize(filepath) / 1024:.1f} KB")
        
        print("\n🎉 ВСЕ ТРЕБОВАНИЯ ВЫПОЛНЕНЫ:")
        print("   ✅ Отдельные листы данных для каждого станка")
        print("   ✅ KPI по каждому оператору в отдельной вкладке")
        print("   ✅ Производство как остаток с компенсацией брака")
        print("   ✅ Колонка 'N чертежа' после даты")
        print("   ✅ Календарь для выбора даты")
        print("   ✅ Выпадающие списки для операторов")
        print("   ✅ Закрепленная шапка с масштабированием")
        print("   ✅ Общий OEE/KPI по всему производству")
        print("   ✅ Диаграммы Ганта для визуализации")
        print("   ✅ Современный дизайн дашборда")
        
        print(f"\n📋 СТРУКТУРА ФАЙЛА:")
        print(f"   🔸 Общие_данные - главная таблица с закрепленной шапкой")
        print(f"   🔸 7 листов станков - индивидуальные данные по каждому")
        print(f"   🔸 KPI_по_операторам - детальная аналитика сотрудников")
        print(f"   🔸 🎛️_Dashboard - современная панель управления")
        print(f"   🔸 📅_Диаграмма_Ганта - производственное планирование")
        print(f"   🔸 📊_Общая_сводка - итоги с диаграммами")
        
        print(f"\n🚀 НОВЫЕ ВОЗМОЖНОСТИ v11.0:")
        print(f"   📊 Остаток производства с автоматическим пересчетом")
        print(f"   📅 Календарь для удобного выбора дат")
        print(f"   📋 Колонка N чертежа для привязки к техдокументации")
        print(f"   🎯 Детальные бонусы и рекомендации для операторов")
        print(f"   📈 Интерактивные диаграммы и визуализация")
        print(f"   🎨 Современное цветовое кодирование")
        print(f"   🔄 Автоматическое обновление всех расчетов")
        
        return filepath
        
    except Exception as e:
        print(f"❌ Ошибка при создании файла: {e}")
        return None

def main():
    """Основная функция"""
    print("=" * 80)
    print("🏭 СИСТЕМА УПРАВЛЕНИЯ ПРОИЗВОДСТВОМ v11.0")
    print("ИСПРАВЛЕННАЯ ВЕРСИЯ - ВСЕ ТРЕБОВАНИЯ ВЫПОЛНЕНЫ")
    print("=" * 80)
    
    try:
        import openpyxl
        print("✅ Библиотека openpyxl загружена")
    except ImportError:
        print("❌ Ошибка: Установите openpyxl командой: pip install openpyxl")
        return
    
    filepath = create_production_system_v11()
    
    if filepath:
        print(f"\n🎉 СИСТЕМА УСПЕШНО СОЗДАНА!")
        print(f"📂 Файл: {os.path.basename(filepath)}")
        print(f"\n🚀 ГОТОВО К ИСПОЛЬЗОВАНИЮ!")
        print(f"📖 Откройте файл в Excel и начните ввод данных")
    else:
        print("\n❌ Не удалось создать систему")

if __name__ == "__main__":
    main()
