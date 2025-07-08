#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ОКОНЧАТЕЛЬНЫЙ ГЕНЕРАТОР ОТЧЕТА OEE/KPI v11.0 - "ПОЛНАЯ ВЕРСИЯ ДЛЯ БОССА"
ИСПРАВЛЕНЫ ВСЕ ОШИБКИ. СОДЕРЖИТ ВСЕ ЛИСТЫ, ФУНКЦИИ И ДИЗАЙН.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from datetime import datetime, date
import os

def создать_окончательный_отчет():
    """
    Создает полный и презентабельный Excel-отчет, объединяющий все функции и дизайн.
    """
    print("👔 Создание ОКОНЧАТЕЛЬНОЙ ПОЛНОЙ 'Версии для Босса' v11.0...")
    
    # --- 1. СТИЛИ И ЦВЕТА ---
    СИНИЙ_ЦВЕТ = "4F81BD"
    ШРИФТ_ЗАГОЛОВКА = Font(bold=True, color="FFFFFF", size=12)
    ЗАЛИВКА_ЗАГОЛОВКА = PatternFill(start_color=СИНИЙ_ЦВЕТ, end_color=СИНИЙ_ЦВЕТ, fill_type="solid")
    ЦЕНТРОВАННОЕ_ВЫРАВНИВАНИЕ = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ТОНКАЯ_ГРАНИЦА = Border(left=Side(style='thin', color="BFBFBF"), right=Side(style='thin', color="BFBFBF"), top=Side(style='thin', color="BFBFBF"), bottom=Side(style='thin', color="BFBFBF"))
    ШРИФТ_KPI = Font(bold=True, size=24, color=СИНИЙ_ЦВЕТ)
    ШРИФТ_ОПИСАНИЯ_KPI = Font(bold=True, size=12, color="000000")

    # --- 2. СОЗДАНИЕ КНИГИ И ВСЕХ ЛИСТОВ ---
    книга = openpyxl.Workbook()
    книга.remove(книга.active)

    лист_панель = книга.create_sheet("Панель Управления", 0)
    лист_основные_данные = книга.create_sheet("Основные Данные")
    лист_kpi = книга.create_sheet("KPI Операторов")
    лист_гант = книга.create_sheet("График Ганта")
    лист_общие_результаты = книга.create_sheet("Общие Результаты")
    лист_переводы = книга.create_sheet("Переводы")

    станки = ["Doosan_Yashana", "Doosan_Hadasha", "Doosan_3", "Pinnacle_Gdola", "Mitsubishi", "JohnFord", "Okuma"]
    операторы = ["Андрей", "Денис", "Даниил", "Кирилл", "Слава", "Аркадий"]

    # ============ ОСНОВНЫЕ ДАННЫЕ ============
    print("📊 Оформление листа 'Основные Данные'...")
    # Код этого блока не менялся
    лист_основные_данные.append([
        "Дата", "Станок", "Оператор", "Смена, мин", "Наладка, мин", "Остаток производства", "Простои, мин",
        "План, шт", "Факт, шт", "Брак, шт", "Годные, шт", "Доступность, %", "Произв-ть, %",
        "Качество, %", "OEE, %", "KPI, %"
    ])
    for ячейка in лист_основные_данные["1:1"]: ячейка.font = ШРИФТ_ЗАГОЛОВКА; ячейка.fill = ЗАЛИВКА_ЗАГОЛОВКА; ячейка.alignment = ЦЕНТРОВАННОЕ_ВЫРАВНИВАНИЕ
    for нс in range(2, 102):
        for нк in range(1, 17): лист_основные_данные.cell(row=нс, column=нк).border = ТОНКАЯ_ГРАНИЦА
        лист_основные_данные.cell(row=нс, column=11, value=f'=I{нс}-J{нс}')
        лист_основные_данные.cell(row=нс, column=12, value=f'=IFERROR((D{нс}-G{нс})/D{нс}, 0)').number_format = '0.0%'
        лист_основные_данные.cell(row=нс, column=13, value=f'=IFERROR(I{нс}/H{нс}, 0)').number_format = '0.0%'
        лист_основные_данные.cell(row=нс, column=14, value=f'=IFERROR(K{нс}/I{нс}, 0)').number_format = '0.0%'
        лист_основные_данные.cell(row=нс, column=15, value=f'=IFERROR(L{нс}*M{нс}*N{нс}, 0)').number_format = '0.0%'
        лист_основные_данные.cell(row=нс, column=16, value=f'=IFERROR((O{нс}*0.7)+(N{нс}*0.3), 0)').number_format = '0.0%'
    лист_основные_данные.conditional_formatting.add(f"J2:J101", ColorScaleRule(start_type='min', start_color='63BE7B', end_type='max', end_color='F8696B'))
    лист_основные_данные.conditional_formatting.add(f"O2:O101", DataBarRule(start_type="min", end_type="max", color=СИНИЙ_ЦВЕТ))
    лист_основные_данные.conditional_formatting.add(f"P2:P101", DataBarRule(start_type="min", end_type="max", color="63BE7B"))
    проверка_операторы = DataValidation(type="list", formula1=f'"{",".join(операторы)}"'); проверка_операторы.add("C2:C101"); лист_основные_данные.add_data_validation(проверка_операторы)
    проверка_станки = DataValidation(type="list", formula1=f'"{",".join(станки)}"'); проверка_станки.add("B2:B101"); лист_основные_данные.add_data_validation(проверка_станки)
    лист_основные_данные.freeze_panes = "A2";
    for б in 'ABCDEFGHIJKLMNOP': лист_основные_данные.column_dimensions[б].width = 15
    тестовые_данные = [ [date.today(), "Doosan_Yashana", "Андрей", 480, 120, None, 60, 150, 120, 5], [date.today(), "Doosan_Hadasha", "Денис", 480, 90, None, 30, 200, 180, 2], [date.today(), "Mitsubishi", "Даниил", 480, 150, None, 45, 250, 225, 22], [date.today(), "Pinnacle_Gdola", "Кирилл", 480, 60, None, 15, 300, 290, 10] ]
    for i, д in enumerate(тестовые_данные, 2):
        for k, з in enumerate(д):
            if k != 5: лист_основные_данные.cell(row=i, column=k + 1, value=з)

    # ============ KPI ОПЕРАТОРОВ ============
    print("👥 Оформление листа 'KPI Операторов'...")
    # Код этого блока не менялся
    лист_kpi.append(["Оператор", "Всего смен", "Средний OEE", "Средний KPI", "Всего годных, шт", "Всего брака, шт", "Доля брака, %", "Среднее время наладки, мин"])
    for ячейка in лист_kpi["1:1"]: ячейка.font = ШРИФТ_ЗАГОЛОВКА; ячейка.fill = ЗАЛИВКА_ЗАГОЛОВКА; ячейка.alignment = ЦЕНТРОВАННОЕ_ВЫРАВНИВАНИЕ
    for нс, оп in enumerate(операторы, 2):
        for нк in range(1, 9): лист_kpi.cell(row=нс, column=нк).border = ТОНКАЯ_ГРАНИЦА
        лист_kpi.cell(row=нс, column=1, value=оп)
        лист_kpi.cell(row=нс, column=2, value=f'=COUNTIF(\'Основные Данные\'!C:C,A{нс})')
        лист_kpi.cell(row=нс, column=3, value=f'=IFERROR(AVERAGEIF(\'Основные Данные\'!C:C,A{нс},\'Основные Данные\'!O:O),0)').number_format = '0.0%'
        лист_kpi.cell(row=нс, column=4, value=f'=IFERROR(AVERAGEIF(\'Основные Данные\'!C:C,A{нс},\'Основные Данные\'!P:P),0)').number_format = '0.0%'
        лист_kpi.cell(row=нс, column=5, value=f'=SUMIF(\'Основные Данные\'!C:C,A{нс},\'Основные Данные\'!K:K)')
        лист_kpi.cell(row=нс, column=6, value=f'=SUMIF(\'Основные Данные\'!C:C,A{нс},\'Основные Данные\'!J:J)')
        лист_kpi.cell(row=нс, column=7, value=f'=IFERROR(F{нс}/(E{нс}+F{нс}),0)').number_format = '0.0%'
        лист_kpi.cell(row=нс, column=8, value=f'=IFERROR(AVERAGEIF(\'Основные Данные\'!C:C,A{нс},\'Основные Данные\'!E:E),0)').number_format = '0'
    for к in 'ABCDEFGH': лист_kpi.column_dimensions[к].width = 25

    # ============ ЛИСТЫ ПО СТАНКАМ ============
    print("🏭 Создание листов по каждому станку...")
    # Код этого блока не менялся
    for станок in станки:
        лист_станка = книга.create_sheet(f"Станок_{станок[:8]}")
        лист_станка.append([f"Сводка по станку: {станок}"]); лист_станка['A1'].font = ШРИФТ_ЗАГОЛОВКА; лист_станка['A1'].fill = ЗАЛИВКА_ЗАГОЛОВКА; лист_станка.merge_cells("A1:B1")
        данные_станка = [("Средний OEE:", f'=IFERROR(AVERAGEIF(\'Основные Данные\'!B:B,"{станок}",\'Основные Данные\'!O:O),0)'), ("Средний KPI:", f'=IFERROR(AVERAGEIF(\'Основные Данные\'!B:B,"{станок}",\'Основные Данные\'!P:P),0)'), ("Всего произведено (факт):", f'=SUMIF(\'Основные Данные\'!B:B,"{станок}",\'Основные Данные\'!I:I)'), ("Всего брака:", f'=SUMIF(\'Основные Данные\'!B:B,"{станок}",\'Основные Данные\'!J:J)')]
        for i, (о, ф) in enumerate(данные_станка, 3):
            лист_станка.cell(row=i, column=1, value=о).font = Font(bold=True); лист_станка.cell(row=i, column=1).border = ТОНКАЯ_ГРАНИЦА
            лист_станка.cell(row=i, column=2, value=ф).border = ТОНКАЯ_ГРАНИЦА
        лист_станка.column_dimensions['A'].width = 30; лист_станка.column_dimensions['B'].width = 15

    # ============ ГРАФИК ГАНТА ============
    print("📅 Создание листа 'График Ганта'...")
    # Код этого блока не менялся
    лист_гант.append(["График Производства"]); лист_гант['A1'].font = ШРИФТ_ЗАГОЛОВКА; лист_гант['A1'].fill = ЗАЛИВКА_ЗАГОЛОВКА; лист_гант.merge_cells("A1:G1")
    заголовки_графика = ["Станок", "Оператор", "Начало", "Конец", "Длительность", "Прогресс", "Статус"]; лист_гант.append(заголовки_графика)
    for ячейка in лист_гант["2:2"]: ячейка.font = ШРИФТ_ЗАГОЛОВКА; ячейка.fill = ЗАЛИВКА_ЗАГОЛОВКА; ячейка.alignment = ЦЕНТРОВАННОЕ_ВЫРАВНИВАНИЕ
    данные_графика = [ ["Doosan_Yashana", "Андрей", "08:00", "16:00", "8h", "75%", "Активно"], ["Doosan_Hadasha", "Денис", "08:00", "14:00", "6h", "100%", "Завершено"], ["Pinnacle_Gdola", "Кирилл", "08:00", "15:30", "7.5h", "80%", "Активно"] ]
    for i, с in enumerate(данные_графика, 3):
        лист_гант.append(с)
        for j in range(1,8): лист_гант.cell(row=i, column=j).border = ТОНКАЯ_ГРАНИЦА
    for c in "ABCDEFG": лист_гант.column_dimensions[c].width = 15
    
    # ============ ОБЩИЕ РЕЗУЛЬТАТЫ (ВОССТАНОВЛЕН И ОФОРМЛЕН) ============
    print("🎯 Создание листа 'Общие Результаты'...")
    лист_общие_результаты.append(["Общие Результаты Производства"]); лист_общие_результаты['A1'].font = ШРИФТ_ЗАГОЛОВКА; лист_общие_результаты['A1'].fill = ЗАЛИВКА_ЗАГОЛОВКА; лист_общие_результаты.merge_cells("A1:B1")
    данные_результатов = [
        ("Общий OEE по предприятию:", "=IFERROR(AVERAGE('Основные Данные'!O:O),0)"),
        ("Общий KPI по предприятию:", "=IFERROR(AVERAGE('Основные Данные'!P:P),0)"),
        ("Общее качество по предприятию:", "=IFERROR(SUM('Основные Данные'!K:K)/SUM('Основные Данные'!I:I),0)"),
        ("Лучший станок (по OEE):", "=IFERROR(INDEX('Основные Данные'!B:B,MATCH(MAX('Основные Данные'!O:O),'Основные Данные'!O:O,0)),\"N/A\")"),
        ("Лучший оператор (по KPI):", "=IFERROR(INDEX('Основные Данные'!C:C,MATCH(MAX('Основные Данные'!P:P),'Основные Данные'!P:P,0)),\"N/A\")")
    ]
    for i, (о, ф) in enumerate(данные_результатов, 3):
        лист_общие_результаты.cell(row=i, column=1, value=о).font = Font(bold=True); лист_общие_результаты.cell(row=i, column=1).border = ТОНКАЯ_ГРАНИЦА
        ячейка_формулы = лист_общие_результаты.cell(row=i, column=2, value=ф)
        ячейка_формулы.border = ТОНКАЯ_ГРАНИЦА
        if "AVERAGE" in ф or "SUM" in ф: ячейка_формулы.number_format = '0.0%'
    лист_общие_результаты.column_dimensions['A'].width = 35; лист_общие_результаты.column_dimensions['B'].width = 20
    
    # ============ ПЕРЕВОДЫ (ВОССТАНОВЛЕН И ОФОРМЛЕН) ============
    print("🌐 Создание листа 'Переводы'...")
    лист_переводы.append(["Английский", "Русский", "Описание"])
    for ячейка in лист_переводы["1:1"]: ячейка.font = ШРИФТ_ЗАГОЛОВКА; ячейка.fill = ЗАЛИВКА_ЗАГОЛОВКА
    переводы = [["MainData", "Основные данные"], ["Date", "Дата"], ["Station", "Станок"], ["Operator", "Оператор"]]
    for п in переводы: лист_переводы.append(п)
    лист_переводы.column_dimensions['A'].width = 20; лист_переводы.column_dimensions['B'].width = 20; лист_переводы.column_dimensions['C'].width = 30
    
    # ============ ПАНЕЛЬ УПРАВЛЕНИЯ ============
    print("🎨✨ Создание главной панели управления...")
    # Код этого блока не менялся
    лист_панель.sheet_view.showGridLines = False; ширины = {'A': 5, 'B': 25, 'C': 25, 'D': 5, 'E': 25, 'F': 25, 'G': 15, 'H': 15};
    for к, ш in ширины.items(): лист_панель.column_dimensions[к].width = ш
    лист_панель.merge_cells("B2:G2"); яз = лист_панель['B2']; яз.value = "Ключевые Показатели Эффективности"; яз.font = Font(bold=True, size=20); яз.alignment = Alignment(horizontal='center')
    лист_панель.merge_cells("B4:C5"); лист_панель['B4'].value = "ОБЩИЙ OEE"; лист_панель['B4'].font = ШРИФТ_ОПИСАНИЯ_KPI; лист_панель['B4'].alignment = Alignment(horizontal='center')
    лист_панель.merge_cells("B6:C7"); лист_панель['B6'].value = f"=AVERAGE('Основные Данные'!O:O)"; лист_панель['B6'].font = ШРИФТ_KPI; лист_панель['B6'].number_format = '0.0%'; лист_панель['B6'].alignment = Alignment(horizontal='center', vertical='center')
    for r in range(4,8):
        for c in range(2,4): лист_панель.cell(row=r, column=c).border = ТОНКАЯ_ГРАНИЦА
    лист_панель.merge_cells("E4:F5"); лист_панель['E4'].value = "ОБЩЕЕ КАЧЕСТВО"; лист_панель['E4'].font = ШРИФТ_ОПИСАНИЯ_KPI; лист_панель['E4'].alignment = Alignment(horizontal='center')
    лист_панель.merge_cells("E6:F7"); лист_панель['E6'].value = f"=IFERROR(SUM('Основные Данные'!K:K)/SUM('Основные Данные'!I:I),0)"; лист_панель['E6'].font = ШРИФТ_KPI; лист_панель['E6'].number_format = '0.0%'; лист_панель['E6'].alignment = Alignment(horizontal='center', vertical='center')
    for r in range(4,8):
        for c in range(5,7): лист_панель.cell(row=r, column=c).border = ТОНКАЯ_ГРАНИЦА

    print("📈 Добавление графиков на панель...")
    до = BarChart(); до.title = "Рейтинг Операторов по KPI"; до.style = 12
    до_данные = Reference(лист_kpi, min_col=4, min_row=1, max_row=len(операторы)+1); до_кат = Reference(лист_kpi, min_col=1, min_row=2, max_row=len(операторы)+1)
    до.add_data(до_данные, titles_from_data=True); до.set_categories(до_кат); до.legend = None; лист_панель.add_chart(до, "B9")

    дк = PieChart(); дк.title = "Соотношение Качества"
    лист_панель['G4'] = "Годные"; лист_панель['H4'] = f"=SUM('Основные Данные'!K:K)"; лист_панель['G5'] = "Брак"; лист_панель['H5'] = f"=SUM('Основные Данные'!J:J)"
    мд = Reference(лист_панель, min_col=7, min_row=4, max_row=5); дд = Reference(лист_панель, min_col=8, min_row=4, max_row=5)
    дк.add_data(дд, titles_from_data=False); дк.set_categories(мд); лист_панель.add_chart(дк, "E9")
    лист_панель.column_dimensions['G'].hidden = True; лист_панель.column_dimensions['H'].hidden = True
    
    лист_панель.merge_cells("F23:G24"); ял = лист_панель['F23']; ял.value = "[ Место для логотипа ]"; ял.font = Font(italic=True, color="808080"); ял.alignment = Alignment(horizontal='center', vertical='center')
    пунктирная_сторона = Side(style='dashed', color="BFBFBF")
    ял.border = Border(top=пунктирная_сторона, bottom=пунктирная_сторона, left=пунктирная_сторона, right=пунктирная_сторона)
    
    # --- 5. СОХРАНЕНИЕ ФАЙЛА ---
    имя_файла = f"Полный_Отчет_OEE_KPI_{datetime.now().strftime('%Y%m%d')}.xlsx"
    путь_к_файлу = os.path.join(os.getcwd(), имя_файла)
    try:
        книга.save(путь_к_файлу)
        print("-" * 60)
        print(f"✅ УСПЕХ! Полный и презентабельный отчет сохранен: {имя_файла}")
        print("   Он содержит ВСЕ листы, функции, дизайн и графики. Все ошибки исправлены.")
        return путь_к_файлу
    except Exception as ошибка:
        print(f"❌ Ошибка при сохранении: {ошибка}")
        return None

if __name__ == "__main__":
    создать_окончательный_отчет()