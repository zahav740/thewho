# -*- coding: utf-8 -*-
"""
СИСТЕМА МОНИТОРИНГА OEE/KPI v13.6 - Полная реализация по ТЗ
Создает профессиональный Excel-файл для мониторинга производства.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
import os

def create_monitoring_system():
    """
    Основная функция для создания Excel-файла мониторинга OEE/KPI.
    """
    print("🚀 Запуск создания ПРОФЕССИОНАЛЬНОЙ системы мониторинга v13.6...")

    # --- 1. Определение стилей и констант (Разделы 2, 5, 6 ТЗ) ---
    print("🎨 Настройка стилей и цветовой схемы...")
    HEADER_COLOR = "1F4E79"
    EXCELLENT_COLOR = "00B050"
    GOOD_COLOR = "92D050"
    OK_COLOR = "FFFF00"
    WARNING_COLOR = "FFC000"
    CRITICAL_COLOR = "FF0000"
    TABLE_BG_COLOR = "E7E6E6"

    HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
    TITLE_FONT = Font(bold=True, size=16)
    HEADER_FILL = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    EXCELLENT_FILL = PatternFill(start_color=EXCELLENT_COLOR, end_color=EXCELLENT_COLOR, fill_type="solid")
    GOOD_FILL = PatternFill(start_color=GOOD_COLOR, end_color=GOOD_COLOR, fill_type="solid")
    OK_FILL = PatternFill(start_color=OK_COLOR, end_color=OK_COLOR, fill_type="solid")
    WARNING_FILL = PatternFill(start_color=WARNING_COLOR, end_color=WARNING_COLOR, fill_type="solid")
    CRITICAL_FILL = PatternFill(start_color=CRITICAL_COLOR, end_color=CRITICAL_COLOR, fill_type="solid")
    TABLE_BG_FILL = PatternFill(start_color=TABLE_BG_COLOR, end_color=TABLE_BG_COLOR, fill_type="solid")
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    MACHINES = ["Doosan Yashana", "Doosan Hadasha", "Doosan 3", "Pinnacle Gdola", "Mitsubishi", "JohnFord", "Okuma"]
    OPERATORS = ["Andrey", "Denis", "Daniel", "Kirill", "Slava", "Arkady"]
    SHIFT_DURATION = 480
    INITIAL_REMAINING = {machine: 300 for machine in MACHINES}

    # --- 2. Создание книги и листов (Раздел 4.1 ТЗ) ---
    print("🗂️ Создание структуры листов Excel...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_dashboard = wb.create_sheet("Dashboard", 0)
    ws_main_data = wb.create_sheet("MainData")
    ws_operator_kpi = wb.create_sheet("OperatorKPI")
    ws_summary = wb.create_sheet("Station_Summary")
    machine_sheets = {machine: wb.create_sheet(f"Station_{machine.split(' ')[0]}") for machine in MACHINES}
    ws_analytics = wb.create_sheet("Analytics")
    ws_gantt = wb.create_sheet("GanttChart")
    ws_results = wb.create_sheet("OverallResults")
    ws_translations = wb.create_sheet("Translations")

    # Словарь для ширин столбцов
    column_widths = {
        'Translations': {'A': 25, 'B': 25, 'C': 25},
        'MainData': {openpyxl.utils.get_column_letter(i): 18 for i in range(1, 21)},
        'OperatorKPI': {openpyxl.utils.get_column_letter(i): 18 for i in range(1, 16)},
        'Station_Summary': {'A': 20, 'B': 20, 'C': 20, 'D': 20},
        'Analytics': {'A': 25, 'B': 25},
        'GanttChart': {'A': 20, 'B': 20, 'C': 20, 'D': 20},
        'OverallResults': {'A': 20, 'B': 20, 'C': 20, 'D': 20, 'E': 20},
    }

    # --- 3. Заполнение листов ---

    # ============ Лист "Translations" (Раздел 4.1.1 ТЗ) ============
    print("📖 Заполнение 'Translations'...")
    ws_translations.append(["English Term", "Russian Translation", "Description"])
    translations_data = [
        ["OEE", "Общая эффективность оборудования", "Ключевой показатель эффективности станка"],
        ["KPI", "Ключевой показатель эффективности", "Оценка производительности оператора"],
        ["Availability", "Доступность", "Доля времени работы оборудования"],
        ["Performance", "Производительность", "Скорость выполнения плана"],
        ["Quality", "Качество", "Доля годных деталей"],
        ["Setup Ratio", "Доля наладки", "Время подготовки оборудования"],
        ["Setup Quality", "Качество наладки", "Качество подготовки оборудования"],
    ]
    for row in translations_data:
        ws_translations.append(row)
    # Форматирование заголовков
    for col_idx in range(1, 4):
        ws_translations.cell(row=1, column=col_idx).font = HEADER_FONT
        ws_translations.cell(row=1, column=col_idx).fill = HEADER_FILL
        ws_translations.cell(row=1, column=col_idx).alignment = CENTER_ALIGN
    # Форматирование тела таблицы
    for row_idx in range(2, 9):
        for col_idx in range(1, 4):
            ws_translations.cell(row=row_idx, column=col_idx).border = THIN_BORDER
            ws_translations.cell(row=row_idx, column=col_idx).fill = TABLE_BG_FILL
    for col, width in column_widths['Translations'].items():
        ws_translations.column_dimensions[col].width = width

    # ============ Лист "MainData" (Раздел 4.1.2 ТЗ) ============
    print("📊 Заполнение 'MainData'...")
    headers_main = [
        "Date", "Station", "Operator", "Shift_min", "Setup_min", "Production_remaining",
        "Downtime_min", "Planned_pcs", "Actual_pcs", "Defect_pcs", "Good_pcs",
        "Availability_pct", "Performance_pct", "Quality_pct", "OEE_pct", "Setup_ratio_pct",
        "Setup_quality_pct", "KPI_pct", "Efficiency_rating", "Status"
    ]
    ws_main_data.append(headers_main)
    for col_idx in range(1, len(headers_main) + 1):
        ws_main_data.cell(row=1, column=col_idx).font = HEADER_FONT
        ws_main_data.cell(row=1, column=col_idx).fill = HEADER_FILL
        ws_main_data.cell(row=1, column=col_idx).alignment = CENTER_ALIGN

    # Валидация данных
    dv_operators = DataValidation(type="list", formula1=f'"{",".join(OPERATORS)}"')
    dv_operators.add("C2:C1001")
    ws_main_data.add_data_validation(dv_operators)
    dv_machines = DataValidation(type="list", formula1=f'"{",".join(MACHINES)}"')
    dv_machines.add("B2:B1001")
    ws_main_data.add_data_validation(dv_machines)
    dv_setup = DataValidation(type="whole", operator="between", formula1="0", formula2="240")
    dv_setup.add("E2:E1001")
    ws_main_data.add_data_validation(dv_setup)
    dv_downtime = DataValidation(type="whole", operator="between", formula1="0", formula2="480")
    dv_downtime.add("G2:G1001")
    ws_main_data.add_data_validation(dv_downtime)
    dv_planned = DataValidation(type="whole", operator="between", formula1="1", formula2="1000")
    dv_planned.add("H2:H1001")
    ws_main_data.add_data_validation(dv_planned)
    dv_actual = DataValidation(type="whole", operator="between", formula1="0", formula2="1000")
    dv_actual.add("I2:I1001")
    ws_main_data.add_data_validation(dv_actual)
    dv_defect = DataValidation(type="custom", formula1="AND(J2>=0,J2<=I2)")
    dv_defect.add("J2:J1001")
    ws_main_data.add_data_validation(dv_defect)

    # Заполнение формул и данных
    for row in range(2, 1002):
        for col in range(1, len(headers_main) + 1):
            ws_main_data.cell(row=row, column=col).border = THIN_BORDER
            ws_main_data.cell(row=row, column=col).fill = TABLE_BG_FILL

        ws_main_data.cell(row=row, column=4).value = SHIFT_DURATION
        ws_main_data.cell(row=row, column=11).value = f'=IFERROR(I{row}-J{row},0)'
        ws_main_data.cell(row=row, column=12).value = f'=IFERROR((D{row}-G{row})/D{row},0)'
        ws_main_data.cell(row=row, column=12).number_format = '0.0%'
        ws_main_data.cell(row=row, column=13).value = f'=IFERROR(I{row}/H{row},0)'
        ws_main_data.cell(row=row, column=13).number_format = '0.0%'
        ws_main_data.cell(row=row, column=14).value = f'=IFERROR(K{row}/I{row},0)'
        ws_main_data.cell(row=row, column=14).number_format = '0.0%'
        ws_main_data.cell(row=row, column=15).value = f'=IFERROR(L{row}*M{row}*N{row},0)'
        ws_main_data.cell(row=row, column=15).number_format = '0.0%'
        ws_main_data.cell(row=row, column=16).value = f'=IFERROR(E{row}/D{row},0)'
        ws_main_data.cell(row=row, column=16).number_format = '0.0%'
        ws_main_data.cell(row=row, column=17).value = f'=N{row}'
        ws_main_data.cell(row=row, column=17).number_format = '0.0%'
        ws_main_data.cell(row=row, column=18).value = f'=IFERROR(O{row}*0.4+(1-P{row})*0.25+Q{row}*0.2+L{row}*0.15,0)'
        ws_main_data.cell(row=row, column=18).number_format = '0.0%'
        ws_main_data.cell(row=row, column=19).value = f'=IF(R{row}>=0.9,"Превосходно",IF(R{row}>=0.85,"Отлично",IF(R{row}>=0.75,"Хорошо",IF(R{row}>=0.65,"Удовлетворительно","Требует развития"))))'
        ws_main_data.cell(row=row, column=20).value = f'=IF(O{row}<0.75,"Требует улучшения","Хороший результат")'

        # Динамические остатки
        if row == 2:
            ws_main_data.cell(row=row, column=6).value = f'=IFERROR(VLOOKUP(B{row},Station_Summary!A:B,2,FALSE),0)'
        else:
            ws_main_data.cell(row=row, column=6).value = f'=IFERROR(VLOOKUP(B{row},Station_Summary!A:B,2,FALSE)-I{row}+J{row},0)'

    # Условное форматирование
    ws_main_data.conditional_formatting.add('O2:O1001', CellIsRule(operator='greaterThanOrEqual', formula=['0.85'], fill=GOOD_FILL))
    ws_main_data.conditional_formatting.add('O2:O1001', CellIsRule(operator='between', formula=['0.75', '0.8499'], fill=OK_FILL))
    ws_main_data.conditional_formatting.add('O2:O1001', CellIsRule(operator='lessThan', formula=['0.75'], fill=CRITICAL_FILL))
    ws_main_data.conditional_formatting.add('R2:R1001', CellIsRule(operator='greaterThanOrEqual', formula=['0.90'], fill=EXCELLENT_FILL))
    ws_main_data.conditional_formatting.add('R2:R1001', CellIsRule(operator='between', formula=['0.85', '0.8999'], fill=GOOD_FILL))
    ws_main_data.conditional_formatting.add('R2:R1001', CellIsRule(operator='between', formula=['0.75', '0.8499'], fill=OK_FILL))
    ws_main_data.conditional_formatting.add('R2:R1001', CellIsRule(operator='between', formula=['0.65', '0.7499'], fill=WARNING_FILL))
    ws_main_data.conditional_formatting.add('R2:R1001', CellIsRule(operator='lessThan', formula=['0.65'], fill=CRITICAL_FILL))
    ws_main_data.conditional_formatting.add('E2:E1001', CellIsRule(operator='greaterThan', formula=['120'], fill=WARNING_FILL))
    ws_main_data.conditional_formatting.add('E2:E1001', CellIsRule(operator='greaterThan', formula=['180'], fill=CRITICAL_FILL))
    ws_main_data.conditional_formatting.add('G2:G1001', CellIsRule(operator='greaterThan', formula=['60'], fill=WARNING_FILL))
    ws_main_data.conditional_formatting.add('G2:G1001', CellIsRule(operator='greaterThan', formula=['120'], fill=CRITICAL_FILL))
    ws_main_data.conditional_formatting.add('F2:F1001', CellIsRule(operator='greaterThan', formula=['100'], fill=GOOD_FILL))
    ws_main_data.conditional_formatting.add('F2:F1001', CellIsRule(operator='between', formula=['50', '100'], fill=OK_FILL))
    ws_main_data.conditional_formatting.add('F2:F1001', CellIsRule(operator='lessThan', formula=['50'], fill=CRITICAL_FILL))
    # Условное форматирование для просроченных дат
    ws_main_data.conditional_formatting.add('A2:A1001', CellIsRule(operator='lessThan', formula=['TODAY()'], fill=WARNING_FILL))

    ws_main_data.freeze_panes = "A2"
    for col, width in column_widths['MainData'].items():
        ws_main_data.column_dimensions[col].width = width

    # ============ Лист "OperatorKPI" (Раздел 4.1.3 ТЗ) ============
    print("👥 Заполнение 'OperatorKPI'...")
    headers_kpi = [
        "Operator", "Total_Shifts", "Avg_OEE", "Avg_KPI", "Best_OEE", "Worst_OEE",
        "Total_Good", "Total_Defects", "Defect_Rate", "Avg_Setup_Time",
        "Efficiency_Score", "Primary_Station", "Versatility", "Performance_Trend", "Overall_Rating"
    ]
    ws_operator_kpi.append(headers_kpi)
    for col_idx in range(1, len(headers_kpi) + 1):
        ws_operator_kpi.cell(row=1, column=col_idx).font = HEADER_FONT
        ws_operator_kpi.cell(row=1, column=col_idx).fill = HEADER_FILL
        ws_operator_kpi.cell(row=1, column=col_idx).alignment = CENTER_ALIGN

    for row, operator in enumerate(OPERATORS, 2):
        ws_operator_kpi.cell(row=row, column=1).value = operator
        condition = f'MainData!C:C,"{operator}"'
        ws_operator_kpi.cell(row=row, column=2).value = f'=COUNTIF({condition})'
        ws_operator_kpi.cell(row=row, column=3).value = f'=IFERROR(AVERAGEIF({condition},MainData!O:O),0)'
        ws_operator_kpi.cell(row=row, column=3).number_format = '0.0%'
        ws_operator_kpi.cell(row=row, column=4).value = f'=IFERROR(AVERAGEIF({condition},MainData!R:R),0)'
        ws_operator_kpi.cell(row=row, column=4).number_format = '0.0%'
        ws_operator_kpi.cell(row=row, column=5).value = f'=IFERROR(MAXIFS(MainData!O:O,{condition}),0)'
        ws_operator_kpi.cell(row=row, column=5).number_format = '0.0%'
        ws_operator_kpi.cell(row=row, column=6).value = f'=IFERROR(MINIFS(MainData!O:O,{condition}),0)'
        ws_operator_kpi.cell(row=row, column=6).number_format = '0.0%'
        ws_operator_kpi.cell(row=row, column=7).value = f'=SUMIF({condition},MainData!K:K)'
        ws_operator_kpi.cell(row=row, column=8).value = f'=SUMIF({condition},MainData!J:J)'
        ws_operator_kpi.cell(row=row, column=9).value = f'=IFERROR(H{row}/(G{row}+H{row}),0)'
        ws_operator_kpi.cell(row=row, column=9).number_format = '0.0%'
        ws_operator_kpi.cell(row=row, column=10).value = f'=IFERROR(AVERAGEIF({condition},MainData!E:E),0)'
        ws_operator_kpi.cell(row=row, column=10).number_format = '0.00 "мин"'
        ws_operator_kpi.cell(row=row, column=11).value = f'=IFERROR(D{row},0)'
        ws_operator_kpi.cell(row=row, column=11).number_format = '0.0%'
        ws_operator_kpi.cell(row=row, column=12).value = f'=IFERROR(INDEX(MainData!B:B,MATCH(MAXIFS(MainData!O:O,{condition}),MainData!O:O,0)),"N/A")'
        ws_operator_kpi.cell(row=row, column=13).value = f'=COUNTIFS(MainData!C:C,"{operator}",MainData!B:B,"<>")'
        ws_operator_kpi.cell(row=row, column=14).value = "N/A"  # Тренд требует временных данных
        ws_operator_kpi.cell(row=row, column=15).value = f'=IF(D{row}>=0.9,"Превосходно",IF(D{row}>=0.85,"Отлично",IF(D{row}>=0.75,"Хорошо",IF(D{row}>=0.65,"Удовлетворительно","Требует развития"))))'

    for row_idx in range(2, len(OPERATORS) + 2):
        for col_idx in range(1, 16):
            ws_operator_kpi.cell(row=row_idx, column=col_idx).border = THIN_BORDER
            ws_operator_kpi.cell(row=row_idx, column=col_idx).fill = TABLE_BG_FILL
    for col, width in column_widths['OperatorKPI'].items():
        ws_operator_kpi.column_dimensions[col].width = width

    # ============ Лист "Station_Summary" ============
    print("🏭 Заполнение 'Station_Summary'...")
    ws_summary.append(["Station", "Remaining_Production", "Avg_OEE", "Avg_KPI"])
    for row, machine in enumerate(MACHINES, 2):
        ws_summary.cell(row=row, column=1).value = machine
        ws_summary.cell(row=row, column=2).value = INITIAL_REMAINING[machine]
        ws_summary.cell(row=row, column=3).value = f'=IFERROR(AVERAGEIF(MainData!B:B,"{machine}",MainData!O:O),0)'
        ws_summary.cell(row=row, column=3).number_format = '0.0%'
        ws_summary.cell(row=row, column=4).value = f'=IFERROR(AVERAGEIF(MainData!B:B,"{machine}",MainData!R:R),0)'
        ws_summary.cell(row=row, column=4).number_format = '0.0%'
    for col_idx in range(1, 5):
        ws_summary.cell(row=1, column=col_idx).font = HEADER_FONT
        ws_summary.cell(row=1, column=col_idx).fill = HEADER_FILL
        ws_summary.cell(row=1, column=col_idx).alignment = CENTER_ALIGN
    for row_idx in range(2, len(MACHINES) + 2):
        for col_idx in range(1, 5):
            ws_summary.cell(row=row_idx, column=col_idx).border = THIN_BORDER
            ws_summary.cell(row=row_idx, column=col_idx).fill = TABLE_BG_FILL
    for col, width in column_widths['Station_Summary'].items():
        ws_summary.column_dimensions[col].width = width
    ws_summary.conditional_formatting.add('B2:B8', CellIsRule(operator='lessThan', formula=['50'], fill=CRITICAL_FILL))

    # ============ Листы по станкам (Раздел 4.1.4 ТЗ) ============
    print("🏭 Генерация отчетов по станкам...")
    for machine in MACHINES:
        ws = machine_sheets[machine]
        ws.append([f"Станок: {machine}"])
        ws.merge_cells("A1:F1")
        ws['A1'].font = TITLE_FONT
        ws['A1'].alignment = CENTER_ALIGN
        ws.append(["Метрика", "Значение"])
        ws.append(["Средний OEE", f'=IFERROR(AVERAGEIF(MainData!B:B,"{machine}",MainData!O:O),0)'])
        ws.append(["Средний KPI", f'=IFERROR(AVERAGEIF(MainData!B:B,"{machine}",MainData!R:R),0)'])
        ws.append(["Лучший OEE", f'=IFERROR(MAXIFS(MainData!O:O,MainData!B:B,"{machine}"),0)'])
        ws.append(["Худший OEE", f'=IFERROR(MINIFS(MainData!O:O,MainData!B:B,"{machine}"),0)'])
        ws.append(["Всего годных деталей", f'=SUMIF(MainData!B:B,"{machine}",MainData!K:K)'])
        ws.append(["Всего брака", f'=SUMIF(MainData!B:B,"{machine}",MainData!J:J)'])
        ws.append(["", ""])
        ws.append(["Дата", "Оператор", "OEE", "KPI", "Время наладки", "Годные детали"])
        for row in range(10, 1002):
            ws.cell(row=row, column=1).value = f'=IFERROR(INDEX(MainData!A:A,SMALL(IF(MainData!B:B="{machine}",ROW(MainData!B:B)-ROW(MainData!B$1)+1),ROW()-ROW($A$9))),"")'
            ws.cell(row=row, column=2).value = f'=IFERROR(INDEX(MainData!C:C,SMALL(IF(MainData!B:B="{machine}",ROW(MainData!B:B)-ROW(MainData!B$1)+1),ROW()-ROW($A$9))),"")'
            ws.cell(row=row, column=3).value = f'=IFERROR(INDEX(MainData!O:O,SMALL(IF(MainData!B:B="{machine}",ROW(MainData!B:B)-ROW(MainData!B$1)+1),ROW()-ROW($A$9))),"")'
            ws.cell(row=row, column=4).value = f'=IFERROR(INDEX(MainData!R:R,SMALL(IF(MainData!B:B="{machine}",ROW(MainData!B:B)-ROW(MainData!B$1)+1),ROW()-ROW($A$9))),"")'
            ws.cell(row=row, column=5).value = f'=IFERROR(INDEX(MainData!E:E,SMALL(IF(MainData!B:B="{machine}",ROW(MainData!B:B)-ROW(MainData!B$1)+1),ROW()-ROW($A$9))),"")'
            ws.cell(row=row, column=6).value = f'=IFERROR(INDEX(MainData!K:K,SMALL(IF(MainData!B:B="{machine}",ROW(MainData!B:B)-ROW(MainData!B$1)+1),ROW()-ROW($A$9))),"")'
        for row_idx in range(2, 1002):
            for col_idx in range(1, 7):
                ws.cell(row=row_idx, column=col_idx).border = THIN_BORDER
                ws.cell(row=row_idx, column=col_idx).fill = TABLE_BG_FILL
        for col in range(1, 7):
            col_letter = openpyxl.utils.get_column_letter(col)
            if not isinstance(ws.cell(row=1, column=col), openpyxl.cell.cell.MergedCell):
                ws.column_dimensions[col_letter].width = 20

    # ============ Лист "Dashboard" (Раздел 4.1.5 ТЗ) ============
    print("🎨 Создание 'Dashboard'...")
    ws_dashboard.sheet_view.showGridLines = False
    ws_dashboard.merge_cells("A1:F2")
    ws_dashboard['A1'].value = "Панель управления производством"
    ws_dashboard['A1'].font = TITLE_FONT
    ws_dashboard['A1'].alignment = CENTER_ALIGN
    ws_dashboard['A4'] = "Критических ситуаций (OEE < 70%):"
    ws_dashboard['B4'] = f'=COUNTIF(MainData!O:O,"<0.7")'
    ws_dashboard['B4'].font = Font(bold=True, color=CRITICAL_COLOR)
    ws_dashboard['A5'] = "Общий OEE:"
    ws_dashboard['B5'] = f'=IFERROR(AVERAGE(MainData!O:O),0)'
    ws_dashboard['B5'].number_format = '0.0%'
    ws_dashboard['A6'] = "Общий KPI:"
    ws_dashboard['B6'] = f'=IFERROR(AVERAGE(MainData!R:R),0)'
    ws_dashboard['B6'].number_format = '0.0%'
    ws_dashboard['A7'] = "Лучший оператор:"
    ws_dashboard['B7'] = f'=INDEX(OperatorKPI!A:A,MATCH(MAX(OperatorKPI!D:D),OperatorKPI!D:D,0))'
    ws_dashboard['A8'] = "Лучший станок:"
    ws_dashboard['B8'] = f'=INDEX(Station_Summary!A:A,MATCH(MAX(Station_Summary!C:C),Station_Summary!C:C,0))'

    # График OEE по станкам
    chart = BarChart()
    chart.title = "OEE по станкам"
    chart.x_axis.title = "Станки"
    chart.y_axis.title = "OEE (%)"
    chart.y_axis.majorGridlines = None
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    data = Reference(ws_summary, min_col=3, min_row=2, max_row=8)
    categories = Reference(ws_summary, min_col=1, min_row=2, max_row=8)
    chart.add_data(data)
    chart.set_categories(categories)
    chart.dataLabels = openpyxl.chart.label.DataLabelList()
    chart.dataLabels.showVal = True
    ws_dashboard.add_chart(chart, "A10")

    # ============ Лист "Analytics" (Раздел 4.1.6 ТЗ) ============
    print("📈 Заполнение 'Analytics'...")
    ws_analytics['A1'] = "Продвинутая аналитика"
    ws_analytics['A1'].font = TITLE_FONT
    ws_analytics.append(["Метрика", "Значение"])
    ws_analytics.append(["Стабильность производства", f'=IFERROR(STDEV(MainData!O:O)/AVERAGE(MainData!O:O),0)'])
    ws_analytics.append(["OEE в пиковые часы", f'=IFERROR(AVERAGEIFS(MainData!O:O,MainData!O:O,">0.85"),0)'])
    ws_analytics.append(["Потенциал улучшения", f'=IFERROR(1-AVERAGE(MainData!O:O),0)'])
    ws_analytics.append(["Среднее время простоя", f'=IFERROR(AVERAGE(MainData!G:G),0)'])
    ws_analytics.append(["Процент брака", f'=IFERROR(SUM(MainData!J:J)/SUM(MainData!I:I),0)'])
    for row_idx in range(1, 7):
        for col_idx in range(1, 3):
            ws_analytics.cell(row=row_idx, column=col_idx).border = THIN_BORDER
            ws_analytics.cell(row=row_idx, column=col_idx).fill = TABLE_BG_FILL
    for col, width in column_widths['Analytics'].items():
        ws_analytics.column_dimensions[col].width = width

    # ============ Лист "GanttChart" (Раздел 4.1.7 ТЗ) ============
    print("📅 Создание 'GanttChart'...")
    ws_gantt['A1'] = "Планирование производства"
    ws_gantt['A1'].font = TITLE_FONT
    ws_gantt.append(["Станок", "Дата начала", "Дата окончания", "Задача"])
    for row, machine in enumerate(MACHINES, 2):
        ws_gantt.cell(row=row, column=1).value = machine
        ws_gantt.cell(row=row, column=2).value = datetime.now().strftime('%d.%m.%Y')
        ws_gantt.cell(row=row, column=3).value = f'=B{row}+7'
        ws_gantt.cell(row=row, column=4).value = "Производство"
        for col in range(1, 5):
            ws_gantt.cell(row=row, column=col).border = THIN_BORDER
            ws_gantt.cell(row=row, column=col).fill = TABLE_BG_FILL
    for col, width in column_widths['GanttChart'].items():
        ws_gantt.column_dimensions[col].width = width
    ws_gantt.conditional_formatting.add('D2:D8', CellIsRule(operator='equal', formula=['"Производство"'], fill=GOOD_FILL))

    # ============ Лист "OverallResults" (Раздел 4.1.8 ТЗ) ============
    print("📊 Заполнение 'OverallResults'...")
    ws_results['A1'] = "Общие результаты производства"
    ws_results['A1'].font = TITLE_FONT
    ws_results.append(["Станок", "Всего годных деталей", "Всего брака", "OEE", "KPI"])
    for row, machine in enumerate(MACHINES, 2):
        ws_results.cell(row=row, column=1).value = machine
        ws_results.cell(row=row, column=2).value = f'=SUMIF(MainData!B:B,"{machine}",MainData!K:K)'
        ws_results.cell(row=row, column=3).value = f'=SUMIF(MainData!B:B,"{machine}",MainData!J:J)'
        ws_results.cell(row=row, column=4).value = f'=IFERROR(AVERAGEIF(MainData!B:B,"{machine}",MainData!O:O),0)'
        ws_results.cell(row=row, column=4).number_format = '0.0%'
        ws_results.cell(row=row, column=5).value = f'=IFERROR(AVERAGEIF(MainData!B:B,"{machine}",MainData!R:R),0)'
        ws_results.cell(row=row, column=5).number_format = '0.0%'
        for col in range(1, 6):
            ws_results.cell(row=row, column=col).border = THIN_BORDER
            ws_results.cell(row=row, column=col).fill = TABLE_BG_FILL
    for col, width in column_widths['OverallResults'].items():
        ws_results.column_dimensions[col].width = width

    # --- 4. Сохранение файла ---
    filename = f"OEE_Monitoring_System_v13.6_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(os.getcwd(), filename)
    try:
        wb.save(filepath)
        print(f"✅ УСПЕХ! Файл создан: {filename}")
        return filepath
    except Exception as e:
        print(f"❌ Ошибка при сохранении: {e}")
        return None

if __name__ == "__main__":
    create_monitoring_system()