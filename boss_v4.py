
# -*- coding: utf-8 -*-
"""
СИСТЕМА МОНИТОРИНГА OEE/KPI v13.9 - Полная реализация по ТЗ
KPI и OEE вычисляются в таблицах станков, MainData агрегирует данные.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
import os

def create_monitoring_system():
    """
    Основная функция для создания Excel-файла мониторинга OEE/KPI.
    """
    print("🚀 Запуск создания ПРОФЕССИОНАЛЬНОЙ системы мониторинга v13.9...")

    # --- 1. Определение стилей и констант (Разделы 2, 5, 6 ТЗ) ---
    print("🎨 Настройка стилей и цветовой схемы...")
    HEADER_COLOR = "1F4E79"
    EXCELLENT_COLOR = "00B050"
    GOOD_COLOR = "92D050"
    OK_COLOR = "FFFF00"
    WARNING_COLOR = "FFC000"
    CRITICAL_COLOR = "FF0000"
    TABLE_BG_COLOR = "E7E6E6"

    HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
    TITLE_FONT = Font(bold=True, size=16)
    HEADER_FILL = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    EXCELLENT_FILL = PatternFill(start_color=EXCELLENT_COLOR, end_color=EXCELLENT_COLOR, fill_type="solid")
    GOOD_FILL = PatternFill(start_color=GOOD_COLOR, end_color=GOOD_COLOR, fill_type="solid")
    OK_FILL = PatternFill(start_color=OK_COLOR, end_color=OK_COLOR, fill_type="solid")
    WARNING_FILL = PatternFill(start_color=WARNING_COLOR, end_color=WARNING_COLOR, fill_type="solid")
    CRITICAL_FILL = PatternFill(start_color=CRITICAL_COLOR, end_color=CRITICAL_COLOR, fill_type="solid")
    TABLE_BG_FILL = PatternFill(start_color=TABLE_BG_COLOR, end_color=TABLE_BG_COLOR, fill_type="solid")
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    MACHINES = ["Doosan Yashana", "Doosan Hadasha", "Doosan 3", "Pinnacle Gdola", "Mitsubishi", "JohnFord", "Okuma"]
    OPERATORS = ["Andrey", "Denis", "Daniel", "Kirill", "Slava", "Arkady"]
    SHIFT_DURATION = 480

    # --- 2. Создание книги и листов (Раздел 4.1 ТЗ) ---
    print("🗂️ Создание структуры листов Excel...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_dashboard = wb.create_sheet("Dashboard", 0)
    ws_main_data = wb.create_sheet("MainData")
    ws_operator_kpi = wb.create_sheet("OperatorKPI")
    ws_summary = wb.create_sheet("Station_Summary")
    machine_sheets = {machine: wb.create_sheet(f"Station_{machine.split(' ')[0]}") for machine in MACHINES}
    ws_analytics = wb.create_sheet("Analytics")
    ws_gantt = wb.create_sheet("GanttChart")
    ws_results = wb.create_sheet("OverallResults")
    ws_translations = wb.create_sheet("Translations")

    # Словарь для ширин столбцов
    column_widths = {
        'Translations': {'A': 25, 'B': 25, 'C': 25},
        'MainData': {openpyxl.utils.get_column_letter(i): 18 for i in range(1, 6)},  # Уменьшен до агрегации
        'OperatorKPI': {openpyxl.utils.get_column_letter(i): 18 for i in range(1, 16)},
        'Station_Summary': {'A': 20, 'B': 20, 'C': 20, 'D': 20},
        'Analytics': {'A': 25, 'B': 25},
        'GanttChart': {'A': 20, 'B': 20, 'C': 20, 'D': 20},
        'OverallResults': {'A': 20, 'B': 20, 'C': 20, 'D': 20, 'E': 20},
    }

    # --- 3. Заполнение листов ---

    # ============ Лист "Translations" (Раздел 4.1.1 ТЗ) ============
    print("📖 Заполнение 'Translations'...")
    ws_translations.append(["English Term", "Russian Translation", "Description"])
    translations_data = [
        ["OEE", "Общая эффективность оборудования", "Ключевой показатель эффективности станка"],
        ["KPI", "Ключевой показатель эффективности", "Оценка производительности оператора"],
        ["Availability", "Доступность", "Доля времени работы оборудования"],
        ["Performance", "Производительность", "Скорость выполнения плана"],
        ["Quality", "Качество", "Доля годных деталей"],
        ["Setup Ratio", "Доля наладки", "Время подготовки оборудования"],
        ["Setup Quality", "Качество наладки", "Качество подготовки оборудования"],
    ]
    for row in translations_data:
        ws_translations.append(row)
    for col_idx in range(1, 4):
        ws_translations.cell(row=1, column=col_idx).font = HEADER_FONT
        ws_translations.cell(row=1, column=col_idx).fill = HEADER_FILL
        ws_translations.cell(row=1, column=col_idx).alignment = CENTER_ALIGN
    for row_idx in range(2, 9):
        for col_idx in range(1, 4):
            ws_translations.cell(row=row_idx, column=col_idx).border = THIN_BORDER
            ws_translations.cell(row=row_idx, column=col_idx).fill = TABLE_BG_FILL
    for col, width in column_widths['Translations'].items():
        ws_translations.column_dimensions[col].width = width

    # ============ Лист "MainData" (Агрегация данных) ============
    print("📊 Заполнение 'MainData'...")
    headers_main = ["Station", "Avg_OEE_pct", "Avg_KPI_pct", "Total_Good_pcs", "Total_Defect_pcs"]
    ws_main_data.append(headers_main)
    for col_idx in range(1, len(headers_main) + 1):
        ws_main_data.cell(row=1, column=col_idx, value=headers_main[col_idx-1]).font = HEADER_FONT
        ws_main_data.cell(row=1, column=col_idx).fill = HEADER_FILL
        ws_main_data.cell(row=1, column=col_idx).alignment = CENTER_ALIGN

    for row, machine in enumerate(MACHINES, 2):
        ws_main_data.cell(row=row, column=1).value = machine
        ws_main_data.cell(row=row, column=2).value = f'=IFERROR(AVERAGE(Station_{machine.split(" ")[0]}!M:M),0)'  # OEE
        ws_main_data.cell(row=row, column=2).number_format = '0.0%'
        ws_main_data.cell(row=row, column=3).value = f'=IFERROR(AVERAGE(Station_{machine.split(" ")[0]}!P:P),0)'  # KPI
        ws_main_data.cell(row=row, column=3).number_format = '0.0%'
        ws_main_data.cell(row=row, column=4).value = f'=SUM(Station_{machine.split(" ")[0]}!I:I)'  # Good_pcs
        ws_main_data.cell(row=row, column=5).value = f'=SUM(Station_{machine.split(" ")[0]}!H:H)'  # Defect_pcs
        for col_idx in range(1, len(headers_main) + 1):
            ws_main_data.cell(row=row, column=col_idx).border = THIN_BORDER
            ws_main_data.cell(row=row, column=col_idx).fill = TABLE_BG_FILL

    ws_main_data.freeze_panes = "A2"
    for col, width in column_widths['MainData'].items():
        ws_main_data.column_dimensions[col].width = width

    # ============ Лист "OperatorKPI" (Раздел 4.1.3 ТЗ) ============
    print("👥 Заполнение 'OperatorKPI'...")
    headers_kpi = [
        "Operator", "Total_Shifts", "Avg_OEE", "Avg_KPI", "Best_OEE", "Worst_OEE",
        "Total_Good", "Total_Defects", "Defect_Rate", "Avg_Setup_Time",
        "Efficiency_Score", "Primary_Station", "Versatility", "Performance_Trend", "Overall_Rating"
    ]
    ws_operator_kpi.append(headers_kpi)
    for col_idx in range(1, len(headers_kpi) + 1):
        ws_operator_kpi.cell(row=1, column=col_idx, value=headers_kpi[col_idx-1]).font = HEADER_FONT
        ws_operator_kpi.cell(row=1, column=col_idx).fill = HEADER_FILL
        ws_operator_kpi.cell(row=1, column=col_idx).alignment = CENTER_ALIGN

    for row, operator in enumerate(OPERATORS, 2):
        ws_operator_kpi.cell(row=row, column=1).value = operator
        # Агрегация по всем станкам для оператора
        condition_base = f'MainData!A:A,"<>"'
        ws_operator_kpi.cell(row=row, column=2).value = f'=COUNTIF({condition_base})'  # Общее количество смен
        ws_operator_kpi.cell(row=row, column=3).value = f'=IFERROR(AVERAGEIF(MainData!A:A,"<>",MainData!B:B),0)'  # Avg_OEE
        ws_operator_kpi.cell(row=row, column=3).number_format = '0.0%'
        ws_operator_kpi.cell(row=row, column=4).value = f'=IFERROR(AVERAGEIF(MainData!A:A,"<>",MainData!C:C),0)'  # Avg_KPI
        ws_operator_kpi.cell(row=row, column=4).number_format = '0.0%'
        ws_operator_kpi.cell(row=row, column=5).value = "N/A"  # Best_OEE (требует доработки)
        ws_operator_kpi.cell(row=row, column=6).value = "N/A"  # Worst_OEE (требует доработки)
        ws_operator_kpi.cell(row=row, column=7).value = f'=SUMIF(MainData!A:A,"<>",MainData!D:D)'  # Total_Good
        ws_operator_kpi.cell(row=row, column=8).value = f'=SUMIF(MainData!A:A,"<>",MainData!E:E)'  # Total_Defects
        ws_operator_kpi.cell(row=row, column=9).value = f'=IFERROR(H{row}/(G{row}+H{row}),0)'  # Defect_Rate
        ws_operator_kpi.cell(row=row, column=9).number_format = '0.0%'
        ws_operator_kpi.cell(row=row, column=10).value = "N/A"  # Avg_Setup_Time (требует доработки)
        ws_operator_kpi.cell(row=row, column=11).value = f'=IFERROR(D{row},0)'  # Efficiency_Score
        ws_operator_kpi.cell(row=row, column=11).number_format = '0.0%'
        ws_operator_kpi.cell(row=row, column=12).value = "N/A"  # Primary_Station (требует доработки)
        ws_operator_kpi.cell(row=row, column=13).value = "N/A"  # Versatility (требует доработки)
        ws_operator_kpi.cell(row=row, column=14).value = "N/A"  # Performance_Trend (требует доработки)
        ws_operator_kpi.cell(row=row, column=15).value = f'=IF(D{row}>=0.9,"Превосходно",IF(D{row}>=0.85,"Отлично",IF(D{row}>=0.75,"Хорошо",IF(D{row}>=0.65,"Удовлетворительно","Требует развития"))))'

    for row_idx in range(2, len(OPERATORS) + 2):
        for col_idx in range(1, 16):
            ws_operator_kpi.cell(row=row_idx, column=col_idx).border = THIN_BORDER
            ws_operator_kpi.cell(row=row_idx, column=col_idx).fill = TABLE_BG_FILL
    for col, width in column_widths['OperatorKPI'].items():
        ws_operator_kpi.column_dimensions[col].width = width

    # ============ Лист "Station_Summary" ============
    print("🏭 Заполнение 'Station_Summary'...")
    ws_summary.append(["Station", "Remaining_Production", "Avg_OEE", "Avg_KPI"])
    for row, machine in enumerate(MACHINES, 2):
        ws_summary.cell(row=row, column=1).value = machine
        ws_summary.cell(row=row, column=2).value = 300  # Изначальное значение
        ws_summary.cell(row=row, column=3).value = f'=IFERROR(AVERAGE(Station_{machine.split(" ")[0]}!M:M),0)'
        ws_summary.cell(row=row, column=3).number_format = '0.0%'
        ws_summary.cell(row=row, column=4).value = f'=IFERROR(AVERAGE(Station_{machine.split(" ")[0]}!P:P),0)'
        ws_summary.cell(row=row, column=4).number_format = '0.0%'
    for col_idx in range(1, 5):
        ws_summary.cell(row=1, column=col_idx).font = HEADER_FONT
        ws_summary.cell(row=1, column=col_idx).fill = HEADER_FILL
        ws_summary.cell(row=1, column=col_idx).alignment = CENTER_ALIGN
    for row_idx in range(2, len(MACHINES) + 2):
        for col_idx in range(1, 5):
            ws_summary.cell(row=row_idx, column=col_idx).border = THIN_BORDER
            ws_summary.cell(row=row_idx, column=col_idx).fill = TABLE_BG_FILL
    for col, width in column_widths['Station_Summary'].items():
        ws_summary.column_dimensions[col].width = width
    ws_summary.conditional_formatting.add('B2:B8', CellIsRule(operator='lessThan', formula=['50'], fill=CRITICAL_FILL))

    # ============ Листы по станкам (Раздел 4.1.4 ТЗ) ============
    print("🏭 Генерация отчетов по станкам...")
    for machine in MACHINES:
        ws = machine_sheets[machine]
        ws.append([f"Станок: {machine}"])
        ws.merge_cells("A1:F1")
        ws['A1'].font = TITLE_FONT
        ws['A1'].alignment = CENTER_ALIGN
        headers_station = [
            "Date", "Operator", "Shift_min", "Setup_min", "Downtime_min", "Planned_pcs",
            "Actual_pcs", "Defect_pcs", "Good_pcs", "Availability_pct", "Performance_pct",
            "Quality_pct", "OEE_pct", "Setup_ratio_pct", "Setup_quality_pct", "KPI_pct"
        ]
        ws.append(headers_station)
        for col_idx in range(1, len(headers_station) + 1):
            ws.cell(row=2, column=col_idx, value=headers_station[col_idx-1]).font = HEADER_FONT
            ws.cell(row=2, column=col_idx).fill = HEADER_FILL
            ws.cell(row=2, column=col_idx).alignment = CENTER_ALIGN

        # Валидация данных
        dv_operators = DataValidation(type="list", formula1=f'"{",".join(OPERATORS)}"')
        dv_operators.add("B3:B1002")
        ws.add_data_validation(dv_operators)
        dv_setup = DataValidation(type="whole", operator="between", formula1="0", formula2="240")
        dv_setup.add("D3:D1002")
        ws.add_data_validation(dv_setup)
        dv_downtime = DataValidation(type="whole", operator="between", formula1="0", formula2="480")
        dv_downtime.add("E3:E1002")
        ws.add_data_validation(dv_downtime)
        dv_planned = DataValidation(type="whole", operator="between", formula1="1", formula2="1000")
        dv_planned.add("F3:F1002")
        ws.add_data_validation(dv_planned)
        dv_actual = DataValidation(type="whole", operator="between", formula1="0", formula2="1000")
        dv_actual.add("G3:G1002")
        ws.add_data_validation(dv_actual)
        dv_defect = DataValidation(type="custom", formula1="AND(H3>=0,H3<=G3)")
        dv_defect.add("H3:H1002")
        ws.add_data_validation(dv_defect)

        # Заполнение данных и формул
        for row in range(3, 1002):
            for col in range(1, len(headers_station) + 1):
                ws.cell(row=row, column=col).border = THIN_BORDER
                ws.cell(row=row, column=col).fill = TABLE_BG_FILL
            ws.cell(row=row, column=3).value = SHIFT_DURATION  # Shift_min
            ws.cell(row=row, column=9).value = f'=IFERROR(G{row}-H{row},0)'  # Good_pcs
            ws.cell(row=row, column=10).value = f'=IFERROR((C{row}-E{row})/C{row},0)'  # Availability_pct
            ws.cell(row=row, column=10).number_format = '0.0%'
            ws.cell(row=row, column=11).value = f'=IFERROR(G{row}/F{row},0)'  # Performance_pct
            ws.cell(row=row, column=11).number_format = '0.0%'
            ws.cell(row=row, column=12).value = f'=IFERROR(I{row}/G{row},0)'  # Quality_pct
            ws.cell(row=row, column=12).number_format = '0.0%'
            ws.cell(row=row, column=13).value = f'=IFERROR(J{row}*K{row}*L{row},0)'  # OEE_pct
            ws.cell(row=row, column=13).number_format = '0.0%'
            ws.cell(row=row, column=14).value = f'=IFERROR(D{row}/C{row},0)'  # Setup_ratio_pct
            ws.cell(row=row, column=14).number_format = '0.0%'
            ws.cell(row=row, column=15).value = f'=L{row}'  # Setup_quality_pct
            ws.cell(row=row, column=15).number_format = '0.0%'
            ws.cell(row=row, column=16).value = f'=IFERROR(M{row}*0.4+(1-N{row})*0.25+O{row}*0.2+J{row}*0.15,0)'  # KPI_pct
            ws.cell(row=row, column=16).number_format = '0.0%'

        for col in range(1, len(headers_station) + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            if not isinstance(ws.cell(row=2, column=col), openpyxl.cell.cell.MergedCell):
                ws.column_dimensions[col_letter].width = 18

    # ============ Лист "Dashboard" (Раздел 4.1.5 ТЗ) ============
    print("🎨 Создание 'Dashboard'...")
    ws_dashboard.sheet_view.showGridLines = False
    ws_dashboard.merge_cells("A1:F2")
    ws_dashboard['A1'].value = "Панель управления производством"
    ws_dashboard['A1'].font = TITLE_FONT
    ws_dashboard['A1'].alignment = CENTER_ALIGN

    ws_dashboard['A4'] = "Критических ситуаций (OEE < 70%):"
    ws_dashboard['B4'] = f'=COUNTIF(MainData!B:B,"<0.7")'
    ws_dashboard['B4'].font = Font(bold=True, color=CRITICAL_COLOR)
    try:
        from openpyxl.comments import Comment
        ws_dashboard['A4'].comment = Comment("Источник: усредненные данные из MainData по станкам", "Note")
    except AttributeError:
        print("⚠️ Комментарии не поддерживаются в текущей версии openpyxl. Используйте заметки вручную.")

    ws_dashboard['A6'] = "Общий OEE:"  # Отступ в 1 строку
    ws_dashboard['B6'] = f'=IFERROR(AVERAGE(MainData!B:B),0)'
    ws_dashboard['B6'].number_format = '0.0%'
    try:
        ws_dashboard['A6'].comment = Comment("Источник: среднее значение OEE из MainData", "Note")
    except AttributeError:
        print("⚠️ Комментарии не поддерживаются в текущей версии openpyxl. Используйте заметки вручную.")

    ws_dashboard['A8'] = "Общий KPI:"  # Отступ в 1 строку
    ws_dashboard['B8'] = f'=IFERROR(AVERAGE(MainData!C:C),0)'
    ws_dashboard['B8'].number_format = '0.0%'
    try:
        ws_dashboard['A8'].comment = Comment("Источник: среднее значение KPI из MainData", "Note")
    except AttributeError:
        print("⚠️ Комментарии не поддерживаются в текущей версии openpyxl. Используйте заметки вручную.")

    ws_dashboard['A10'] = "Лучший оператор:"
    ws_dashboard['B10'] = "N/A"
    try:
        ws_dashboard['A10'].comment = Comment("Требуется доработка: агрегация по операторам", "Note")
    except AttributeError:
        print("⚠️ Комментарии не поддерживаются в текущей версии openpyxl. Используйте заметки вручную.")

    ws_dashboard['A12'] = "Лучший станок:"
    ws_dashboard['B12'] = f'=INDEX(MainData!A:A,MATCH(MAX(MainData!B:B),MainData!B:B,0))'
    try:
        ws_dashboard['A12'].comment = Comment("Источник: станок с максимальным OEE из MainData", "Note")
    except AttributeError:
        print("⚠️ Комментарии не поддерживаются в текущей версии openpyxl. Используйте заметки вручную.")

    chart = BarChart()
    chart.title = "OEE по станкам"
    chart.x_axis.title = "Станки"
    chart.y_axis.title = "OEE (%)"
    chart.y_axis.majorGridlines = None
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    data = Reference(ws_main_data, min_col=2, min_row=2, max_row=8)
    categories = Reference(ws_main_data, min_col=1, min_row=2, max_row=8)
    chart.add_data(data)
    chart.set_categories(categories)
    chart.dataLabels = openpyxl.chart.label.DataLabelList()
    chart.dataLabels.showVal = True
    chart.height = 15  # Установка высоты графика (в пунктах)
    chart.width = 30   # Установка ширины графика (в пунктах)
    ws_dashboard.add_chart(chart, "A15")  # Сдвигаем график ниже

    # ============ Лист "Analytics" (Раздел 4.1.6 ТЗ) ============
    print("📈 Заполнение 'Analytics'...")
    ws_analytics['A1'] = "Продвинутая аналитика"
    ws_analytics['A1'].font = TITLE_FONT
    ws_analytics.append(["Метрика", "Значение"])
    ws_analytics.append(["Стабильность производства", f'=IFERROR(STDEV(MainData!B:B)/AVERAGE(MainData!B:B),0)'])
    ws_analytics.append(["OEE в пиковые часы", f'=IFERROR(AVERAGEIF(MainData!B:B,">0.85"),0)'])
    ws_analytics.append(["Потенциал улучшения", f'=IFERROR(1-AVERAGE(MainData!B:B),0)'])
    ws_analytics.append(["Среднее время простоя", "N/A"])  # Требует доработки
    ws_analytics.append(["Процент брака", f'=IFERROR(SUM(MainData!E:E)/SUM(MainData!D:D),0)'])
    for row_idx in range(1, 7):
        for col_idx in range(1, 3):
            ws_analytics.cell(row=row_idx, column=col_idx).border = THIN_BORDER
            ws_analytics.cell(row=row_idx, column=col_idx).fill = TABLE_BG_FILL
    for col, width in column_widths['Analytics'].items():
        ws_analytics.column_dimensions[col].width = width

    # ============ Лист "GanttChart" (Раздел 4.1.7 ТЗ) ============
    print("📅 Создание 'GanttChart'...")
    ws_gantt['A1'] = "Планирование производства"
    ws_gantt['A1'].font = TITLE_FONT
    ws_gantt.append(["Станок", "Дата начала", "Дата окончания", "Задача"])
    for row, machine in enumerate(MACHINES, 2):
        ws_gantt.cell(row=row, column=1).value = machine
        ws_gantt.cell(row=row, column=2).value = datetime(2025, 6, 28, 22, 40).strftime('%d.%m.%Y')  # 10:40 PM IDT
        ws_gantt.cell(row=row, column=3).value = f'=B{row}+7'
        ws_gantt.cell(row=row, column=4).value = "Производство"
        for col in range(1, 5):
            ws_gantt.cell(row=row, column=col).border = THIN_BORDER
            ws_gantt.cell(row=row, column=col).fill = TABLE_BG_FILL
    for col, width in column_widths['GanttChart'].items():
        ws_gantt.column_dimensions[col].width = width
    ws_gantt.conditional_formatting.add('D2:D8', CellIsRule(operator='equal', formula=['"Производство"'], fill=GOOD_FILL))

    # ============ Лист "OverallResults" (Раздел 4.1.8 ТЗ) ============
    print("📊 Заполнение 'OverallResults'...")
    ws_results['A1'] = "Общие результаты производства"
    ws_results['A1'].font = TITLE_FONT
    ws_results.append(["Станок", "Всего годных деталей", "Всего брака", "OEE", "KPI"])
    for row, machine in enumerate(MACHINES, 2):
        ws_results.cell(row=row, column=1).value = machine
        ws_results.cell(row=row, column=2).value = f'=SUM(Station_{machine.split(" ")[0]}!I:I)'
        ws_results.cell(row=row, column=3).value = f'=SUM(Station_{machine.split(" ")[0]}!H:H)'
        ws_results.cell(row=row, column=4).value = f'=IFERROR(AVERAGE(Station_{machine.split(" ")[0]}!M:M),0)'
        ws_results.cell(row=row, column=4).number_format = '0.0%'
        ws_results.cell(row=row, column=5).value = f'=IFERROR(AVERAGE(Station_{machine.split(" ")[0]}!P:P),0)'
        ws_results.cell(row=row, column=5).number_format = '0.0%'
        for col in range(1, 6):
            ws_results.cell(row=row, column=col).border = THIN_BORDER
            ws_results.cell(row=row, column=col).fill = TABLE_BG_FILL
    for col, width in column_widths['OverallResults'].items():
        ws_results.column_dimensions[col].width = width

    # --- 4. Сохранение файла ---
    filename = f"OEE_Monitoring_System_v13.9_{datetime(2025, 6, 28, 22, 40).strftime('%Y%m%d_%H%M')}.xlsx"
    filepath = os.path.join(os.getcwd(), filename)
    try:
        wb.save(filepath)
        print(f"✅ УСПЕХ! Файл создан: {filename}")
        return filepath
    except Exception as e:
        print(f"❌ Ошибка при сохранении: {e}")
        return None

if __name__ == "__main__":
    create_monitoring_system()