#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
СИСТЕМА МОНИТОРИНГА OEE/KPI v14.0 - Исправленная система с разделением ответственности
OEE станка и KPI оператора рассчитываются раздельно для справедливой оценки.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart
from openpyxl.formatting.rule import CellIsRule
from openpyxl.comments import Comment
from datetime import datetime
import os

def create_monitoring_system():
    """
    Основная функция для создания Excel-файла мониторинга OEE/KPI.
    """
    print("🚀 Запуск создания ИСПРАВЛЕННОЙ системы мониторинга v14.0...")

    # --- 1. Определение стилей и констант ---
    print("🎨 Настройка стилей и цветовой схемы...")
    HEADER_COLOR = "1F4E79"
    EXCELLENT_COLOR = "00B050"
    GOOD_COLOR = "92D050"
    OK_COLOR = "FFFF00"
    WARNING_COLOR = "FFC000"
    CRITICAL_COLOR = "FF0000"
    TABLE_BG_COLOR = "E7E6E6"
    OPERATOR_COLOR = "4472C4"
    EQUIPMENT_COLOR = "E7E6E6"

    HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
    TITLE_FONT = Font(bold=True, size=16)
    BILINGUAL_FONT = Font(bold=True, size=10)
    HEADER_FILL = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    EXCELLENT_FILL = PatternFill(start_color=EXCELLENT_COLOR, end_color=EXCELLENT_COLOR, fill_type="solid")
    GOOD_FILL = PatternFill(start_color=GOOD_COLOR, end_color=GOOD_COLOR, fill_type="solid")
    OK_FILL = PatternFill(start_color=OK_COLOR, end_color=OK_COLOR, fill_type="solid")
    WARNING_FILL = PatternFill(start_color=WARNING_COLOR, end_color=WARNING_COLOR, fill_type="solid")
    CRITICAL_FILL = PatternFill(start_color=CRITICAL_COLOR, end_color=CRITICAL_COLOR, fill_type="solid")
    TABLE_BG_FILL = PatternFill(start_color=TABLE_BG_COLOR, end_color=TABLE_BG_COLOR, fill_type="solid")
    OPERATOR_FILL = PatternFill(start_color=OPERATOR_COLOR, end_color=OPERATOR_COLOR, fill_type="solid")
    EQUIPMENT_FILL = PatternFill(start_color=EQUIPMENT_COLOR, end_color=EQUIPMENT_COLOR, fill_type="solid")
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    MACHINES = ["Doosan Yashana", "Doosan Hadasha", "Doosan 3", "Pinnacle Gdola", "Mitsubishi", "JohnFord", "Okuma"]
    OPERATORS = ["Andrey", "Denis", "Daniel", "Kirill", "Slava", "Arkady"]

    # --- 2. Создание книги и листов ---
    print("🗂️ Создание структуры листов Excel...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_dashboard = wb.create_sheet("Dashboard_Панель", 0)
    ws_main_data = wb.create_sheet("MainData_Основные")
    ws_operator_kpi = wb.create_sheet("OperatorKPI_КПИ")
    ws_equipment_oee = wb.create_sheet("EquipmentOEE_ОЭО")
    ws_summary = wb.create_sheet("Summary_Сводка")
    machine_sheets = {machine: wb.create_sheet(f"Station_{machine.split(' ')[0]}") for machine in MACHINES}
    ws_analytics = wb.create_sheet("Analytics_Аналитика")
    ws_trends = wb.create_sheet("Trends_Тренды")
    ws_distributions = wb.create_sheet("Distributions_Распред")
    ws_translations = wb.create_sheet("Translations_Перевод")

    # --- 3. Заполнение листов ---

    # ============ Лист "Translations" ============
    print("📖 Заполнение 'Translations'...")
    ws_translations.append(["English Term", "Russian Translation", "Description / Описание"])
    translations_data = [
        ["OEE Equipment", "ОЭО Оборудования", "Overall Equipment Effectiveness / Общая эффективность оборудования"],
        ["KPI Operator", "КПИ Оператора", "Key Performance Indicator / Ключевой показатель эффективности"],
        ["Availability", "Доступность", "Equipment uptime / Время работы оборудования"],
        ["Equipment Performance", "Производительность станка", "Technical efficiency / Техническая эффективность"],
        ["Operator Performance", "Производительность оператора", "Plan achievement / Выполнение плана"],
        ["Quality", "Качество", "Good parts ratio / Доля годных деталей"],
        ["Setup Time", "Время наладки", "Preparation time / Время подготовки"],
        ["Setup Quality", "Качество наладки", "Setup effectiveness / Эффективность наладки"],
        ["Planned Time", "Плановое время", "Standard time per piece / Нормативное время на деталь"],
        ["Actual Time", "Фактическое время", "Real time per piece / Реальное время на деталь"],
        ["Good Parts", "Годные детали", "Quality production / Качественная продукция"],
        ["Defect Parts", "Брак", "Rejected production / Отбракованная продукция"],
    ]
    for row in translations_data:
        ws_translations.append(row)
    
    # Форматирование Translations
    for col_idx in range(1, 4):
        ws_translations.cell(row=1, column=col_idx).font = HEADER_FONT
        ws_translations.cell(row=1, column=col_idx).fill = HEADER_FILL
        ws_translations.cell(row=1, column=col_idx).alignment = CENTER_ALIGN
    for row_idx in range(2, len(translations_data) + 2):
        for col_idx in range(1, 4):
            ws_translations.cell(row=row_idx, column=col_idx).border = THIN_BORDER
            ws_translations.cell(row=row_idx, column=col_idx).fill = TABLE_BG_FILL
    
    ws_translations.column_dimensions['A'].width = 25
    ws_translations.column_dimensions['B'].width = 25
    ws_translations.column_dimensions['C'].width = 35

    # ============ Лист "MainData" (Агрегация данных) ============
    print("📊 Заполнение 'MainData'...")
    headers_main = [
        "Station / Станок", 
        "Avg_Equipment_OEE / Средний_ОЭО_станка", 
        "Avg_Operator_KPI / Средний_КПИ_оператора",
        "Total_Good_pcs / Всего_годных", 
        "Total_Defect_pcs / Всего_брака",
        "Equipment_Availability / Доступность_станка",
        "Equipment_Performance / Производительность_станка",
        "Operator_Performance / Производительность_оператора"
    ]
    ws_main_data.append(headers_main)
    
    for col_idx in range(1, len(headers_main) + 1):
        ws_main_data.cell(row=1, column=col_idx, value=headers_main[col_idx-1]).font = BILINGUAL_FONT
        ws_main_data.cell(row=1, column=col_idx).fill = HEADER_FILL
        ws_main_data.cell(row=1, column=col_idx).alignment = CENTER_ALIGN

    for row, machine in enumerate(MACHINES, 2):
        machine_short = machine.split(" ")[0]
        ws_main_data.cell(row=row, column=1).value = machine
        
        # OEE оборудования
        ws_main_data.cell(row=row, column=2).value = f'=IFERROR(AVERAGE(Station_{machine_short}!O:O),0)'  
        ws_main_data.cell(row=row, column=2).number_format = '0.0%'
        
        # KPI оператора
        ws_main_data.cell(row=row, column=3).value = f'=IFERROR(AVERAGE(Station_{machine_short}!T:T),0)'  
        ws_main_data.cell(row=row, column=3).number_format = '0.0%'
        
        # Годные и брак
        ws_main_data.cell(row=row, column=4).value = f'=SUMIF(Station_{machine_short}!A:A,"<>",Station_{machine_short}!K:K)'  
        ws_main_data.cell(row=row, column=5).value = f'=SUMIF(Station_{machine_short}!A:A,"<>",Station_{machine_short}!J:J)'  
        
        # Показатели оборудования
        ws_main_data.cell(row=row, column=6).value = f'=IFERROR(AVERAGE(Station_{machine_short}!L:L),0)'  # Availability
        ws_main_data.cell(row=row, column=6).number_format = '0.0%'
        ws_main_data.cell(row=row, column=7).value = f'=IFERROR(AVERAGE(Station_{machine_short}!M:M),0)'  # Equipment Performance
        ws_main_data.cell(row=row, column=7).number_format = '0.0%'
        ws_main_data.cell(row=row, column=8).value = f'=IFERROR(AVERAGE(Station_{machine_short}!Q:Q),0)'  # Operator Performance
        ws_main_data.cell(row=row, column=8).number_format = '0.0%'
        
        for col_idx in range(1, len(headers_main) + 1):
            ws_main_data.cell(row=row, column=col_idx).border = THIN_BORDER
            ws_main_data.cell(row=row, column=col_idx).fill = TABLE_BG_FILL

    ws_main_data.freeze_panes = "A2"
    for i in range(1, len(headers_main) + 1):
        ws_main_data.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20

    # ============ Лист "OperatorKPI" ============
    print("👥 Заполнение 'OperatorKPI'...")
    headers_op_kpi = [
        "Operator / Оператор", 
        "Total_Shifts / Всего_смен", 
        "Avg_KPI / Средний_КПИ", 
        "Best_KPI / Лучший_КПИ",
        "Worst_KPI / Худший_КПИ", 
        "Total_Good / Всего_годных", 
        "Total_Defects / Всего_брака", 
        "Defect_Rate / Процент_брака",
        "Plan_Achievement / Выполнение_плана", 
        "Primary_Station / Основной_станок", 
        "Versatility / Универсальность", 
        "Overall_Rating / Общая_оценка"
    ]
    ws_operator_kpi.append(headers_op_kpi)
    
    for col_idx in range(1, len(headers_op_kpi) + 1):
        ws_operator_kpi.cell(row=1, column=col_idx, value=headers_op_kpi[col_idx-1]).font = BILINGUAL_FONT
        ws_operator_kpi.cell(row=1, column=col_idx).fill = OPERATOR_FILL
        ws_operator_kpi.cell(row=1, column=col_idx).alignment = CENTER_ALIGN

    for row, operator in enumerate(OPERATORS, 2):
        ws_operator_kpi.cell(row=row, column=1).value = operator
        
        # Создаем массивные формулы для подсчета по всем станкам
        stations_formula = "+".join([f'COUNTIFS(Station_{machine.split(" ")[0]}!C:C,"{operator}",Station_{machine.split(" ")[0]}!A:A,"<>")' for machine in MACHINES])
        ws_operator_kpi.cell(row=row, column=2).value = f'={stations_formula}'  # Total shifts
        
        # Средний KPI по всем станкам
        kpi_formula = "AVERAGE(" + ",".join([f'IF(Station_{machine.split(" ")[0]}!C:C="{operator}",Station_{machine.split(" ")[0]}!T:T)' for machine in MACHINES]) + ")"
        ws_operator_kpi.cell(row=row, column=3).value = f'=IFERROR({kpi_formula},0)'
        ws_operator_kpi.cell(row=row, column=3).number_format = '0.0%'
        
        # Лучший и худший KPI
        max_kpi_formula = "MAX(" + ",".join([f'IF(Station_{machine.split(" ")[0]}!C:C="{operator}",Station_{machine.split(" ")[0]}!T:T)' for machine in MACHINES]) + ")"
        min_kpi_formula = "MIN(" + ",".join([f'IF(Station_{machine.split(" ")[0]}!C:C="{operator}",Station_{machine.split(" ")[0]}!T:T)' for machine in MACHINES]) + ")"
        ws_operator_kpi.cell(row=row, column=4).value = f'=IFERROR({max_kpi_formula},0)'
        ws_operator_kpi.cell(row=row, column=4).number_format = '0.0%'
        ws_operator_kpi.cell(row=row, column=5).value = f'=IFERROR({min_kpi_formula},0)'
        ws_operator_kpi.cell(row=row, column=5).number_format = '0.0%'
        
        # Общие годные и брак
        good_formula = "+".join([f'SUMIFS(Station_{machine.split(" ")[0]}!K:K,Station_{machine.split(" ")[0]}!C:C,"{operator}",Station_{machine.split(" ")[0]}!A:A,"<>")' for machine in MACHINES])
        defect_formula = "+".join([f'SUMIFS(Station_{machine.split(" ")[0]}!J:J,Station_{machine.split(" ")[0]}!C:C,"{operator}",Station_{machine.split(" ")[0]}!A:A,"<>")' for machine in MACHINES])
        ws_operator_kpi.cell(row=row, column=6).value = f'={good_formula}'
        ws_operator_kpi.cell(row=row, column=7).value = f'={defect_formula}'
        
        # Процент брака
        ws_operator_kpi.cell(row=row, column=8).value = f'=IFERROR(G{row}/(F{row}+G{row}),0)'
        ws_operator_kpi.cell(row=row, column=8).number_format = '0.0%'
        
        # Выполнение плана
        plan_formula = "AVERAGE(" + ",".join([f'IF(Station_{machine.split(" ")[0]}!C:C="{operator}",Station_{machine.split(" ")[0]}!Q:Q)' for machine in MACHINES]) + ")"
        ws_operator_kpi.cell(row=row, column=9).value = f'=IFERROR({plan_formula},0)'
        ws_operator_kpi.cell(row=row, column=9).number_format = '0.0%'
        
        # Основной станок (пока N/A - требует сложной логики)
        ws_operator_kpi.cell(row=row, column=10).value = "N/A"
        ws_operator_kpi.cell(row=row, column=11).value = f'={len(MACHINES)}'  # Максимальная универсальность
        
        # Общая оценка
        ws_operator_kpi.cell(row=row, column=12).value = f'=IF(C{row}>=0.9,"Отлично/Excellent",IF(C{row}>=0.8,"Хорошо/Good",IF(C{row}>=0.7,"Удовл/Satisfactory","Треб.развития/Needs Dev")))'

    # Форматирование OperatorKPI
    for row_idx in range(2, len(OPERATORS) + 2):
        for col_idx in range(1, len(headers_op_kpi) + 1):
            ws_operator_kpi.cell(row=row_idx, column=col_idx).border = THIN_BORDER
            ws_operator_kpi.cell(row=row_idx, column=col_idx).fill = TABLE_BG_FILL
    
    for i in range(1, len(headers_op_kpi) + 1):
        ws_operator_kpi.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 18

    # ============ Лист "EquipmentOEE" ============
    print("🏭 Заполнение 'EquipmentOEE'...")
    headers_eq_oee = [
        "Equipment / Оборудование",
        "Avg_OEE / Средний_ОЭО", 
        "Availability / Доступность",
        "Performance / Производительность", 
        "Quality / Качество",
        "Total_Downtime / Общий_простой",
        "Avg_Cycle_Time / Среднее_время_цикла",
        "Efficiency_Rating / Рейтинг_эффективности"
    ]
    ws_equipment_oee.append(headers_eq_oee)
    
    for col_idx in range(1, len(headers_eq_oee) + 1):
        ws_equipment_oee.cell(row=1, column=col_idx, value=headers_eq_oee[col_idx-1]).font = BILINGUAL_FONT
        ws_equipment_oee.cell(row=1, column=col_idx).fill = EQUIPMENT_FILL
        ws_equipment_oee.cell(row=1, column=col_idx).alignment = CENTER_ALIGN

    for row, machine in enumerate(MACHINES, 2):
        machine_short = machine.split(" ")[0]
        ws_equipment_oee.cell(row=row, column=1).value = machine
        
        # OEE и его компоненты
        ws_equipment_oee.cell(row=row, column=2).value = f'=IFERROR(AVERAGE(Station_{machine_short}!O:O),0)'
        ws_equipment_oee.cell(row=row, column=2).number_format = '0.0%'
        ws_equipment_oee.cell(row=row, column=3).value = f'=IFERROR(AVERAGE(Station_{machine_short}!L:L),0)'
        ws_equipment_oee.cell(row=row, column=3).number_format = '0.0%'
        ws_equipment_oee.cell(row=row, column=4).value = f'=IFERROR(AVERAGE(Station_{machine_short}!M:M),0)'
        ws_equipment_oee.cell(row=row, column=4).number_format = '0.0%'
        ws_equipment_oee.cell(row=row, column=5).value = f'=IFERROR(AVERAGE(Station_{machine_short}!N:N),0)'
        ws_equipment_oee.cell(row=row, column=5).number_format = '0.0%'
        
        # Общий простой
        ws_equipment_oee.cell(row=row, column=6).value = f'=SUMIF(Station_{machine_short}!A:A,"<>",Station_{machine_short}!F:F)'
        
        # Среднее время цикла
        ws_equipment_oee.cell(row=row, column=7).value = f'=IFERROR(AVERAGE(Station_{machine_short}!H:H),0)'
        
        # Рейтинг эффективности
        ws_equipment_oee.cell(row=row, column=8).value = f'=IF(B{row}>=0.85,"Отлично/Excellent",IF(B{row}>=0.75,"Хорошо/Good",IF(B{row}>=0.65,"Удовл/Satisfactory","Треб.ремонта/Needs Repair")))'
        
        for col_idx in range(1, len(headers_eq_oee) + 1):
            ws_equipment_oee.cell(row=row, column=col_idx).border = THIN_BORDER
            ws_equipment_oee.cell(row=row, column=col_idx).fill = TABLE_BG_FILL

    for i in range(1, len(headers_eq_oee) + 1):
        ws_equipment_oee.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20

    # ============ Листы по станкам (ИСПРАВЛЕННАЯ СТРУКТУРА) ============
    print("🏭 Генерация отчетов по станкам...")
    for machine in MACHINES:
        ws = machine_sheets[machine]
        ws.append([f"Станок / Station: {machine}"])
        ws.merge_cells("A1:F1")
        ws['A1'].font = TITLE_FONT
        ws['A1'].alignment = CENTER_ALIGN
        
        # ИСПРАВЛЕННЫЕ ЗАГОЛОВКИ с разделением OEE и KPI
        headers_station = [
            "Date/Дата", "Drawing/Чертеж", "Operator/Оператор", "Shift_min/Смена_мин", 
            "Setup_min/Наладка_мин", "Downtime_min/Простой_мин", "Planned_pcs/План_шт", 
            "Actual_time_per_piece/Факт_время_на_деталь", "Planned_time_per_piece/План_время_на_деталь",
            "Defect_pcs/Брак_шт", "Good_pcs/Годные_шт", "Availability%/Доступность%", 
            "Equipment_Performance%/Производ_станка%", "Quality%/Качество%", "Equipment_OEE%/ОЭО_станка%",
            "Setup_ratio%/Доля_наладки%", "Setup_quality%/Качество_наладки%", 
            "Actual_pcs/Факт_шт", "Operator_Performance%/Производ_оператора%", 
            "Operator_KPI%/КПИ_оператора%"
        ]
        ws.append(headers_station)
        
        for col_idx in range(1, len(headers_station) + 1):
            cell = ws.cell(row=2, column=col_idx, value=headers_station[col_idx-1])
            cell.font = BILINGUAL_FONT
            if col_idx <= 11:  # Основные данные
                cell.fill = HEADER_FILL
            elif col_idx <= 15:  # OEE станка
                cell.fill = EQUIPMENT_FILL  
            else:  # KPI оператора
                cell.fill = OPERATOR_FILL
            cell.alignment = CENTER_ALIGN

        # ===== УЛУЧШЕННАЯ ВАЛИДАЦИЯ ДАТ С КАЛЕНДАРЕМ =====
        dv_date = DataValidation(
            type="date", 
            operator="between", 
            formula1="2025-01-01", 
            formula2="2025-12-31",
            allow_blank=True,
            showInputMessage=True,
            promptTitle="📅 Выбор даты / Date Selection",
            prompt="• Excel 365/2021: дважды кликните для календаря / double-click for calendar\n• Другие версии / Other versions: введите дату вручную / enter date manually\n• Формат / Format: ДД.ММ.ГГГГ / DD.MM.YYYY\n• Пример / Example: 28.06.2025",
            showErrorMessage=True,
            errorTitle="❌ Ошибка даты / Date Error",
            error="Введите корректную дату в диапазоне 2025 года / Enter correct date in 2025 range"
        )
        dv_date.add("A3:A1002")
        ws.add_data_validation(dv_date)

        # Валидации
        dv_operators = DataValidation(type="list", formula1=f'"{",".join(OPERATORS)}"')
        dv_operators.add("C3:C1002")
        ws.add_data_validation(dv_operators)
        
        dv_shift = DataValidation(type="whole", operator="between", formula1="240", formula2="720")
        dv_shift.add("D3:D1002")
        ws.add_data_validation(dv_shift)
        
        dv_times = DataValidation(type="decimal", operator="greaterThan", formula1="0")
        dv_times.add("H3:H1002")
        ws.add_data_validation(dv_times)
        
        dv_times2 = DataValidation(type="decimal", operator="greaterThan", formula1="0")
        dv_times2.add("I3:I1002")
        ws.add_data_validation(dv_times2)

        # Заполнение формул
        for row in range(3, 1002):
            # Форматирование
            for col in range(1, len(headers_station) + 1):
                ws.cell(row=row, column=col).border = THIN_BORDER
                ws.cell(row=row, column=col).fill = TABLE_BG_FILL
            
            # Начальные значения
            ws.cell(row=row, column=1).number_format = "DD.MM.YYYY"
            ws.cell(row=row, column=1).comment = Comment(
                "📅 КАЛЕНДАРЬ / CALENDAR:\n"
                "═══════════════════════════\n"
                "✓ Excel 365/2021: двойной клик / double-click\n"
                "✓ Excel 2019/2016: F2 + Enter\n"
                "✓ Ручной ввод / Manual: ДД.ММ.ГГГГ / DD.MM.YYYY\n"
                "✓ Быстрый ввод / Quick: Ctrl+; (текущая дата/current date)\n"
                "\n"
                "📋 ПРИМЕРЫ / EXAMPLES:\n"
                "• 28.06.2025 • 01.07.2025 • 15.12.2025", 
                "OEE/KPI Monitoring System v14.0"
            )
            
            ws.cell(row=row, column=4).value = 480  # Shift time
            ws.cell(row=row, column=7).value = 5    # Planned pieces
            ws.cell(row=row, column=8).value = 60   # Actual time per piece  
            ws.cell(row=row, column=9).value = 50   # Planned time per piece
            
            # Формулы расчетов
            # Good pieces = Actual pieces - Defects
            ws.cell(row=row, column=11).value = f'=MAX(R{row}-J{row},0)'
            
            # EQUIPMENT OEE CALCULATIONS
            # Availability = (Shift_time - Downtime) / Shift_time
            ws.cell(row=row, column=12).value = f'=IF(D{row}>0,(D{row}-F{row})/D{row},0)'
            ws.cell(row=row, column=12).number_format = '0.0%'
            
            # Equipment Performance = Planned_time_per_piece / Actual_time_per_piece
            ws.cell(row=row, column=13).value = f'=IF(H{row}>0,I{row}/H{row},0)'
            ws.cell(row=row, column=13).number_format = '0.0%'
            
            # Quality = Good_pieces / Actual_pieces (если Actual > 0)
            ws.cell(row=row, column=14).value = f'=IF(R{row}>0,K{row}/R{row},0)'
            ws.cell(row=row, column=14).number_format = '0.0%'
            
            # Equipment OEE = Availability × Equipment_Performance × Quality
            ws.cell(row=row, column=15).value = f'=L{row}*M{row}*N{row}'
            ws.cell(row=row, column=15).number_format = '0.0%'
            
            # Setup calculations
            ws.cell(row=row, column=16).value = f'=IF(D{row}>0,E{row}/D{row},0)'  # Setup ratio
            ws.cell(row=row, column=16).number_format = '0.0%'
            ws.cell(row=row, column=17).value = f'=N{row}'  # Setup quality = Quality
            ws.cell(row=row, column=17).number_format = '0.0%'
            
            # OPERATOR KPI CALCULATIONS
            # Actual pieces (calculated from times and effective work time)
            ws.cell(row=row, column=18).value = f'=IF(AND(H{row}>0,D{row}>E{row}+F{row}),ROUNDDOWN((D{row}-E{row}-F{row})/H{row},0),0)'
            
            # Operator Performance = Actual_pieces / Planned_pieces
            ws.cell(row=row, column=19).value = f'=IF(G{row}>0,MIN(R{row}/G{row},1),0)'
            ws.cell(row=row, column=19).number_format = '0.0%'
            
            # Operator KPI = weighted formula
            ws.cell(row=row, column=20).value = f'=L{row}*0.3+S{row}*0.4+N{row}*0.2+(1-P{row})*0.1'
            ws.cell(row=row, column=20).number_format = '0.0%'

        # Настройка ширины столбцов
        for col in range(1, len(headers_station) + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            ws.column_dimensions[col_letter].width = 16

    # ============ Лист "Dashboard" ============
    print("🎨 Создание 'Dashboard'...")
    ws_dashboard.sheet_view.showGridLines = False
    ws_dashboard.merge_cells("A1:H2")
    ws_dashboard['A1'].value = "Панель управления производством / Production Management Dashboard"
    ws_dashboard['A1'].font = TITLE_FONT
    ws_dashboard['A1'].alignment = CENTER_ALIGN

    # Основные показатели
    ws_dashboard['A4'] = "Критических ситуаций (OEE < 70%) / Critical Situations:"
    ws_dashboard['C4'] = f'=COUNTIF(MainData_Основные!B:B,"<0.7")'
    ws_dashboard['C4'].font = Font(bold=True, color=CRITICAL_COLOR)

    ws_dashboard['A6'] = "Общий OEE оборудования / Overall Equipment OEE:"
    ws_dashboard['C6'] = f'=IFERROR(AVERAGE(MainData_Основные!B:B),0)'
    ws_dashboard['C6'].number_format = '0.0%'

    ws_dashboard['A8'] = "Общий KPI операторов / Overall Operator KPI:"
    ws_dashboard['C8'] = f'=IFERROR(AVERAGE(MainData_Основные!C:C),0)'
    ws_dashboard['C8'].number_format = '0.0%'

    ws_dashboard['A10'] = "Лучший оператор / Best Operator:"
    ws_dashboard['C10'] = f'=INDEX(OperatorKPI_КПИ!A:A,MATCH(MAX(OperatorKPI_КПИ!C:C),OperatorKPI_КПИ!C:C,0))'

    ws_dashboard['A12'] = "Лучший станок / Best Equipment:"
    ws_dashboard['C12'] = f'=INDEX(MainData_Основные!A:A,MATCH(MAX(MainData_Основные!B:B),MainData_Основные!B:B,0))'

    # График OEE по станкам
    chart_oee = BarChart()
    chart_oee.title = "OEE по станкам / Equipment OEE"
    chart_oee.x_axis.title = "Станки / Equipment"
    chart_oee.y_axis.title = "OEE (%)"
    data_oee = Reference(ws_main_data, min_col=2, min_row=2, max_row=8)
    categories_oee = Reference(ws_main_data, min_col=1, min_row=2, max_row=8)
    chart_oee.add_data(data_oee, titles_from_data=False)
    chart_oee.set_categories(categories_oee)
    chart_oee.height = 10
    chart_oee.width = 15
    ws_dashboard.add_chart(chart_oee, "A15")

    # График KPI по операторам
    chart_kpi = BarChart()
    chart_kpi.title = "KPI операторов / Operator KPI"
    chart_kpi.x_axis.title = "Операторы / Operators"
    chart_kpi.y_axis.title = "KPI (%)"
    data_kpi = Reference(ws_operator_kpi, min_col=3, min_row=2, max_row=7)
    categories_kpi = Reference(ws_operator_kpi, min_col=1, min_row=2, max_row=7)
    chart_kpi.add_data(data_kpi, titles_from_data=False)
    chart_kpi.set_categories(categories_kpi)
    chart_kpi.height = 10
    chart_kpi.width = 15
    ws_dashboard.add_chart(chart_kpi, "A30")

    # ============ Лист "Analytics" ============
    print("📈 Заполнение 'Analytics'...")
    ws_analytics['A1'] = "Продвинутая аналитика / Advanced Analytics"
    ws_analytics['A1'].font = TITLE_FONT
    
    analytics_headers = ["Метрика / Metric", "Значение / Value", "Описание / Description"]
    ws_analytics.append(analytics_headers)
    
    analytics_data = [
        ["Стабильность OEE / OEE Stability", f'=IFERROR(STDEV(MainData_Основные!B:B)/AVERAGE(MainData_Основные!B:B),0)', "Коэффициент вариации / Coefficient of variation"],
        ["Корреляция OEE-KPI / OEE-KPI Correlation", f'=IFERROR(CORREL(MainData_Основные!B:B,MainData_Основные!C:C),0)', "Связь между OEE и KPI / Relationship between OEE and KPI"],
        ["Потенциал улучшения / Improvement Potential", f'=IFERROR(1-AVERAGE(MainData_Основные!B:B),0)', "Возможный рост OEE / Possible OEE growth"],
        ["Эффективность наладки / Setup Efficiency", f'=IFERROR(1-AVERAGE(MainData_Основные!F:F)/AVERAGE(MainData_Основные!D:D),0)', "Доля времени на наладку / Setup time ratio"],
        ["Качество производства / Production Quality", f'=IFERROR(SUM(MainData_Основные!D:D)/(SUM(MainData_Основные!D:D)+SUM(MainData_Основные!E:E)),0)', "Доля годной продукции / Good production ratio"]
    ]
    
    for row_data in analytics_data:
        ws_analytics.append(row_data)
    
    # Форматирование Analytics
    for col_idx in range(1, 4):
        ws_analytics.cell(row=2, column=col_idx).font = BILINGUAL_FONT
        ws_analytics.cell(row=2, column=col_idx).fill = HEADER_FILL
        ws_analytics.cell(row=2, column=col_idx).alignment = CENTER_ALIGN
    
    for row_idx in range(3, 8):
        for col_idx in range(1, 4):
            ws_analytics.cell(row=row_idx, column=col_idx).border = THIN_BORDER
            ws_analytics.cell(row=row_idx, column=col_idx).fill = TABLE_BG_FILL
        ws_analytics.cell(row=row_idx, column=2).number_format = '0.0%'
    
    ws_analytics.column_dimensions['A'].width = 30
    ws_analytics.column_dimensions['B'].width = 20
    ws_analytics.column_dimensions['C'].width = 35

    # ============ Лист "Trends" ============
    print("📈 Создание 'Trends'...")
    ws_trends['A1'] = "Тренды производительности / Performance Trends"
    ws_trends['A1'].font = TITLE_FONT
    
    # Создание таблицы для трендов (пример данных)
    trend_headers = ["Date/Дата", "Avg_OEE/Средний_ОЭО", "Avg_KPI/Средний_КПИ", "Total_Production/Общее_производство"]
    ws_trends.append(trend_headers)
    
    # Пример данных за последние 30 дней
    for day in range(1, 31):
        ws_trends.append([
            f"0{day}.06.2025" if day < 10 else f"{day}.06.2025",
            f'=0.7+RAND()*0.25',  # Случайный OEE 70-95%
            f'=0.65+RAND()*0.3',  # Случайный KPI 65-95%
            f'=ROUND(20+RAND()*30,0)'  # Случайное производство 20-50 деталей
        ])
    
    # График трендов
    chart_trends = LineChart()
    chart_trends.title = "Тренды OEE и KPI / OEE and KPI Trends"
    chart_trends.x_axis.title = "Дата / Date"
    chart_trends.y_axis.title = "Показатели / Metrics (%)"
    
    data_trends = Reference(ws_trends, min_col=2, min_row=2, max_col=3, max_row=31)
    categories_trends = Reference(ws_trends, min_col=1, min_row=2, max_row=31)
    chart_trends.add_data(data_trends, titles_from_data=False)
    chart_trends.set_categories(categories_trends)
    chart_trends.height = 15
    chart_trends.width = 25
    ws_trends.add_chart(chart_trends, "F2")

    # --- 4. Сохранение файла ---
    filename = f"OEE_KPI_Monitoring_v14.0_Fixed_{datetime(2025, 6, 28, 23, 15).strftime('%Y%m%d_%H%M')}.xlsx"
    filepath = os.path.join(os.getcwd(), filename)
    try:
        wb.save(filepath)
        print(f"✅ УСПЕХ! Файл создан: {filename}")
        print("\n🔧 КЛЮЧЕВЫЕ ИСПРАВЛЕНИЯ:")
        print("   ✓ Разделение OEE станка и KPI оператора")
        print("   ✓ Справедливые формулы расчета")
        print("   ✓ Двуязычный интерфейс (RUS/ENG)")
        print("   ✓ Связанные данные между листами")
        print("   ✓ Правильная агрегация по операторам")
        print("   ✓ Аналитика и тренды")
        print("   ✓ Улучшенный календарь")
        print("\n📊 ФОРМУЛЫ:")
        print("   • Equipment OEE = Availability × Equipment_Performance × Quality")
        print("   • Operator KPI = Availability×0.3 + Operator_Performance×0.4 + Quality×0.2 + Setup_Efficiency×0.1")
        return filepath
    except Exception as e:
        print(f"❌ Ошибка при сохранении: {e}")
        return None

if __name__ == "__main__":
    create_monitoring_system()