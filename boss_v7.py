#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
СИСТЕМА МОНИТОРИНГА OEE/KPI v7.0 - УПРОЩЕННАЯ ЛОГИЧНАЯ ВЕРСИЯ
Основано на классической формуле OEE с учетом наладки
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.comments import Comment
from datetime import datetime
import os

def create_monitoring_system():
    """
    Основная функция для создания Excel-файла мониторинга OEE/KPI.
    """
    print("🚀 Запуск создания УПРОЩЕННОЙ системы мониторинга v7.0...")

    # --- 1. Определение стилей и констант ---
    print("🎨 Настройка стилей и цветовой схемы...")
    HEADER_COLOR = "1F4E79"
    EXCELLENT_COLOR = "00B050"  # 85%+ (Мировой класс)
    GOOD_COLOR = "92D050"       # 70-84% (Хороший уровень)
    AVERAGE_COLOR = "FFFF00"    # 60-69% (Средний уровень)
    POOR_COLOR = "FF0000"       # <60% (Требует улучшения)
    TABLE_BG_COLOR = "E7E6E6"

    HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
    TITLE_FONT = Font(bold=True, size=16)
    BILINGUAL_FONT = Font(bold=True, size=10)
    HEADER_FILL = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    EXCELLENT_FILL = PatternFill(start_color=EXCELLENT_COLOR, end_color=EXCELLENT_COLOR, fill_type="solid")
    GOOD_FILL = PatternFill(start_color=GOOD_COLOR, end_color=GOOD_COLOR, fill_type="solid")
    AVERAGE_FILL = PatternFill(start_color=AVERAGE_COLOR, end_color=AVERAGE_COLOR, fill_type="solid")
    POOR_FILL = PatternFill(start_color=POOR_COLOR, end_color=POOR_COLOR, fill_type="solid")
    TABLE_BG_FILL = PatternFill(start_color=TABLE_BG_COLOR, end_color=TABLE_BG_COLOR, fill_type="solid")
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    MACHINES = ["Doosan Yashana", "Doosan Hadasha", "Doosan 3", "Pinnacle Gdola", "Mitsubishi", "JohnFord", "Okuma"]
    OPERATORS = ["Andrey", "Denis", "Daniel", "Kirill", "Slava", "Arkady"]

    # --- 2. Создание книги и листов ---
    print("🗂️ Создание структуры листов Excel...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_dashboard = wb.create_sheet("Dashboard_Панель", 0)
    ws_main_data = wb.create_sheet("MainData_Основные")
    ws_operator_summary = wb.create_sheet("OperatorSummary_Операторы")
    machine_sheets = {machine: wb.create_sheet(f"Station_{machine.split(' ')[0]}") for machine in MACHINES}
    ws_analytics = wb.create_sheet("Analytics_Аналитика")
    ws_trends = wb.create_sheet("Trends_Тренды")
    ws_translations = wb.create_sheet("Translations_Перевод")

    # --- 3. Заполнение листов ---

    # ============ Лист "Translations" ============
    print("📖 Заполнение 'Translations'...")
    ws_translations.append(["English Term", "Russian Translation", "Description / Описание"])
    translations_data = [
        ["OEE", "Общая эффективность оборудования", "Overall Equipment Effectiveness"],
        ["Availability", "Доступность", "Equipment uptime percentage"],
        ["Performance", "Производительность", "Speed efficiency ratio"],
        ["Quality", "Качество", "Good parts ratio"],
        ["Setup Time", "Время наладки", "Time to first good part"],
        ["Downtime", "Простои", "Unplanned stops"],
        ["Planned Output", "Плановый выпуск", "Target production quantity"],
        ["Actual Output", "Фактический выпуск", "Real production quantity"],
        ["Defects", "Брак", "Rejected parts"],
        ["Good Parts", "Годные детали", "Quality production"],
        ["World Class", "Мировой класс", "85%+ OEE"],
        ["Good Level", "Хороший уровень", "70-84% OEE"],
        ["Average Level", "Средний уровень", "60-69% OEE"],
        ["Needs Improvement", "Требует улучшения", "<60% OEE"],
    ]
    for row in translations_data:
        ws_translations.append(row)
    
    # Форматирование Translations
    for col_idx in range(1, 4):
        ws_translations.cell(row=1, column=col_idx).font = HEADER_FONT
        ws_translations.cell(row=1, column=col_idx).fill = HEADER_FILL
        ws_translations.cell(row=1, column=col_idx).alignment = CENTER_ALIGN
    for row_idx in range(2, len(translations_data) + 2):
        for col_idx in range(1, 4):
            ws_translations.cell(row=row_idx, column=col_idx).border = THIN_BORDER
            ws_translations.cell(row=row_idx, column=col_idx).fill = TABLE_BG_FILL
    
    ws_translations.column_dimensions['A'].width = 25
    ws_translations.column_dimensions['B'].width = 30
    ws_translations.column_dimensions['C'].width = 35

    # ============ Лист "MainData" (Агрегация данных) ============
    print("📊 Заполнение 'MainData'...")
    headers_main = [
        "Station / Станок", 
        "Avg_OEE / Средний_ОЭО", 
        "Avg_Availability / Средняя_доступность",
        "Avg_Performance / Средняя_производительность", 
        "Avg_Quality / Среднее_качество",
        "Total_Good_Parts / Всего_годных", 
        "Total_Defects / Всего_брака",
        "Total_Setup_Time / Общее_время_наладки",
        "Avg_Shift_Time / Среднее_время_смены",
        "OEE_Rating / Рейтинг_ОЭО"
    ]
    ws_main_data.append(headers_main)
    
    for col_idx in range(1, len(headers_main) + 1):
        ws_main_data.cell(row=1, column=col_idx, value=headers_main[col_idx-1]).font = BILINGUAL_FONT
        ws_main_data.cell(row=1, column=col_idx).fill = HEADER_FILL
        ws_main_data.cell(row=1, column=col_idx).alignment = CENTER_ALIGN

    for row, machine in enumerate(MACHINES, 2):
        machine_short = machine.split(" ")[0]
        ws_main_data.cell(row=row, column=1).value = machine
        
        # Основные показатели OEE
        ws_main_data.cell(row=row, column=2).value = f'=IFERROR(AVERAGE(Station_{machine_short}!L:L),0)'  # OEE
        ws_main_data.cell(row=row, column=2).number_format = '0.0%'
        
        ws_main_data.cell(row=row, column=3).value = f'=IFERROR(AVERAGE(Station_{machine_short}!I:I),0)'  # Availability
        ws_main_data.cell(row=row, column=3).number_format = '0.0%'
        
        ws_main_data.cell(row=row, column=4).value = f'=IFERROR(AVERAGE(Station_{machine_short}!J:J),0)'  # Performance
        ws_main_data.cell(row=row, column=4).number_format = '0.0%'
        
        ws_main_data.cell(row=row, column=5).value = f'=IFERROR(AVERAGE(Station_{machine_short}!K:K),0)'  # Quality
        ws_main_data.cell(row=row, column=5).number_format = '0.0%'
        
        # Суммарные показатели
        ws_main_data.cell(row=row, column=6).value = f'=SUMIF(Station_{machine_short}!A:A,"<>",Station_{machine_short}!H:H)'  # Good parts
        ws_main_data.cell(row=row, column=7).value = f'=SUMIF(Station_{machine_short}!A:A,"<>",Station_{machine_short}!G:G)'  # Defects
        ws_main_data.cell(row=row, column=8).value = f'=SUMIF(Station_{machine_short}!A:A,"<>",Station_{machine_short}!E:E)'  # Setup time
        ws_main_data.cell(row=row, column=9).value = f'=IFERROR(AVERAGE(Station_{machine_short}!D:D),0)'  # Shift time
        
        # Рейтинг OEE
        ws_main_data.cell(row=row, column=10).value = f'=IF(B{row}>=0.85,"Мировой класс / World Class",IF(B{row}>=0.7,"Хороший / Good",IF(B{row}>=0.6,"Средний / Average","Требует улучшения / Needs Improvement")))'
        
        for col_idx in range(1, len(headers_main) + 1):
            ws_main_data.cell(row=row, column=col_idx).border = THIN_BORDER
            ws_main_data.cell(row=row, column=col_idx).fill = TABLE_BG_FILL

    # Условное форматирование для OEE
    ws_main_data.conditional_formatting.add('B2:B8', CellIsRule(operator='greaterThanOrEqual', formula=['0.85'], fill=EXCELLENT_FILL))
    ws_main_data.conditional_formatting.add('B2:B8', CellIsRule(operator='between', formula=['0.7', '0.84'], fill=GOOD_FILL))
    ws_main_data.conditional_formatting.add('B2:B8', CellIsRule(operator='between', formula=['0.6', '0.69'], fill=AVERAGE_FILL))
    ws_main_data.conditional_formatting.add('B2:B8', CellIsRule(operator='lessThan', formula=['0.6'], fill=POOR_FILL))

    ws_main_data.freeze_panes = "A2"
    for i in range(1, len(headers_main) + 1):
        ws_main_data.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 22

    # ============ Лист "OperatorSummary" ============
    print("👥 Заполнение 'OperatorSummary'...")
    headers_op = [
        "Operator / Оператор", 
        "Total_Shifts / Всего_смен", 
        "Avg_OEE / Средний_ОЭО", 
        "Best_OEE / Лучший_ОЭО",
        "Total_Good_Parts / Всего_годных", 
        "Total_Defects / Всего_брака", 
        "Defect_Rate / Процент_брака",
        "Avg_Setup_Time / Среднее_время_наладки",
        "Performance_Rating / Рейтинг_работы"
    ]
    ws_operator_summary.append(headers_op)
    
    for col_idx in range(1, len(headers_op) + 1):
        ws_operator_summary.cell(row=1, column=col_idx, value=headers_op[col_idx-1]).font = BILINGUAL_FONT
        ws_operator_summary.cell(row=1, column=col_idx).fill = HEADER_FILL
        ws_operator_summary.cell(row=1, column=col_idx).alignment = CENTER_ALIGN

    for row, operator in enumerate(OPERATORS, 2):
        ws_operator_summary.cell(row=row, column=1).value = operator
        
        # Подсчет смен по всем станкам
        shifts_formula = "+".join([f'COUNTIFS(Station_{machine.split(" ")[0]}!C:C,"{operator}",Station_{machine.split(" ")[0]}!A:A,"<>")' for machine in MACHINES])
        ws_operator_summary.cell(row=row, column=2).value = f'={shifts_formula}'
        
        # Средний OEE по всем станкам (используем массивную формулу)
        oee_ranges = [f'IF(Station_{machine.split(" ")[0]}!C:C="{operator}",Station_{machine.split(" ")[0]}!L:L)' for machine in MACHINES]
        avg_oee_formula = f'=IFERROR(AVERAGE({",".join(oee_ranges)}),0)'
        ws_operator_summary.cell(row=row, column=3).value = avg_oee_formula
        ws_operator_summary.cell(row=row, column=3).number_format = '0.0%'
        
        # Лучший OEE
        max_oee_formula = f'=IFERROR(MAX({",".join(oee_ranges)}),0)'
        ws_operator_summary.cell(row=row, column=4).value = max_oee_formula
        ws_operator_summary.cell(row=row, column=4).number_format = '0.0%'
        
        # Общее количество годных и брака
        good_formula = "+".join([f'SUMIFS(Station_{machine.split(" ")[0]}!H:H,Station_{machine.split(" ")[0]}!C:C,"{operator}",Station_{machine.split(" ")[0]}!A:A,"<>")' for machine in MACHINES])
        defect_formula = "+".join([f'SUMIFS(Station_{machine.split(" ")[0]}!G:G,Station_{machine.split(" ")[0]}!C:C,"{operator}",Station_{machine.split(" ")[0]}!A:A,"<>")' for machine in MACHINES])
        ws_operator_summary.cell(row=row, column=5).value = f'={good_formula}'
        ws_operator_summary.cell(row=row, column=6).value = f'={defect_formula}'
        
        # Процент брака
        ws_operator_summary.cell(row=row, column=7).value = f'=IFERROR(F{row}/(E{row}+F{row}),0)'
        ws_operator_summary.cell(row=row, column=7).number_format = '0.0%'
        
        # Среднее время наладки
        setup_ranges = [f'IF(Station_{machine.split(" ")[0]}!C:C="{operator}",Station_{machine.split(" ")[0]}!E:E)' for machine in MACHINES]
        avg_setup_formula = f'=IFERROR(AVERAGE({",".join(setup_ranges)}),0)'
        ws_operator_summary.cell(row=row, column=8).value = avg_setup_formula
        
        # Рейтинг работы
        ws_operator_summary.cell(row=row, column=9).value = f'=IF(C{row}>=0.85,"Отлично / Excellent",IF(C{row}>=0.75,"Хорошо / Good",IF(C{row}>=0.65,"Удовлетворительно / Satisfactory","Требует развития / Needs Development")))'

    # Форматирование OperatorSummary
    for row_idx in range(2, len(OPERATORS) + 2):
        for col_idx in range(1, len(headers_op) + 1):
            ws_operator_summary.cell(row=row_idx, column=col_idx).border = THIN_BORDER
            ws_operator_summary.cell(row=row_idx, column=col_idx).fill = TABLE_BG_FILL
    
    for i in range(1, len(headers_op) + 1):
        ws_operator_summary.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20

    # ============ Листы по станкам (УПРОЩЕННАЯ СТРУКТУРА) ============
    print("🏭 Генерация отчетов по станкам...")
    for machine in MACHINES:
        ws = machine_sheets[machine]
        ws.append([f"Станок / Station: {machine}"])
        ws.merge_cells("A1:F1")
        ws['A1'].font = TITLE_FONT
        ws['A1'].alignment = CENTER_ALIGN
        
        # ПРОСТЫЕ И ЛОГИЧНЫЕ ЗАГОЛОВКИ
        headers_station = [
            "Date / Дата", 
            "Drawing / Чертеж", 
            "Operator / Оператор", 
            "Shift_Time_min / Время_смены_мин", 
            "Setup_Time_min / Время_наладки_мин", 
            "Downtime_min / Простои_мин",
            "Defects / Брак_шт", 
            "Good_Parts / Годные_шт", 
            "Availability% / Доступность%", 
            "Performance% / Производительность%", 
            "Quality% / Качество%", 
            "OEE% / ОЭО%"
        ]
        ws.append(headers_station)
        
        for col_idx in range(1, len(headers_station) + 1):
            cell = ws.cell(row=2, column=col_idx, value=headers_station[col_idx-1])
            cell.font = BILINGUAL_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN

        # ===== ВАЛИДАЦИИ =====
        # Даты
        dv_date = DataValidation(
            type="date", 
            operator="between", 
            formula1="2025-01-01", 
            formula2="2025-12-31",
            allow_blank=True,
            showInputMessage=True,
            promptTitle="📅 Выбор даты / Date Selection",
            prompt="• Excel 365/2021: дважды кликните для календаря / double-click for calendar\n• Другие версии / Other versions: введите дату вручную / enter date manually\n• Формат / Format: ДД.ММ.ГГГГ / DD.MM.YYYY\n• Пример / Example: 28.06.2025",
            showErrorMessage=True,
            errorTitle="❌ Ошибка даты / Date Error",
            error="Введите корректную дату в диапазоне 2025 года / Enter correct date in 2025 range"
        )
        dv_date.add("A3:A1002")
        ws.add_data_validation(dv_date)

        # Операторы
        dv_operators = DataValidation(type="list", formula1=f'"{",".join(OPERATORS)}"')
        dv_operators.add("C3:C1002")
        ws.add_data_validation(dv_operators)
        
        # Время смены (4-12 часов)
        dv_shift = DataValidation(type="whole", operator="between", formula1="240", formula2="720")
        dv_shift.add("D3:D1002")
        ws.add_data_validation(dv_shift)
        
        # Время наладки (0-4 часа)
        dv_setup = DataValidation(type="whole", operator="between", formula1="0", formula2="240")
        dv_setup.add("E3:E1002")
        ws.add_data_validation(dv_setup)
        
        # Простои (0-8 часов)
        dv_downtime = DataValidation(type="whole", operator="between", formula1="0", formula2="480")
        dv_downtime.add("F3:F1002")
        ws.add_data_validation(dv_downtime)

        # Заполнение формул и данных
        for row in range(3, 1002):
            # Форматирование
            for col in range(1, len(headers_station) + 1):
                ws.cell(row=row, column=col).border = THIN_BORDER
                ws.cell(row=row, column=col).fill = TABLE_BG_FILL
            
            # Форматирование даты
            ws.cell(row=row, column=1).number_format = "DD.MM.YYYY"
            ws.cell(row=row, column=1).comment = Comment(
                "📅 КАЛЕНДАРЬ / CALENDAR:\n"
                "═══════════════════════════\n"
                "✓ Excel 365/2021: двойной клик / double-click\n"
                "✓ Excel 2019/2016: F2 + Enter\n"
                "✓ Ручной ввод / Manual: ДД.ММ.ГГГГ / DD.MM.YYYY\n"
                "✓ Быстрый ввод / Quick: Ctrl+; (текущая дата/current date)\n"
                "\n"
                "📋 ПРИМЕРЫ / EXAMPLES:\n"
                "• 28.06.2025 • 01.07.2025 • 15.12.2025", 
                "OEE Monitoring System v7.0"
            )
            
            # Начальные значения
            ws.cell(row=row, column=4).value = 480  # 8-часовая смена (единственное значение по умолчанию)
            
            # КЛАССИЧЕСКИЕ ФОРМУЛЫ OEE
            
            # Availability = (Shift_Time - Downtime - Setup_Time) / Shift_Time
            ws.cell(row=row, column=9).value = f'=IF(D{row}>0,(D{row}-F{row}-E{row})/D{row},0)'
            ws.cell(row=row, column=9).number_format = '0.0%'
            
            # Performance = Actual_Output / Planned_Output (будет рассчитываться через скорость)
            # Пока упрощенно: если нет простоев и наладки, то 100%
            ws.cell(row=row, column=10).value = f'=IF(AND(H{row}>0,D{row}>E{row}+F{row}),MIN(H{row}*60/(D{row}-E{row}-F{row}),1),0)'
            ws.cell(row=row, column=10).number_format = '0.0%'
            
            # Quality = Good_Parts / (Good_Parts + Defects)
            ws.cell(row=row, column=11).value = f'=IF(H{row}+G{row}>0,H{row}/(H{row}+G{row}),0)'
            ws.cell(row=row, column=11).number_format = '0.0%'
            
            # OEE = Availability × Performance × Quality
            ws.cell(row=row, column=12).value = f'=I{row}*J{row}*K{row}'
            ws.cell(row=row, column=12).number_format = '0.0%'

        # Условное форматирование OEE
        ws.conditional_formatting.add('L3:L1002', CellIsRule(operator='greaterThanOrEqual', formula=['0.85'], fill=EXCELLENT_FILL))
        ws.conditional_formatting.add('L3:L1002', CellIsRule(operator='between', formula=['0.7', '0.84'], fill=GOOD_FILL))
        ws.conditional_formatting.add('L3:L1002', CellIsRule(operator='between', formula=['0.6', '0.69'], fill=AVERAGE_FILL))
        ws.conditional_formatting.add('L3:L1002', CellIsRule(operator='lessThan', formula=['0.6'], fill=POOR_FILL))

        # Настройка ширины столбцов
        for col in range(1, len(headers_station) + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            ws.column_dimensions[col_letter].width = 18

    # ============ Лист "Dashboard" ============
    print("🎨 Создание 'Dashboard'...")
    ws_dashboard.sheet_view.showGridLines = False
    ws_dashboard.merge_cells("A1:H2")
    ws_dashboard['A1'].value = "Панель управления производством / Production Management Dashboard"
    ws_dashboard['A1'].font = TITLE_FONT
    ws_dashboard['A1'].alignment = CENTER_ALIGN

    # Основные показатели
    ws_dashboard['A4'] = "Станков с OEE < 60% / Stations with OEE < 60%:"
    ws_dashboard['C4'] = f'=COUNTIF(MainData_Основные!B:B,"<0.6")'
    ws_dashboard['C4'].font = Font(bold=True, color=POOR_COLOR)

    ws_dashboard['A6'] = "Станков мирового класса (85%+) / World Class Stations:"
    ws_dashboard['C6'] = f'=COUNTIF(MainData_Основные!B:B,">=0.85")'
    ws_dashboard['C6'].font = Font(bold=True, color=EXCELLENT_COLOR)

    ws_dashboard['A8'] = "Средний OEE / Average OEE:"
    ws_dashboard['C8'] = f'=IFERROR(AVERAGE(MainData_Основные!B:B),0)'
    ws_dashboard['C8'].number_format = '0.0%'
    ws_dashboard['C8'].font = Font(bold=True, size=14)

    ws_dashboard['A10'] = "Общее количество годных деталей / Total Good Parts:"
    ws_dashboard['C10'] = f'=SUM(MainData_Основные!F:F)'
    ws_dashboard['C10'].font = Font(bold=True, color=EXCELLENT_COLOR)

    ws_dashboard['A12'] = "Общий процент брака / Overall Defect Rate:"
    ws_dashboard['C12'] = f'=IFERROR(SUM(MainData_Основные!G:G)/(SUM(MainData_Основные!F:F)+SUM(MainData_Основные!G:G)),0)'
    ws_dashboard['C12'].number_format = '0.0%'

    ws_dashboard['A14'] = "Лучший станок / Best Station:"
    ws_dashboard['C14'] = f'=INDEX(MainData_Основные!A:A,MATCH(MAX(MainData_Основные!B:B),MainData_Основные!B:B,0))'

    ws_dashboard['A16'] = "Лучший оператор / Best Operator:"
    ws_dashboard['C16'] = f'=INDEX(OperatorSummary_Операторы!A:A,MATCH(MAX(OperatorSummary_Операторы!C:C),OperatorSummary_Операторы!C:C,0))'

    # График OEE по станкам
    chart_oee = BarChart()
    chart_oee.title = "OEE по станкам / Equipment OEE"
    chart_oee.x_axis.title = "Станки / Equipment"
    chart_oee.y_axis.title = "OEE (%)"
    data_oee = Reference(ws_main_data, min_col=2, min_row=2, max_row=8)
    categories_oee = Reference(ws_main_data, min_col=1, min_row=2, max_row=8)
    chart_oee.add_data(data_oee, titles_from_data=False)
    chart_oee.set_categories(categories_oee)
    chart_oee.height = 12
    chart_oee.width = 20
    ws_dashboard.add_chart(chart_oee, "A18")

    # График по операторам
    chart_operators = BarChart()
    chart_operators.title = "OEE операторов / Operator OEE"
    chart_operators.x_axis.title = "Операторы / Operators"
    chart_operators.y_axis.title = "OEE (%)"
    data_operators = Reference(ws_operator_summary, min_col=3, min_row=2, max_row=7)
    categories_operators = Reference(ws_operator_summary, min_col=1, min_row=2, max_row=7)
    chart_operators.add_data(data_operators, titles_from_data=False)
    chart_operators.set_categories(categories_operators)
    chart_operators.height = 12
    chart_operators.width = 20
    ws_dashboard.add_chart(chart_operators, "A35")

    # ============ Лист "Analytics" ============
    print("📈 Заполнение 'Analytics'...")
    ws_analytics['A1'] = "Аналитика OEE / OEE Analytics"
    ws_analytics['A1'].font = TITLE_FONT
    
    analytics_headers = ["Метрика / Metric", "Значение / Value", "Бенчмарк / Benchmark", "Статус / Status"]
    ws_analytics.append(analytics_headers)
    
    analytics_data = [
        ["Средний OEE / Average OEE", f'=IFERROR(AVERAGE(MainData_Основные!B:B),0)', "70%", f'=IF(B3>=0.7,"✅ Хорошо","⚠️ Ниже нормы")'],
        ["Стабильность OEE / OEE Stability", f'=IFERROR(STDEV(MainData_Основные!B:B),0)', "<5%", f'=IF(B4<=0.05,"✅ Стабильно","⚠️ Нестабильно")'],
        ["Средняя доступность / Avg Availability", f'=IFERROR(AVERAGE(MainData_Основные!C:C),0)', "90%", f'=IF(B5>=0.9,"✅ Отлично","⚠️ Ниже нормы")'],
        ["Средняя производительность / Avg Performance", f'=IFERROR(AVERAGE(MainData_Основные!D:D),0)', "95%", f'=IF(B6>=0.95,"✅ Отлично","⚠️ Ниже нормы")'],
        ["Среднее качество / Avg Quality", f'=IFERROR(AVERAGE(MainData_Основные!E:E),0)', "99%", f'=IF(B7>=0.99,"✅ Отлично","⚠️ Ниже нормы")'],
        ["Процент брака / Defect Rate", f'=IFERROR(SUM(MainData_Основные!G:G)/(SUM(MainData_Основные!F:F)+SUM(MainData_Основные!G:G)),0)', "<2%", f'=IF(B8<=0.02,"✅ Хорошо","⚠️ Высокий")'],
        ["Среднее время наладки / Avg Setup Time", f'=IFERROR(AVERAGE(MainData_Основные!H:H),0)', "<60 мин", f'=IF(B9<=60,"✅ Быстро","⚠️ Медленно")']
    ]
    
    for row_data in analytics_data:
        ws_analytics.append(row_data)
    
    # Форматирование Analytics
    for col_idx in range(1, 5):
        ws_analytics.cell(row=2, column=col_idx).font = BILINGUAL_FONT
        ws_analytics.cell(row=2, column=col_idx).fill = HEADER_FILL
        ws_analytics.cell(row=2, column=col_idx).alignment = CENTER_ALIGN
    
    for row_idx in range(3, 10):
        for col_idx in range(1, 5):
            ws_analytics.cell(row=row_idx, column=col_idx).border = THIN_BORDER
            ws_analytics.cell(row=row_idx, column=col_idx).fill = TABLE_BG_FILL
        # Форматирование процентов для нужных ячеек
        if row_idx in [3, 4, 5, 6, 7, 8]:
            ws_analytics.cell(row=row_idx, column=2).number_format = '0.0%'
    
    ws_analytics.column_dimensions['A'].width = 35
    ws_analytics.column_dimensions['B'].width = 20
    ws_analytics.column_dimensions['C'].width = 15
    ws_analytics.column_dimensions['D'].width = 20

    # ============ Лист "Trends" ============
    print("📈 Создание 'Trends'...")
    ws_trends['A1'] = "Тренды OEE / OEE Trends"
    ws_trends['A1'].font = TITLE_FONT
    
    # Заголовки для трендов
    trend_headers = ["Date / Дата", "Overall_OEE / Общий_OEE", "Best_Station / Лучший_станок", "Worst_Station / Худший_станок"]
    ws_trends.append(trend_headers)
    
    # Форматирование заголовков
    for col_idx in range(1, 5):
        ws_trends.cell(row=2, column=col_idx).font = BILINGUAL_FONT
        ws_trends.cell(row=2, column=col_idx).fill = HEADER_FILL
        ws_trends.cell(row=2, column=col_idx).alignment = CENTER_ALIGN
    
    # Пустые строки для ручного заполнения (без моковых данных)
    for row in range(3, 33):  # 30 пустых строк
        for col in range(1, 5):
            ws_trends.cell(row=row, column=col).border = THIN_BORDER
            ws_trends.cell(row=row, column=col).fill = TABLE_BG_FILL
        # Форматирование для будущих процентов
        for col in range(2, 5):
            ws_trends.cell(row=row, column=col).number_format = '0.0%'
    
    # График трендов (будет работать после заполнения данных)
    chart_trends = LineChart()
    chart_trends.title = "Тренды OEE / OEE Trends"
    chart_trends.x_axis.title = "Дата / Date"
    chart_trends.y_axis.title = "OEE (%)"
    
    # График будет пустым до заполнения данных
    data_trends = Reference(ws_trends, min_col=2, min_row=2, max_col=4, max_row=10)  # Меньший диапазон
    categories_trends = Reference(ws_trends, min_col=1, min_row=3, max_row=10)
    chart_trends.add_data(data_trends, titles_from_data=True)
    chart_trends.set_categories(categories_trends)
    chart_trends.height = 15
    chart_trends.width = 25
    ws_trends.add_chart(chart_trends, "F2")
    
    # Подсказка для пользователя
    ws_trends['F20'] = "Подсказка: Заполните данные слева для отображения графика"
    ws_trends['F21'] = "Tip: Fill data on the left to display the chart"
    ws_trends['F20'].font = Font(italic=True)
    ws_trends['F21'].font = Font(italic=True)

    # --- 4. Сохранение файла ---
    filename = f"OEE_Monitoring_Simple_v7.0_{datetime(2025, 6, 28, 23, 15).strftime('%Y%m%d_%H%M')}.xlsx"
    filepath = os.path.join(os.getcwd(), filename)
    try:
        wb.save(filepath)
        print(f"✅ УСПЕХ! Файл создан: {filename}")
        print("\n🎯 УПРОЩЕННАЯ СИСТЕМА ГОТОВА:")
        print("   ✓ Классическая формула OEE")
        print("   ✓ Наладка учтена в доступности")
        print("   ✓ Простая структура данных")
        print("   ✓ Двуязычный интерфейс")
        print("   ✓ Цветовая индикация по бенчмаркам")
        print("   ✓ Связанная аналитика")
        print("\n📊 БЕНЧМАРКИ OEE:")
        print("   🏆 85%+ - Мировой класс")
        print("   ✅ 70-84% - Хороший уровень")
        print("   ⚠️ 60-69% - Средний уровень")
        print("   ❌ <60% - Требует улучшения")
        print("\n🔧 ФОРМУЛА OEE:")
        print("   OEE = Доступность × Производительность × Качество")
        print("   Доступность = (Время смены - Простои - Наладка) / Время смены")
        return filepath
    except Exception as e:
        print(f"❌ Ошибка при сохранении: {e}")
        return None

if __name__ == "__main__":
    create_monitoring_system()
