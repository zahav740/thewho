#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Быстрая генерация Excel файла OEE/KPI на русском языке
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import date
import os

def quick_excel():
    """Быстрое создание Excel файла с русскими названиями"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OEE_KPI_Данные"
    
    # Заголовки на русском
    headers = ["Дата", "Станок", "Оператор", "Смена", "Наладка", "Производство", 
              "Простои", "План_шт", "Факт_шт", "Брак", "Годные", "Доступность_%", 
              "Производительность_%", "Качество_%", "OEE_%", "Доля_наладки_%", 
              "Качество_наладки_%", "KPI_%"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Формулы
    formulas = {
        'K2': '=I2-J2',
        'L2': '=IF(D2>0,(D2-G2)/D2*100,0)',
        'M2': '=IF(H2>0,I2/H2*100,0)',
        'N2': '=IF(I2>0,K2/I2*100,0)',
        'O2': '=L2*M2*N2/10000',
        'P2': '=IF(D2>0,E2/D2*100,0)',
        'Q2': '=IF(I2>0,(I2-J2)/I2*100,0)',
        'R2': '=O2*0.5+(100-P2)*0.2+Q2*0.15+90*0.15'
    }
    
    for cell_ref, formula in formulas.items():
        ws[cell_ref] = formula
    
    # Тестовые данные
    ws['A2'] = date.today()
    ws['B2'] = "CNC-1"
    ws['C2'] = "Профи-фрезеровщик"
    ws['D2'] = 480
    ws['E2'] = 120
    ws['F2'] = 300
    ws['G2'] = 60
    ws['H2'] = 15
    ws['I2'] = 12
    ws['J2'] = 1
    
    # Условное форматирование
    from openpyxl.formatting.rule import CellIsRule
    
    # OEE форматирование
    green_rule = CellIsRule(operator='greaterThan', formula=['80'], fill=PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'))
    yellow_rule = CellIsRule(operator='between', formula=['70', '79'], fill=PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid'))
    red_rule = CellIsRule(operator='lessThan', formula=['70'], fill=PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid'))
    
    ws.conditional_formatting.add("O2:O100", green_rule)
    ws.conditional_formatting.add("O2:O100", yellow_rule)
    ws.conditional_formatting.add("O2:O100", red_rule)
    
    # KPI форматирование
    kpi_green = CellIsRule(operator='greaterThan', formula=['85'], fill=PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'))
    kpi_yellow = CellIsRule(operator='between', formula=['75', '84'], fill=PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid'))
    kpi_red = CellIsRule(operator='lessThan', formula=['75'], fill=PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid'))
    
    ws.conditional_formatting.add("R2:R100", kpi_green)
    ws.conditional_formatting.add("R2:R100", kpi_yellow)
    ws.conditional_formatting.add("R2:R100", kpi_red)
    
    # Автоширина столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    filename = f"OEE_KPI_Быстрый_{date.today().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"✅ Файл создан: {filename}")
    print("📊 Тестовые данные:")
    print("   - OEE должен быть ~64%")
    print("   - Доля наладки: 25%")
    print("   - KPI: ~79%")
    return filename

if __name__ == "__main__":
    quick_excel()
