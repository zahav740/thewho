#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ИСПРАВЛЕННЫЙ СКРИПТ OEE/KPI v12.0 - ПРАВИЛЬНЫЕ ФОРМУЛЫ
Исправления согласно требованиям:
1. OEE станка = (Время наладки + Время производства) / Общее время смены * 100%
2. KPI оператора НЕ штрафуется за сложность наладки  
3. Наладка включает: наладку + ОТК + поправки на ошибки
4. Наладка = полезное время станка (НЕ простой)
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from datetime import datetime, date
import os

def create_corrected_oee_excel():
    """
    Создает Excel файл с ИСПРАВЛЕННЫМИ формулами OEE и KPI
    """
    print("🔧 Создаем Excel файл с ИСПРАВЛЕННЫМИ формулами OEE/KPI...")
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Цветовая палитра
    colors = {
        'header': '3B82F6',      # Синий заголовков
        'success': '10B981',     # Зеленый успех  
        'warning': 'F59E0B',     # Желтый предупреждение
        'danger': 'EF4444',      # Красный опасность
        'info': '06B6D4',        # Голубой информация
        'correct': '10B981',     # Зеленый для правильных формул
        'wrong': 'EF4444'        # Красный для неправильных
    }
    
    # Стили
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True, color="FFFFFF", size=11)
    header_style.fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type="solid")
    header_style.alignment = Alignment(horizontal="center", vertical="center")
    wb.add_named_style(header_style)
    
    # ============ ЛИСТ СРАВНЕНИЯ ФОРМУЛ ============
    print("📋 Создаем лист сравнения формул...")
    ws_compare = wb.create_sheet("Сравнение_формул", 0)
    
    # Заголовок
    ws_compare['A1'] = "🔧 СРАВНЕНИЕ СТАРЫХ И НОВЫХ ФОРМУЛ"
    ws_compare['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws_compare['A1'].fill = PatternFill(start_color=colors['info'], end_color=colors['info'], fill_type="solid")
    ws_compare.merge_cells('A1:F1')
    ws_compare['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_compare.row_dimensions[1].height = 35
    
    # Заголовки сравнительной таблицы
    comparison_headers = ["Метрика", "❌ Старая формула", "✅ Новая формула", "Проблема старой", "Преимущество новой", "Пример Кирилла"]
    
    for col, header in enumerate(comparison_headers, 1):
        cell = ws_compare.cell(row=2, column=col, value=header)
        cell.style = "header_style"
    
    # Сравнительные данные
    comparison_data = [
        [
            "OEE станка",
            "=(Смена-Простои)/Смена*100",
            "=(Наладка+Производство)/Смена*100",
            "Штрафует за наладку как за простой",
            "Наладка = полезное время станка",
            "❌ 87.5% → ✅ 87.5% (но правильная логика!)"
        ],
        [
            "KPI оператора",
            "=OEE*50%+(100-Доля_наладки)*20%+...",
            "=Эффективность*60%+Качество*30%+Нормы*10%",
            "Штрафует за сложность наладки",
            "Оценивает только работу оператора",
            "❌ 79% → ✅ 89% (справедливая оценка!)"
        ],
        [
            "Доступность",
            "=(Смена-Простои)/Смена*100",
            "Не используется в новой схеме",
            "Наладка считается простоем",
            "Заменена на загрузку станка",
            "Устранена неправильная логика"
        ],
        [
            "Время наладки",
            "Вычитается из полезного времени",
            "Включается в полезное время",
            "Несправедливо к операторам",
            "Наладка + ОТК + поправки = работа",
            "120 мин = полезное время, а не простой"
        ]
    ]
    
    for row_idx, row_data in enumerate(comparison_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_compare.cell(row=row_idx, column=col_idx, value=value)
            
            # Цветовое кодирование
            if col_idx == 2:  # Старые формулы
                cell.fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type="solid")
                cell.font = Font(color=colors['wrong'])
            elif col_idx == 3:  # Новые формулы
                cell.fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type="solid")
                cell.font = Font(color=colors['correct'], bold=True)
            elif col_idx == 6:  # Примеры
                cell.font = Font(bold=True)
    
    # Настройка ширины столбцов
    ws_compare.column_dimensions['A'].width = 15
    ws_compare.column_dimensions['B'].width = 30
    ws_compare.column_dimensions['C'].width = 35
    ws_compare.column_dimensions['D'].width = 25
    ws_compare.column_dimensions['E'].width = 25
    ws_compare.column_dimensions['F'].width = 30
    
    # ============ ОСНОВНЫЕ ДАННЫЕ С ПРАВИЛЬНЫМИ ФОРМУЛАМИ ============
    print("📊 Создаем основной лист данных...")
    ws_data = wb.create_sheet("Данные_исправленные")
    
    # Заголовок
    ws_data['A1'] = "🏭 ПРОИЗВОДСТВЕННЫЕ ДАННЫЕ - ИСПРАВЛЕННЫЕ ФОРМУЛЫ"
    ws_data['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws_data['A1'].fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type="solid")
    ws_data.merge_cells('A1:S1')
    ws_data['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_data.row_dimensions[1].height = 35
    
    # Заголовки столбцов
    headers = [
        "Дата", "Станок", "Оператор", "Смена_мин", "Наладка_мин", "Производство_мин", 
        "Простои_мин", "План_шт", "Факт_шт", "Брак_шт", "Годные_шт", 
        "OEE_станка_%", "Эффективность_производства_%", "Качество_%", "Соблюдение_норм_%",
        "KPI_оператора_%", "Загрузка_станка_%", "Разбивка_наладка_%", "Разбивка_простои_%"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws_data.cell(row=2, column=col, value=header)
        cell.style = "header_style"
    
    # ПРАВИЛЬНЫЕ ФОРМУЛЫ
    corrected_formulas = {
        # Годные детали
        'K3': '=I3-J3',
        
        # ✅ ПРАВИЛЬНЫЙ OEE станка = (Наладка + Производство) / Смена * 100%
        'L3': '=IF(D3>0,(E3+F3)/D3*100,0)',
        
        # ✅ Эффективность производства (только в рабочее время)
        'M3': '=IF(F3>0,MIN(100,(I3*25)/F3*100),0)',  # 25 мин = норма на деталь
        
        # ✅ Качество (процент годных)
        'N3': '=IF(I3>0,K3/I3*100,0)',
        
        # ✅ Соблюдение норм времени
        'O3': '=IF(I3>0,MIN(100,25/(F3/I3)*100),0)',  # норма 25 мин vs факт
        
        # ✅ ПРАВИЛЬНЫЙ KPI оператора (БЕЗ штрафа за наладку!)
        'P3': '=M3*0.6+N3*0.3+O3*0.1',
        
        # Загрузка станка (то же что OEE)
        'Q3': '=L3',
        
        # Разбивка времени - наладка %
        'R3': '=IF(D3>0,E3/D3*100,0)',
        
        # Разбивка времени - простои %
        'S3': '=IF(D3>0,G3/D3*100,0)'
    }
    
    # Вставляем формулы
    for cell_ref, formula in corrected_formulas.items():
        ws_data[cell_ref] = formula
        ws_data[cell_ref].font = Font(color=colors['correct'], bold=True)
    
    # Копируем формулы на 50 строк
    for row in range(4, 51):
        for col_letter, base_formula in [
            ('K', '=I{}-J{}'),
            ('L', '=IF(D{}>0,(E{}+F{})/D{}*100,0)'),
            ('M', '=IF(F{}>0,MIN(100,(I{}*25)/F{}*100),0)'),
            ('N', '=IF(I{}>0,K{}/I{}*100,0)'),
            ('O', '=IF(I{}>0,MIN(100,25/(F{}/I{})*100),0)'),
            ('P', '=M{}*0.6+N{}*0.3+O{}*0.1'),
            ('Q', '=L{}'),
            ('R', '=IF(D{}>0,E{}/D{}*100,0)'),
            ('S', '=IF(D{}>0,G{}/D{}*100,0)')
        ]:
            formula = base_formula.format(*[row] * base_formula.count('{}'))
            ws_data[f'{col_letter}{row}'] = formula
    
    # Тестовые данные
    test_data = [
        # Дата, Станок, Оператор, Смена, Наладка, Производство, Простои, План, Факт, Брак
        [date.today(), "Doosan Yashana", "Kirill", 480, 120, 300, 60, 15, 12, 1],
        [date.today(), "Doosan Hadasha", "Arkady", 480, 80, 350, 50, 18, 20, 0],
        [date.today(), "Mitsubishi", "Denis", 480, 150, 200, 130, 10, 8, 2],
        [date.today(), "Pinnacle", "Daniel", 480, 100, 320, 60, 16, 15, 1]
    ]
    
    for row_idx, row_data in enumerate(test_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            ws_data.cell(row=row_idx, column=col_idx, value=value)
    
    # Условное форматирование для OEE станка
    oee_range = "L3:L50"
    oee_excellent = CellIsRule(operator='greaterThanOrEqual', formula=['85'], 
                              fill=PatternFill(start_color=colors['success'], end_color=colors['success'], fill_type='solid'))
    oee_good = CellIsRule(operator='between', formula=['75', '84'], 
                         fill=PatternFill(start_color=colors['warning'], end_color=colors['warning'], fill_type='solid'))
    oee_poor = CellIsRule(operator='lessThan', formula=['75'], 
                         fill=PatternFill(start_color=colors['danger'], end_color=colors['danger'], fill_type='solid'))
    
    ws_data.conditional_formatting.add(oee_range, oee_excellent)
    ws_data.conditional_formatting.add(oee_range, oee_good)
    ws_data.conditional_formatting.add(oee_range, oee_poor)
    
    # Условное форматирование для KPI оператора
    kpi_range = "P3:P50"
    kpi_excellent = CellIsRule(operator='greaterThanOrEqual', formula=['90'], 
                              fill=PatternFill(start_color=colors['success'], end_color=colors['success'], fill_type='solid'))
    kpi_good = CellIsRule(operator='between', formula=['80', '89'], 
                         fill=PatternFill(start_color=colors['warning'], end_color=colors['warning'], fill_type='solid'))
    kpi_poor = CellIsRule(operator='lessThan', formula=['80'], 
                         fill=PatternFill(start_color=colors['danger'], end_color=colors['danger'], fill_type='solid'))
    
    ws_data.conditional_formatting.add(kpi_range, kpi_excellent)
    ws_data.conditional_formatting.add(kpi_range, kpi_good)
    ws_data.conditional_formatting.add(kpi_range, kpi_poor)
    
    # Настройка ширины столбцов
    column_widths = {
        'A': 12, 'B': 14, 'C': 12, 'D': 10, 'E': 12, 'F': 15, 'G': 10, 'H': 8, 'I': 8, 'J': 8, 'K': 10,
        'L': 12, 'M': 20, 'N': 10, 'O': 15, 'P': 15, 'Q': 15, 'R': 15, 'S': 15
    }
    
    for col, width in column_widths.items():
        ws_data.column_dimensions[col].width = width
    
    # ============ ПРИМЕРЫ РАСЧЕТОВ ============
    print("📋 Создаем лист с примерами...")
    ws_examples = wb.create_sheet("Примеры_расчетов")
    
    # Заголовок
    ws_examples['A1'] = "📋 ПРИМЕРЫ РАСЧЕТОВ ПО НОВОЙ ЛОГИКЕ"
    ws_examples['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws_examples['A1'].fill = PatternFill(start_color=colors['info'], end_color=colors['info'], fill_type="solid")
    ws_examples.merge_cells('A1:G1')
    ws_examples['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_examples.row_dimensions[1].height = 35
    
    # Пример для Кирилла
    example_data = [
        ("", "ПРИМЕР: Смена Кирилла", "", "", "", "", ""),
        ("Параметр", "Значение", "Единица", "Пояснение", "", "", ""),
        ("Общее время смены", "480", "минут", "8 часов", "", "", ""),
        ("Время наладки", "120", "минут", "Сложная наладка + ОТК + поправки", "", "", ""),
        ("Время производства", "300", "минут", "Чистое производство деталей", "", "", ""),
        ("Простои", "60", "минут", "Поломки, ожидание", "", "", ""),
        ("План деталей", "15", "штук", "Плановое задание", "", "", ""),
        ("Факт произведено", "12", "штук", "Реально сделано", "", "", ""),
        ("Брак", "1", "штук", "Бракованные детали", "", "", ""),
        ("", "", "", "", "", "", ""),
        ("РАСЧЕТЫ ПО НОВОЙ ЛОГИКЕ:", "", "", "", "", "", ""),
        ("", "", "", "", "", "", ""),
        ("OEE станка", "=(120+300)/480*100", "87.5%", "Загруженность станка", "", "", ""),
        ("Эффективность произв.", "=(12*25)/300*100", "100%", "Норма выполнена", "", "", ""),
        ("Качество", "=(12-1)/12*100", "91.7%", "Процент годных", "", "", ""),
        ("Соблюдение норм", "=25/(300/12)*100", "100%", "Время в норме", "", "", ""),
        ("KPI оператора", "=100*0.6+91.7*0.3+100*0.1", "97.5%", "БЕЗ штрафа за наладку!", "", "", ""),
        ("", "", "", "", "", "", ""),
        ("СРАВНЕНИЕ СО СТАРОЙ ЛОГИКОЙ:", "", "", "", "", "", ""),
        ("Старый KPI", "~79%", "", "Штрафовал за наладку", "", "", ""),
        ("Новый KPI", "97.5%", "", "Справедливая оценка", "", "", ""),
        ("Разница", "+18.5%", "", "Значительное улучшение!", "", "", "")
    ]
    
    for row_idx, row_data in enumerate(example_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            if value:
                cell = ws_examples.cell(row=row_idx, column=col_idx, value=value)
                
                # Форматирование
                if "ПРИМЕР:" in str(value) or "РАСЧЕТЫ:" in str(value) or "СРАВНЕНИЕ:" in str(value):
                    cell.font = Font(bold=True, size=14, color=colors['info'])
                    cell.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type="solid")
                elif row_idx == 4:  # Заголовки
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type="solid")
                elif "97.5%" in str(value) or "+18.5%" in str(value):
                    cell.font = Font(bold=True, color=colors['success'])
                elif "БЕЗ штрафа" in str(value) or "Справедливая" in str(value):
                    cell.font = Font(bold=True, color=colors['success'])
    
    # Настройка ширины столбцов
    ws_examples.column_dimensions['A'].width = 20
    ws_examples.column_dimensions['B'].width = 20
    ws_examples.column_dimensions['C'].width = 10
    ws_examples.column_dimensions['D'].width = 35
    
    # ============ СОХРАНЕНИЕ ФАЙЛА ============
    filename = f"OEE_KPI_ИСПРАВЛЕННЫЕ_ФОРМУЛЫ_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(os.getcwd(), filename)
    
    try:
        wb.save(filepath)
        print(f"✅ Excel файл с исправленными формулами создан: {filename}")
        print(f"📁 Путь к файлу: {filepath}")
        print(f"📊 Размер файла: {os.path.getsize(filepath) / 1024:.1f} KB")
        
        print("\n🔧 КЛЮЧЕВЫЕ ИСПРАВЛЕНИЯ:")
        print("   ✅ OEE = (Наладка + Производство) / Смена * 100%")
        print("   ✅ KPI оператора БЕЗ штрафа за сложность наладки")
        print("   ✅ Наладка включает: наладку + ОТК + поправки")
        print("   ✅ Наладка = полезное время станка")
        
        print("\n📊 РЕЗУЛЬТАТЫ ДЛЯ КИРИЛЛА:")
        print("   ❌ Старый KPI: ~79% (несправедливо)")
        print("   ✅ Новый KPI: 97.5% (справедливо)")
        print("   📈 Улучшение: +18.5%")
        
        print("\n📋 СТРУКТУРА ФАЙЛА:")
        print("   🔸 Сравнение_формул - анализ старых vs новых формул")
        print("   🔸 Данные_исправленные - основная таблица с правильными расчетами")
        print("   🔸 Примеры_расчетов - детальный разбор расчетов для Кирилла")
        
        return filepath
        
    except Exception as e:
        print(f"❌ Ошибка при создании файла: {e}")
        return None

def main():
    """Основная функция"""
    print("=" * 80)
    print("🔧 ГЕНЕРАТОР EXCEL С ИСПРАВЛЕННЫМИ ФОРМУЛАМИ OEE/KPI")
    print("=" * 80)
    
    try:
        import openpyxl
        print("✅ Библиотека openpyxl доступна")
    except ImportError:
        print("❌ Ошибка: Установите openpyxl командой: pip install openpyxl")
        return
    
    filepath = create_corrected_oee_excel()
    
    if filepath:
        print(f"\n🎉 ФАЙЛ С ИСПРАВЛЕННЫМИ ФОРМУЛАМИ ГОТОВ!")
        print(f"📂 Откройте: {os.path.basename(filepath)}")
        print(f"\n🚀 ПРЕИМУЩЕСТВА НОВОЙ ЛОГИКИ:")
        print(f"   • OEE отражает РЕАЛЬНУЮ загрузку станка")
        print(f"   • KPI оператора справедливо оценивает работу")
        print(f"   • Наладка не штрафуется как простой")
        print(f"   • Учитывается сложность наладочных работ")
    else:
        print("\n❌ Не удалось создать файл")

if __name__ == "__main__":
    main()
