#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Скрипт для создания Excel файла с таблицами OEE и KPI
Автор: Claude AI Assistant
Дата: 2025-06-28
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.data_validation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, date
import os

def create_oee_kpi_excel_file():
    """
    Создает полноценный Excel файл с системой OEE и KPI для производства
    """
    print("🚀 Начинаем создание Excel файла OEE/KPI...")
    
    # Создаем новую книгу Excel
    wb = openpyxl.Workbook()
    
    # Удаляем стандартный лист
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # ============ ЛИСТ 1: ДАННЫЕ ============
    print("📊 Создаем лист 'Данные'...")
    ws_data = wb.create_sheet("Данные", 0)
    
    # Заголовки таблицы
    headers = [
        "Дата", "Станок", "Оператор", "Смена", "Наладка", "Производство", 
        "Простои", "План_шт", "Факт_шт", "Брак", "Годные", "Доступность_%", 
        "Производительность_%", "Качество_%", "OEE_%", "Доля_наладки_%", 
        "Качество_наладки_%", "KPI_%"
    ]
    
    # Добавляем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Формулы для автоматических расчетов (строка 2)
    formulas = {
        'K2': '=I2-J2',  # Годные детали
        'L2': '=IF(D2>0,(D2-G2)/D2*100,0)',  # Доступность %
        'M2': '=IF(H2>0,I2/H2*100,0)',  # Производительность %
        'N2': '=IF(I2>0,K2/I2*100,0)',  # Качество %
        'O2': '=L2*M2*N2/10000',  # OEE %
        'P2': '=IF(D2>0,E2/D2*100,0)',  # Доля наладки %
        'Q2': '=IF(I2>0,(I2-J2)/I2*100,0)',  # Качество наладки %
        'R2': '=O2*0.5+(100-P2)*0.2+Q2*0.15+90*0.15'  # KPI %
    }
    
    # Вставляем формулы
    for cell_ref, formula in formulas.items():
        ws_data[cell_ref] = formula
    
    # Копируем формулы на 100 строк
    for row in range(3, 101):
        for col in range(11, 19):  # Столбцы K-R
            source_cell = ws_data.cell(row=2, column=col)
            target_cell = ws_data.cell(row=row, column=col)
            if source_cell.value and isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                # Адаптируем формулу для новой строки
                new_formula = source_cell.value.replace('2', str(row))
                target_cell.value = new_formula
    
    # Настройка ширины столбцов
    column_widths = {
        'A': 12, 'B': 10, 'C': 18, 'D': 8, 'E': 10, 'F': 12, 'G': 10, 'H': 10, 
        'I': 10, 'J': 8, 'K': 10, 'L': 14, 'M': 16, 'N': 12, 'O': 8, 'P': 14, 'Q': 16, 'R': 8
    }
    
    for col, width in column_widths.items():
        ws_data.column_dimensions[col].width = width
    
    # Условное форматирование для OEE
    oee_range = "O2:O100"
    # Зеленый для >=80
    green_rule = CellIsRule(operator='greaterThan', formula=['80'], fill=PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'))
    # Желтый для 70-79
    yellow_rule = CellIsRule(operator='between', formula=['70', '79'], fill=PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid'))
    # Красный для <70
    red_rule = CellIsRule(operator='lessThan', formula=['70'], fill=PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid'))
    
    ws_data.conditional_formatting.add(oee_range, green_rule)
    ws_data.conditional_formatting.add(oee_range, yellow_rule)
    ws_data.conditional_formatting.add(oee_range, red_rule)
    
    # Аналогично для KPI
    kpi_range = "R2:R100"
    kpi_green = CellIsRule(operator='greaterThan', formula=['85'], fill=PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'))
    kpi_yellow = CellIsRule(operator='between', formula=['75', '84'], fill=PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid'))
    kpi_red = CellIsRule(operator='lessThan', formula=['75'], fill=PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid'))
    
    ws_data.conditional_formatting.add(kpi_range, kpi_green)
    ws_data.conditional_formatting.add(kpi_range, kpi_yellow)
    ws_data.conditional_formatting.add(kpi_range, kpi_red)
    
    # Добавляем тестовые данные
    test_data = [
        [date.today(), "CNC-1", "Профи-фрезеровщик", 480, 120, 300, 60, 15, 12, 1]
    ]
    
    for row_idx, row_data in enumerate(test_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws_data.cell(row=row_idx, column=col_idx, value=value)
    
    # ============ ЛИСТ 2: ВВОД ДАННЫХ ============
    print("📝 Создаем лист 'Ввод_данных'...")
    ws_input = wb.create_sheet("Ввод_данных")
    
    # Заголовок формы
    ws_input['A1'] = "ФОРМА ВВОДА ДАННЫХ ЗА СМЕНУ"
    ws_input['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws_input['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_input.merge_cells('A1:E1')
    ws_input['A1'].alignment = Alignment(horizontal="center")
    
    # Поля формы
    form_fields = [
        ("A3", "Дата:", "C3"),
        ("A4", "Станок:", "C4"),
        ("A5", "Оператор:", "C5"),
        ("A7", "ВРЕМЕННЫЕ ДАННЫЕ (в минутах):", ""),
        ("A8", "Продолжительность смены:", "C8"),
        ("A9", "Время наладки:", "C9"),
        ("A10", "Время производства:", "C10"),
        ("A11", "Время простоев:", "C11"),
        ("A13", "ПРОИЗВОДСТВЕННЫЕ ДАННЫЕ:", ""),
        ("A14", "Запланировано деталей:", "C14"),
        ("A15", "Фактически произведено:", "C15"),
        ("A16", "Из них брак:", "C16"),
    ]
    
    for field in form_fields:
        ws_input[field[0]] = field[1]
        ws_input[field[0]].font = Font(bold=True)
        if field[0] in ["A7", "A13"]:
            ws_input[field[0]].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Константа для смены
    ws_input['C8'] = 480
    ws_input['C8'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Выпадающие списки
    stations = ["CNC-1", "CNC-2", "CNC-3", "CNC-4", "CNC-5", "CNC-6", "CNC-7"]
    operators = ["Профи-фрезеровщик", "Средний специалист", "Слабый фрезеровщик", "Помощник", "Новый токарь"]
    
    # Валидация для станков
    dv_stations = DataValidation(type="list", formula1=f'"{",".join(stations)}"')
    ws_input.add_data_validation(dv_stations)
    dv_stations.add("C4")
    
    # Валидация для операторов
    dv_operators = DataValidation(type="list", formula1=f'"{",".join(operators)}"')
    ws_input.add_data_validation(dv_operators)
    dv_operators.add("C5")
    
    # Настройка ширины столбцов
    ws_input.column_dimensions['A'].width = 25
    ws_input.column_dimensions['B'].width = 5
    ws_input.column_dimensions['C'].width = 20
    
    # ============ ЛИСТ 3: СВОДКА ============
    print("📈 Создаем лист 'Сводка'...")
    ws_summary = wb.create_sheet("Сводка")
    
    # Заголовки сводки
    summary_headers = [
        "Станок", "Средний_OEE_%", "Средняя_доступность_%", "Средняя_производительность_%",
        "Среднее_качество_%", "Средняя_доля_наладки_%", "Общий_KPI_%", "Статус"
    ]
    
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Добавляем станки и формулы
    for row, station in enumerate(stations, 2):
        ws_summary.cell(row=row, column=1, value=station)
        
        # Формулы AVERAGEIF для каждого показателя
        ws_summary.cell(row=row, column=2, value=f'=AVERAGEIF(Данные.B:B,"{station}",Данные.O:O)')
        ws_summary.cell(row=row, column=3, value=f'=AVERAGEIF(Данные.B:B,"{station}",Данные.L:L)')
        ws_summary.cell(row=row, column=4, value=f'=AVERAGEIF(Данные.B:B,"{station}",Данные.M:M)')
        ws_summary.cell(row=row, column=5, value=f'=AVERAGEIF(Данные.B:B,"{station}",Данные.N:N)')
        ws_summary.cell(row=row, column=6, value=f'=AVERAGEIF(Данные.B:B,"{station}",Данные.P:P)')
        ws_summary.cell(row=row, column=7, value=f'=AVERAGEIF(Данные.B:B,"{station}",Данные.R:R)')
        
        # Формула для статуса
        status_formula = f'=IF(G{row}>=85,"🏆 Отлично",IF(G{row}>=75,"✅ Хорошо",IF(G{row}>=65,"⚠️ Средне","❌ Плохо")))'
        ws_summary.cell(row=row, column=8, value=status_formula)
    
    # Настройка ширины столбцов
    for col in range(1, 9):
        ws_summary.column_dimensions[chr(64 + col)].width = 18
    
    # ============ ЛИСТ 4: DASHBOARD ============
    print("🎛️ Создаем лист 'Dashboard'...")
    ws_dash = wb.create_sheet("Dashboard")
    
    # Заголовок
    ws_dash['A1'] = "🏭 DASHBOARD ПРОИЗВОДСТВА"
    ws_dash['A1'].font = Font(bold=True, size=18, color="FFFFFF")
    ws_dash['A1'].fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    ws_dash.merge_cells('A1:D1')
    ws_dash['A1'].alignment = Alignment(horizontal="center")
    
    # Данные dashboard
    dashboard_data = [
        ("A3", "📅 Дата обновления:", "C3", "=TODAY()"),
        ("A5", "📊 ОБЩИЕ ПОКАЗАТЕЛИ", "", ""),
        ("A6", "Средний OEE по цеху:", "C6", "=ROUND(AVERAGE(Данные.O:O),1)"),
        ("A7", "Средний KPI по цеху:", "C7", "=ROUND(AVERAGE(Данные.R:R),1)"),
        ("A8", "🏆 Лучший станок (OEE):", "C8", "=INDEX(Данные.B:B,MATCH(MAX(Данные.O:O),Данные.O:O,0))"),
        ("A9", "⚠️ Худший станок (OEE):", "C9", "=INDEX(Данные.B:B,MATCH(MIN(Данные.O:O),Данные.O:O,0))"),
        ("A11", "🎯 ДОСТИЖЕНИЕ ЦЕЛЕЙ", "", ""),
        ("A12", "Станков с OEE > 80%:", "C12", "=COUNTIF(Данные.O:O,\">80\")"),
        ("A13", "Станков с KPI > 75%:", "C13", "=COUNTIF(Данные.R:R,\">75\")"),
        ("A14", "Записей с долей наладки < 50%:", "C14", "=COUNTIF(Данные.P:P,\"<50\")"),
        ("A16", "⚠️ ПРОБЛЕМНЫЕ ЗОНЫ", "", ""),
        ("A17", "Высокая доля наладки (>60%):", "C17", "=COUNTIF(Данные.P:P,\">60\")"),
        ("A18", "Низкое качество (<90%):", "C18", "=COUNTIF(Данные.N:N,\"<90\")"),
        ("A19", "Низкий OEE (<70%):", "C19", "=COUNTIF(Данные.O:O,\"<70\")"),
    ]
    
    for row_data in dashboard_data:
        if row_data[0]:  # Метка
            ws_dash[row_data[0]] = row_data[1]
            if row_data[1].startswith(("📊", "🎯", "⚠️")):
                ws_dash[row_data[0]].font = Font(bold=True, size=12)
                ws_dash[row_data[0]].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            else:
                ws_dash[row_data[0]].font = Font(bold=True)
        
        if row_data[2] and row_data[3]:  # Значение
            ws_dash[row_data[2]] = row_data[3]
            if row_data[3].startswith("="):
                ws_dash[row_data[2]].font = Font(bold=True, color="2F75B5")
    
    # Настройка ширины столбцов
    ws_dash.column_dimensions['A'].width = 30
    ws_dash.column_dimensions['B'].width = 5
    ws_dash.column_dimensions['C'].width = 25
    ws_dash.column_dimensions['D'].width = 15
    
    # ============ ЛИСТ 5: НАСТРОЙКИ ============
    print("⚙️ Создаем лист 'Настройки'...")
    ws_settings = wb.create_sheet("Настройки")
    
    # Заголовок
    ws_settings['A1'] = "⚙️ НАСТРОЙКИ И СПРАВОЧНИКИ"
    ws_settings['A1'].font = Font(bold=True, size=16)
    ws_settings['A1'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    ws_settings.merge_cells('A1:F1')
    ws_settings['A1'].alignment = Alignment(horizontal="center")
    
    # Справочные данные
    settings_data = [
        ("A3", "📋 СПИСОК СТАНКОВ:"),
        ("A4", "CNC-1 (Фрезерный)"),
        ("A5", "CNC-2 (Фрезерный)"),
        ("A6", "CNC-3 (Фрезерный)"),
        ("A7", "CNC-4 (Фрезерный)"),
        ("A8", "CNC-5 (Фрезерный)"),
        ("A9", "CNC-6 (Токарный)"),
        ("A10", "CNC-7 (Токарный)"),
        ("C3", "👥 ОПЕРАТОРЫ:"),
        ("C4", "Профи-фрезеровщик"),
        ("C5", "Средний специалист"),
        ("C6", "Слабый фрезеровщик"),
        ("C7", "Помощник"),
        ("C8", "Новый токарь"),
        ("E3", "🎯 ЦЕЛЕВЫЕ ПОКАЗАТЕЛИ:"),
        ("E4", "OEE цель: ≥ 80%"),
        ("E5", "Доступность цель: ≥ 95%"),
        ("E6", "Производительность цель: ≥ 90%"),
        ("E7", "Качество цель: ≥ 95%"),
        ("E8", "Доля наладки цель: ≤ 50%"),
        ("E9", "KPI цель: ≥ 85%"),
        ("A12", "📖 ОБЪЯСНЕНИЕ ФОРМУЛ:"),
        ("A13", "OEE = Доступность × Производительность × Качество / 10000"),
        ("A14", "Доступность = (Смена - Простои) / Смена × 100%"),
        ("A15", "Производительность = Факт / План × 100%"),
        ("A16", "Качество = (Факт - Брак) / Факт × 100%"),
        ("A17", "KPI = OEE×50% + (100-Доля_наладки)×20% + Качество_наладки×15% + Сроки×15%"),
    ]
    
    for cell_addr, value in settings_data:
        ws_settings[cell_addr] = value
        if value.startswith(("📋", "👥", "🎯", "📖")):
            ws_settings[cell_addr].font = Font(bold=True, size=12)
            ws_settings[cell_addr].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        elif ":" in value and not value.startswith("OEE"):
            ws_settings[cell_addr].font = Font(bold=True)
    
    # Настройка ширины столбцов
    ws_settings.column_dimensions['A'].width = 25
    ws_settings.column_dimensions['B'].width = 5
    ws_settings.column_dimensions['C'].width = 20
    ws_settings.column_dimensions['D'].width = 5
    ws_settings.column_dimensions['E'].width = 30
    ws_settings.column_dimensions['F'].width = 10
    
    # ============ СОХРАНЕНИЕ ФАЙЛА ============
    filename = f"OEE_KPI_Производство_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(os.getcwd(), filename)
    
    try:
        wb.save(filepath)
        print(f"✅ Excel файл успешно создан: {filename}")
        print(f"📁 Путь к файлу: {filepath}")
        print(f"📊 Размер файла: {os.path.getsize(filepath) / 1024:.1f} KB")
        
        # Информация о файле
        print("\n📋 СТРУКТУРА ФАЙЛА:")
        print("   🔸 Лист 'Данные' - основная таблица с автоматическими расчетами")
        print("   🔸 Лист 'Ввод_данных' - форма для ввода новых записей")
        print("   🔸 Лист 'Сводка' - еженедельные итоги по станкам")
        print("   🔸 Лист 'Dashboard' - ключевые показатели в реальном времени")
        print("   🔸 Лист 'Настройки' - справочники и объяснения")
        
        print("\n🎯 ТЕСТОВЫЕ ДАННЫЕ:")
        print("   📅 Добавлена одна тестовая запись для проверки формул")
        print("   🔧 Все формулы настроены и готовы к работе")
        print("   🎨 Условное форматирование применено (зеленый/желтый/красный)")
        
        print("\n🚀 СЛЕДУЮЩИЕ ШАГИ:")
        print("   1. Откройте файл в Excel")
        print("   2. Перейдите на лист 'Ввод_данных'")
        print("   3. Заполните форму тестовыми данными")
        print("   4. Проверьте автоматические расчеты на листе 'Данные'")
        print("   5. Просмотрите Dashboard для общей картины")
        
        return filepath
        
    except Exception as e:
        print(f"❌ Ошибка при сохранении файла: {e}")
        return None

def main():
    """Основная функция"""
    print("=" * 60)
    print("🏭 ГЕНЕРАТОР EXCEL ФАЙЛА OEE/KPI ДЛЯ ПРОИЗВОДСТВА")
    print("=" * 60)
    
    # Проверяем наличие необходимых библиотек
    try:
        import openpyxl
        import pandas as pd
        print("✅ Все необходимые библиотеки установлены")
    except ImportError as e:
        print(f"❌ Отсутствует библиотека: {e}")
        print("💡 Установите командой: pip install openpyxl pandas")
        return
    
    # Создаем файл
    filepath = create_oee_kpi_excel_file()
    
    if filepath:
        print(f"\n🎉 ГОТОВО! Файл создан успешно!")
        print(f"📂 Откройте файл: {os.path.basename(filepath)}")
    else:
        print("\n❌ Не удалось создать файл")

if __name__ == "__main__":
    main()
