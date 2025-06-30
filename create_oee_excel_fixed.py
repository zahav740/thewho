#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Fixed Excel OEE/KPI Generator
Author: Claude AI Assistant
Date: 2025-06-28
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, date
import os

def create_oee_kpi_excel_file():
    """
    Creates complete Excel file with OEE and KPI system for production
    """
    print("🚀 Starting Excel OEE/KPI file creation...")
    
    # Create new Excel workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # ============ SHEET 1: DATA ============
    print("📊 Creating 'Data' sheet...")
    ws_data = wb.create_sheet("Data", 0)
    
    # Table headers
    headers = [
        "Date", "Station", "Operator", "Shift", "Setup", "Production", 
        "Downtime", "Plan_pcs", "Actual_pcs", "Defects", "Good_pcs", "Availability_%", 
        "Performance_%", "Quality_%", "OEE_%", "Setup_ratio_%", 
        "Setup_quality_%", "KPI_%"
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Formulas for automatic calculations (row 2)
    formulas = {
        'K2': '=I2-J2',  # Good pieces
        'L2': '=IF(D2>0,(D2-G2)/D2*100,0)',  # Availability %
        'M2': '=IF(H2>0,I2/H2*100,0)',  # Performance %
        'N2': '=IF(I2>0,K2/I2*100,0)',  # Quality %
        'O2': '=L2*M2*N2/10000',  # OEE %
        'P2': '=IF(D2>0,E2/D2*100,0)',  # Setup ratio %
        'Q2': '=IF(I2>0,(I2-J2)/I2*100,0)',  # Setup quality %
        'R2': '=O2*0.5+(100-P2)*0.2+Q2*0.15+90*0.15'  # KPI %
    }
    
    # Insert formulas
    for cell_ref, formula in formulas.items():
        ws_data[cell_ref] = formula
    
    # Copy formulas to 100 rows
    for row in range(3, 101):
        for col in range(11, 19):  # Columns K-R
            source_cell = ws_data.cell(row=2, column=col)
            target_cell = ws_data.cell(row=row, column=col)
            if source_cell.value and isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                # Adapt formula for new row
                new_formula = source_cell.value.replace('2', str(row))
                target_cell.value = new_formula
    
    # Set column widths
    column_widths = {
        'A': 12, 'B': 10, 'C': 18, 'D': 8, 'E': 10, 'F': 12, 'G': 10, 'H': 10, 
        'I': 10, 'J': 8, 'K': 10, 'L': 14, 'M': 16, 'N': 12, 'O': 8, 'P': 14, 'Q': 16, 'R': 8
    }
    
    for col, width in column_widths.items():
        ws_data.column_dimensions[col].width = width
    
    # Conditional formatting for OEE
    oee_range = "O2:O100"
    # Green for >=80
    green_rule = CellIsRule(operator='greaterThan', formula=['80'], fill=PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'))
    # Yellow for 70-79
    yellow_rule = CellIsRule(operator='between', formula=['70', '79'], fill=PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid'))
    # Red for <70
    red_rule = CellIsRule(operator='lessThan', formula=['70'], fill=PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid'))
    
    ws_data.conditional_formatting.add(oee_range, green_rule)
    ws_data.conditional_formatting.add(oee_range, yellow_rule)
    ws_data.conditional_formatting.add(oee_range, red_rule)
    
    # Same for KPI
    kpi_range = "R2:R100"
    kpi_green = CellIsRule(operator='greaterThan', formula=['85'], fill=PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'))
    kpi_yellow = CellIsRule(operator='between', formula=['75', '84'], fill=PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid'))
    kpi_red = CellIsRule(operator='lessThan', formula=['75'], fill=PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid'))
    
    ws_data.conditional_formatting.add(kpi_range, kpi_green)
    ws_data.conditional_formatting.add(kpi_range, kpi_yellow)
    ws_data.conditional_formatting.add(kpi_range, kpi_red)
    
    # Add test data
    test_data = [
        [date.today(), "CNC-1", "Expert-Miller", 480, 120, 300, 60, 15, 12, 1]
    ]
    
    for row_idx, row_data in enumerate(test_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws_data.cell(row=row_idx, column=col_idx, value=value)
    
    # ============ SHEET 2: INPUT FORM ============
    print("📝 Creating 'Input_Form' sheet...")
    ws_input = wb.create_sheet("Input_Form")
    
    # Form header
    ws_input['A1'] = "DATA INPUT FORM FOR SHIFT"
    ws_input['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws_input['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_input.merge_cells('A1:E1')
    ws_input['A1'].alignment = Alignment(horizontal="center")
    
    # Form fields
    form_fields = [
        ("A3", "Date:", "C3"),
        ("A4", "Station:", "C4"),
        ("A5", "Operator:", "C5"),
        ("A7", "TIME DATA (in minutes):", ""),
        ("A8", "Shift duration:", "C8"),
        ("A9", "Setup time:", "C9"),
        ("A10", "Production time:", "C10"),
        ("A11", "Downtime:", "C11"),
        ("A13", "PRODUCTION DATA:", ""),
        ("A14", "Planned pieces:", "C14"),
        ("A15", "Actual pieces:", "C15"),
        ("A16", "Defects:", "C16"),
    ]
    
    for field in form_fields:
        ws_input[field[0]] = field[1]
        ws_input[field[0]].font = Font(bold=True)
        if field[0] in ["A7", "A13"]:
            ws_input[field[0]].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Constant for shift
    ws_input['C8'] = 480
    ws_input['C8'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Set column widths
    ws_input.column_dimensions['A'].width = 25
    ws_input.column_dimensions['B'].width = 5
    ws_input.column_dimensions['C'].width = 20
    
    # ============ SHEET 3: SUMMARY ============
    print("📈 Creating 'Summary' sheet...")
    ws_summary = wb.create_sheet("Summary")
    
    # Summary headers
    summary_headers = [
        "Station", "Avg_OEE_%", "Avg_Availability_%", "Avg_Performance_%",
        "Avg_Quality_%", "Avg_Setup_ratio_%", "Overall_KPI_%", "Status"
    ]
    
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add stations and formulas
    stations = ["CNC-1", "CNC-2", "CNC-3", "CNC-4", "CNC-5", "CNC-6", "CNC-7"]
    for row, station in enumerate(stations, 2):
        ws_summary.cell(row=row, column=1, value=station)
        
        # AVERAGEIF formulas for each metric
        ws_summary.cell(row=row, column=2, value=f'=AVERAGEIF(Data.B:B,"{station}",Data.O:O)')
        ws_summary.cell(row=row, column=3, value=f'=AVERAGEIF(Data.B:B,"{station}",Data.L:L)')
        ws_summary.cell(row=row, column=4, value=f'=AVERAGEIF(Data.B:B,"{station}",Data.M:M)')
        ws_summary.cell(row=row, column=5, value=f'=AVERAGEIF(Data.B:B,"{station}",Data.N:N)')
        ws_summary.cell(row=row, column=6, value=f'=AVERAGEIF(Data.B:B,"{station}",Data.P:P)')
        ws_summary.cell(row=row, column=7, value=f'=AVERAGEIF(Data.B:B,"{station}",Data.R:R)')
        
        # Status formula
        status_formula = f'=IF(G{row}>=85,"🏆 Excellent",IF(G{row}>=75,"✅ Good",IF(G{row}>=65,"⚠️ Average","❌ Poor")))'
        ws_summary.cell(row=row, column=8, value=status_formula)
    
    # Set column widths
    for col in range(1, 9):
        ws_summary.column_dimensions[chr(64 + col)].width = 18
    
    # ============ SHEET 4: DASHBOARD ============
    print("🎛️ Creating 'Dashboard' sheet...")
    ws_dash = wb.create_sheet("Dashboard")
    
    # Header
    ws_dash['A1'] = "🏭 PRODUCTION DASHBOARD"
    ws_dash['A1'].font = Font(bold=True, size=18, color="FFFFFF")
    ws_dash['A1'].fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    ws_dash.merge_cells('A1:D1')
    ws_dash['A1'].alignment = Alignment(horizontal="center")
    
    # Dashboard data
    dashboard_data = [
        ("A3", "📅 Last updated:", "C3", "=TODAY()"),
        ("A5", "📊 OVERALL METRICS", "", ""),
        ("A6", "Average OEE across shop:", "C6", "=ROUND(AVERAGE(Data.O:O),1)"),
        ("A7", "Average KPI across shop:", "C7", "=ROUND(AVERAGE(Data.R:R),1)"),
        ("A8", "🏆 Best station (OEE):", "C8", "=INDEX(Data.B:B,MATCH(MAX(Data.O:O),Data.O:O,0))"),
        ("A9", "⚠️ Worst station (OEE):", "C9", "=INDEX(Data.B:B,MATCH(MIN(Data.O:O),Data.O:O,0))"),
        ("A11", "🎯 TARGET ACHIEVEMENT", "", ""),
        ("A12", "Stations with OEE > 80%:", "C12", "=COUNTIF(Data.O:O,\">80\")"),
        ("A13", "Stations with KPI > 75%:", "C13", "=COUNTIF(Data.R:R,\">75\")"),
        ("A14", "Records with setup < 50%:", "C14", "=COUNTIF(Data.P:P,\"<50\")"),
        ("A16", "⚠️ PROBLEM AREAS", "", ""),
        ("A17", "High setup ratio (>60%):", "C17", "=COUNTIF(Data.P:P,\">60\")"),
        ("A18", "Low quality (<90%):", "C18", "=COUNTIF(Data.N:N,\"<90\")"),
        ("A19", "Low OEE (<70%):", "C19", "=COUNTIF(Data.O:O,\"<70\")"),
    ]
    
    for row_data in dashboard_data:
        if row_data[0]:  # Label
            ws_dash[row_data[0]] = row_data[1]
            if row_data[1].startswith(("📊", "🎯", "⚠️")):
                ws_dash[row_data[0]].font = Font(bold=True, size=12)
                ws_dash[row_data[0]].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            else:
                ws_dash[row_data[0]].font = Font(bold=True)
        
        if row_data[2] and row_data[3]:  # Value
            ws_dash[row_data[2]] = row_data[3]
            if row_data[3].startswith("="):
                ws_dash[row_data[2]].font = Font(bold=True, color="2F75B5")
    
    # Set column widths
    ws_dash.column_dimensions['A'].width = 30
    ws_dash.column_dimensions['B'].width = 5
    ws_dash.column_dimensions['C'].width = 25
    ws_dash.column_dimensions['D'].width = 15
    
    # ============ SHEET 5: SETTINGS ============
    print("⚙️ Creating 'Settings' sheet...")
    ws_settings = wb.create_sheet("Settings")
    
    # Header
    ws_settings['A1'] = "⚙️ SETTINGS AND REFERENCES"
    ws_settings['A1'].font = Font(bold=True, size=16)
    ws_settings['A1'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    ws_settings.merge_cells('A1:F1')
    ws_settings['A1'].alignment = Alignment(horizontal="center")
    
    # Reference data
    settings_data = [
        ("A3", "📋 STATION LIST:"),
        ("A4", "CNC-1 (Milling)"),
        ("A5", "CNC-2 (Milling)"),
        ("A6", "CNC-3 (Milling)"),
        ("A7", "CNC-4 (Milling)"),
        ("A8", "CNC-5 (Milling)"),
        ("A9", "CNC-6 (Turning)"),
        ("A10", "CNC-7 (Turning)"),
        ("C3", "👥 OPERATORS:"),
        ("C4", "Expert-Miller"),
        ("C5", "Average-Specialist"),
        ("C6", "Junior-Miller"),
        ("C7", "Assistant"),
        ("C8", "New-Turner"),
        ("E3", "🎯 TARGET VALUES:"),
        ("E4", "OEE target: ≥ 80%"),
        ("E5", "Availability target: ≥ 95%"),
        ("E6", "Performance target: ≥ 90%"),
        ("E7", "Quality target: ≥ 95%"),
        ("E8", "Setup ratio target: ≤ 50%"),
        ("E9", "KPI target: ≥ 85%"),
        ("A12", "📖 FORMULA EXPLANATIONS:"),
        ("A13", "OEE = Availability × Performance × Quality / 10000"),
        ("A14", "Availability = (Shift - Downtime) / Shift × 100%"),
        ("A15", "Performance = Actual / Plan × 100%"),
        ("A16", "Quality = (Actual - Defects) / Actual × 100%"),
        ("A17", "KPI = OEE×50% + (100-Setup_ratio)×20% + Setup_quality×15% + Schedule×15%"),
    ]
    
    for cell_addr, value in settings_data:
        ws_settings[cell_addr] = value
        if value.startswith(("📋", "👥", "🎯", "📖")):
            ws_settings[cell_addr].font = Font(bold=True, size=12)
            ws_settings[cell_addr].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        elif ":" in value and not value.startswith("OEE"):
            ws_settings[cell_addr].font = Font(bold=True)
    
    # Set column widths
    ws_settings.column_dimensions['A'].width = 25
    ws_settings.column_dimensions['B'].width = 5
    ws_settings.column_dimensions['C'].width = 20
    ws_settings.column_dimensions['D'].width = 5
    ws_settings.column_dimensions['E'].width = 30
    ws_settings.column_dimensions['F'].width = 10
    
    # ============ SAVE FILE ============
    filename = f"OEE_KPI_Production_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(os.getcwd(), filename)
    
    try:
        wb.save(filepath)
        print(f"✅ Excel file successfully created: {filename}")
        print(f"📁 File path: {filepath}")
        print(f"📊 File size: {os.path.getsize(filepath) / 1024:.1f} KB")
        
        # File information
        print("\n📋 FILE STRUCTURE:")
        print("   🔸 Sheet 'Data' - main table with automatic calculations")
        print("   🔸 Sheet 'Input_Form' - form for entering new records")
        print("   🔸 Sheet 'Summary' - weekly totals by station")
        print("   🔸 Sheet 'Dashboard' - key metrics in real time")
        print("   🔸 Sheet 'Settings' - references and explanations")
        
        print("\n🎯 TEST DATA:")
        print("   📅 One test record added to verify formulas")
        print("   🔧 All formulas configured and ready to work")
        print("   🎨 Conditional formatting applied (green/yellow/red)")
        
        print("\n🚀 NEXT STEPS:")
        print("   1. Open file in Excel")
        print("   2. Go to 'Input_Form' sheet")
        print("   3. Fill form with test data")
        print("   4. Check automatic calculations on 'Data' sheet")
        print("   5. View Dashboard for overall picture")
        
        return filepath
        
    except Exception as e:
        print(f"❌ Error saving file: {e}")
        return None

def main():
    """Main function"""
    print("=" * 60)
    print("🏭 EXCEL OEE/KPI GENERATOR FOR PRODUCTION")
    print("=" * 60)
    
    # Check required libraries
    try:
        import openpyxl
        import pandas as pd
        print("✅ All required libraries installed")
    except ImportError as e:
        print(f"❌ Missing library: {e}")
        print("💡 Install with command: pip install openpyxl pandas")
        return
    
    # Create file
    filepath = create_oee_kpi_excel_file()
    
    if filepath:
        print(f"\n🎉 SUCCESS! File created successfully!")
        print(f"📂 Open file: {os.path.basename(filepath)}")
    else:
        print("\n❌ Failed to create file")

if __name__ == "__main__":
    main()
