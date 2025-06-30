import json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

print("Начало выполнения скрипта...")

# Загрузка JSON
try:
    with open('oee_kpi.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    print("JSON-файл успешно загружен")
except Exception as e:
    print(f"Ошибка при загрузке JSON: {e}")
    exit()

# Создание новой книги Excel
wb = openpyxl.Workbook()
wb.remove(wb.active)
print("Новая книга Excel создана")

# Функция для условного форматирования
def apply_conditional_formatting(ws, range_str, rules):
    for rule in rules:
        if rule['type'] == 'cell_value':
            if rule['operator'] == '>=':
                ws.conditional_formatting.add(range_str, 
                    openpyxl.formatting.rule.CellIsRule(operator='greaterThanOrEqual', 
                                                        formula=[str(rule['value'])], 
                                                        fill=PatternFill(start_color=rule['format']['fill_color'], end_color=rule['format']['fill_color'], fill_type='solid')))
            elif rule['operator'] == 'between':
                ws.conditional_formatting.add(range_str, 
                    openpyxl.formatting.rule.CellIsRule(operator='between', 
                                                        formula=[str(rule['value'][0]), str(rule['value'][1])], 
                                                        fill=PatternFill(start_color=rule['format']['fill_color'], end_color=rule['format']['fill_color'], fill_type='solid')))
            elif rule['operator'] == '<':
                ws.conditional_formatting.add(range_str, 
                    openpyxl.formatting.rule.CellIsRule(operator='lessThan', 
                                                        formula=[str(rule['value'])], 
                                                        fill=PatternFill(start_color=rule['format']['fill_color'], end_color=rule['format']['fill_color'], fill_type='solid')))

# Обработка листов
for sheet_data in data['workbook']['sheets']:
    ws = wb.create_sheet(title=sheet_data['name'])
    print(f"Создан лист: {sheet_data['name']}")
    
    if 'columns' in sheet_data:
        for i, col in enumerate(sheet_data['columns'], 1):
            ws.column_dimensions[get_column_letter(i)].width = col['width']
    
    if 'header_format' in sheet_data:
        for col_idx, col in enumerate(sheet_data['columns'], 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col['name']
            cell.font = Font(bold=sheet_data['header_format']['font_bold'], size=sheet_data['header_format']['font_size'])
            cell.alignment = Alignment(horizontal=sheet_data['header_format']['alignment'])
            cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    
    if 'content' in sheet_data:
        for item in sheet_data['content']:
            ws[item['cell']] = item['value']
    
    if 'formulas' in sheet_data:
        for formula in sheet_data['formulas']:
            ws[formula['cell']] = formula['value']
    
    if 'formula_copy' in sheet_data:
        for row in range(int(sheet_data['formula_copy']['copy_to'].split(':')[0][1:]), 
                        int(sheet_data['formula_copy']['copy_to'].split(':')[1][1:]) + 1):
            for col in range(ord(sheet_data['formula_copy']['range'].split(':')[0][0]) - ord('A') + 1, 
                            ord(sheet_data['formula_copy']['range'].split(':')[1][0]) - ord('A') + 2):
                cell = ws.cell(row=row, column=col)
                source_cell = ws.cell(row=int(sheet_data['formula_copy']['range'].split(':')[0][1:]), column=col)
                cell.value = source_cell.value
    
    if 'conditional_formatting' in sheet_data:
        for cf in sheet_data['conditional_formatting']:
            apply_conditional_formatting(ws, cf['range'], cf['rules'])
    
    if 'data_validation' in sheet_data:
        for dv in sheet_data['data_validation']:
            data_val = DataValidation(type='list', formula1=f'"{dv["source"]}"', allow_blank=True)
            data_val.add(dv['cell'])
            ws.add_data_validation(data_val)

# Сохранение файла
try:
    wb.save('OEE_KPI_Производство.xlsx')
    print("Excel-файл успешно создан: OEE_KPI_Производство.xlsx")
except Exception as e:
    print(f"Ошибка при сохранении файла: {e}")

print("Скрипт завершён.")