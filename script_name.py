#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
УЛУЧШЕННАЯ СИСТЕМА OEE/KPI v10.0
Исправления по требованиям пользователя:
1. Отдельные листы данных для каждого станка
2. KPI по каждому оператору в отдельной вкладке  
3. Производство как остаток (не фиксированное)
4. Выпадающие списки для операторов
5. Закрепленная шапка
6. Общий OEE/KPI по всему производству
7. График Гаугуса (Gantt)
8. Современный дизайн дашборда
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from datetime import datetime, date
import os

def create_advanced_oee_kpi():
    """
    Создает улучшенную систему OEE/KPI с отдельными листами для каждого станка
    """
    print("🚀 Создаем улучшенную систему OEE/KPI v10.0...")
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Обновленные списки
    stations = ["Doosan_Yashana", "Doosan_Hadasha", "Doosan_3", "Pinnacle_Gdola", "Mitsubishi", "JohnFord", "Okuma"]
    operators = ["Andrey", "Denis", "Daniel", "Kirill", "Slava", "Arkady"]
    
    # Современная цветовая схема
    colors = {
        'header': '1F2937',      # Темно-серый
        'primary': '3B82F6',     # Синий
        'success': '10B981',     # Зеленый
        'warning': 'F59E0B',     # Оранжевый
        'danger': 'EF4444',      # Красный
        'light': 'F3F4F6',      # Светло-серый
        'accent': '8B5CF6'       # Фиолетовый
    }
    
    # ============ ОБЩИЕ ДАННЫЕ ============
    print("📊 Создаем лист общих данных...")
    ws_main = wb.create_sheet("Общие_данные", 0)
    
    # Заголовок
    ws_main['A1'] = "СИСТЕМА КОНТРОЛЯ ПРОИЗВОДСТВА"
    ws_main['A1'].font = Font(bold=True, size=18, color="FFFFFF")
    ws_main['A1'].fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type="solid")
    ws_main.merge_cells('A1:P1')
    ws_main['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_main.row_dimensions[1].height = 30
    
    # Заголовки колонок (строка 2 - закрепленная шапка)
    headers = [
        "Дата", "Станок", "Оператор", "Смена_мин", "Наладка_мин", "Остаток_производства", 
        "Простои_мин", "План_шт", "Факт_шт", "Брак_шт", "Годные_шт", "Доступность_%", 
        "Производительность_%", "Качество_%", "OEE_%", "KPI_%"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws_main.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color=colors['primary'], end_color=colors['primary'], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF')
        )
    
    # ЗАКРЕПЛЕНИЕ ШАПКИ
    ws_main.freeze_panes = "A3"
    
    # Выпадающие списки
    dv_operators = DataValidation(type="list", formula1='"' + ','.join(operators) + '"')
    dv_operators.add("C3:C1000")
    ws_main.add_data_validation(dv_operators)
    
    dv_stations = DataValidation(type="list", formula1='"' + ','.join(stations) + '"')
    dv_stations.add("B3:B1000")
    ws_main.add_data_validation(dv_stations)
    
    # Начальное значение остатка производства
    ws_main['F3'] = 1000  # Стартовый остаток
    
    # Формулы для расчетов
    formulas_row3 = {
        'K3': '=I3-J3',  # Годные
        'L3': '=IFERROR((D3-G3)/D3*100,0)',  # Доступность
        'M3': '=IFERROR(I3/H3*100,0)',  # Производительность  
        'N3': '=IFERROR(K3/I3*100,0)',  # Качество
        'O3': '=IFERROR(L3*M3*N3/10000,0)',  # OEE
        'P3': '=IFERROR(O3*0.5+(100-(E3/D3*100))*0.2+N3*0.15+90*0.15,0)'  # KPI
    }
    
    for cell_ref, formula in formulas_row3.items():
        ws_main[cell_ref] = formula
    
    # Копируем формулы на 50 строк
    for row in range(4, 51):
        # Остаток производства = предыдущий остаток - произведено + брак
        ws_main.cell(row=row, column=6, value=f'=IFERROR(MAX(0,F{row-1}-I{row-1}+J{row-1}),F{row-1})')
        
        # Остальные формулы
        ws_main.cell(row=row, column=11, value=f'=I{row}-J{row}')  # Годные
        ws_main.cell(row=row, column=12, value=f'=IFERROR((D{row}-G{row})/D{row}*100,0)')  # Доступность
        ws_main.cell(row=row, column=13, value=f'=IFERROR(I{row}/H{row}*100,0)')  # Производительность
        ws_main.cell(row=row, column=14, value=f'=IFERROR(K{row}/I{row}*100,0)')  # Качество
        ws_main.cell(row=row, column=15, value=f'=IFERROR(L{row}*M{row}*N{row}/10000,0)')  # OEE
        ws_main.cell(row=row, column=16, value=f'=IFERROR(O{row}*0.5+(100-(E{row}/D{row}*100))*0.2+N{row}*0.15+90*0.15,0)')  # KPI
    
    # Условное форматирование
    # OEE
    oee_range = "O3:O50"
    oee_rule = ColorScaleRule(
        start_type='num', start_value=0, start_color='EF4444',
        mid_type='num', mid_value=75, mid_color='F59E0B', 
        end_type='num', end_value=100, end_color='10B981'
    )
    ws_main.conditional_formatting.add(oee_range, oee_rule)
    
    # KPI
    kpi_range = "P3:P50"
    kpi_rule = ColorScaleRule(
        start_type='num', start_value=0, start_color='EF4444',
        mid_type='num', mid_value=80, mid_color='F59E0B',
        end_type='num', end_value=100, end_color='10B981'
    )
    ws_main.conditional_formatting.add(kpi_range, kpi_rule)
    
    # Ширина колонок
    column_widths = {
        'A': 12, 'B': 16, 'C': 14, 'D': 10, 'E': 12, 'F': 18, 'G': 12, 'H': 10, 
        'I': 10, 'J': 10, 'K': 12, 'L': 14, 'M': 16, 'N': 12, 'O': 8, 'P': 8
    }
    
    for col, width in column_widths.items():
        ws_main.column_dimensions[col].width = width
    
    # Тестовые данные
    test_data = [
        [date.today(), "Doosan_Yashana", "Andrey", 480, 60, None, 20, 50, 45, 2],
        [date.today(), "Doosan_Hadasha", "Denis", 480, 45, None, 15, 40, 38, 1],
        [date.today(), "Doosan_3", "Daniel", 480, 90, None, 30, 30, 28, 3]
    ]
    
    for row_idx, row_data in enumerate(test_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            if value is not None:
                ws_main.cell(row=row_idx, column=col_idx, value=value)
    
    # ============ ОТДЕЛЬНЫЕ ЛИСТЫ ДЛЯ КАЖДОГО СТАНКА ============
    print("🏭 Создаем отдельные листы для станков...")
    
    for station in stations:
        print(f"   📊 Создаем лист для {station}...")
        ws_station = wb.create_sheet(station)
        
        # Заголовок станка
        ws_station['A1'] = f"СТАНОК: {station.replace('_', ' ')}"
        ws_station['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws_station['A1'].fill = PatternFill(start_color=colors['accent'], end_color=colors['accent'], fill_type="solid")
        ws_station.merge_cells('A1:M1')
        ws_station['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws_station.row_dimensions[1].height = 25
        
        # Заголовки для станка
        station_headers = [
            "Дата", "Оператор", "Смена_мин", "Наладка_мин", "Остаток_произв", 
            "Простои_мин", "План_шт", "Факт_шт", "Брак_шт", "OEE_%", "KPI_%", "Эффективность", "Статус"
        ]
        
        for col, header in enumerate(station_headers, 1):
            cell = ws_station.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=colors['primary'], end_color=colors['primary'], fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Закрепление шапки
        ws_station.freeze_panes = "A3"
        
        # Выпадающий список операторов
        dv_ops = DataValidation(type="list", formula1='"' + ','.join(operators) + '"')
        dv_ops.add("B3:B100")
        ws_station.add_data_validation(dv_ops)
        
        # Формулы для станка
        ws_station['E3'] = 500  # Начальный остаток для станка
        ws_station['J3'] = f'=IFERROR(AVERAGEIF(Общие_данные!B:B,"{station}",Общие_данные!O:O),0)'  # Средний OEE
        ws_station['K3'] = f'=IFERROR(AVERAGEIF(Общие_данные!B:B,"{station}",Общие_данные!P:P),0)'  # Средний KPI
        ws_station['L3'] = '=IF(J3>=85,"Отлично",IF(J3>=75,"Хорошо",IF(J3>=65,"Средне","Плохо")))'
        ws_station['M3'] = '=IF(J3>=85,"🟢",IF(J3>=75,"🟡",IF(J3>=65,"🟠","🔴")))'
        
        # Копируем формулы
        for row in range(4, 31):
            ws_station.cell(row=row, column=5, value=f'=IFERROR(MAX(0,E{row-1}-H{row-1}+I{row-1}),E{row-1})')
            ws_station.cell(row=row, column=10, value=f'=IFERROR((C{row}-F{row})/C{row}*H{row}/G{row}*(H{row}-I{row})/H{row}*100,0)')
            ws_station.cell(row=row, column=11, value=f'=IFERROR(J{row}*0.8+20,0)')
            ws_station.cell(row=row, column=12, value=f'=IF(J{row}>=85,"Отлично",IF(J{row}>=75,"Хорошо",IF(J{row}>=65,"Средне","Плохо")))')
            ws_station.cell(row=row, column=13, value=f'=IF(J{row}>=85,"🟢",IF(J{row}>=75,"🟡",IF(J{row}>=65,"🟠","🔴")))')
        
        # Ширина колонок для станка
        for col in range(1, 14):
            ws_station.column_dimensions[chr(64 + col)].width = 14
    
    # ============ KPI ПО ОПЕРАТОРАМ ============
    print("👥 Создаем KPI по операторам...")
    ws_operators = wb.create_sheet("KPI_Операторов")
    
    # Заголовок
    ws_operators['A1'] = "KPI ПО ОПЕРАТОРАМ"
    ws_operators['A1'].font = Font(bold=True, size=18, color="FFFFFF")
    ws_operators['A1'].fill = PatternFill(start_color=colors['success'], end_color=colors['success'], fill_type="solid")
    ws_operators.merge_cells('A1:J1')
    ws_operators['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_operators.row_dimensions[1].height = 30
    
    # Заголовки KPI операторов
    op_headers = [
        "Оператор", "Всего_смен", "Средний_OEE_%", "Средний_KPI_%", "Всего_деталей", 
        "Общий_брак", "Лучший_станок", "Эффективность", "Рейтинг", "Статус"
    ]
    
    for col, header in enumerate(op_headers, 1):
        cell = ws_operators.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=colors['primary'], end_color=colors['primary'], fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Данные по операторам
    for row, operator in enumerate(operators, 3):
        ws_operators.cell(row=row, column=1, value=operator)
        ws_operators.cell(row=row, column=2, value=f'=COUNTIF(Общие_данные!C:C,"{operator}")')
        ws_operators.cell(row=row, column=3, value=f'=IFERROR(AVERAGEIF(Общие_данные!C:C,"{operator}",Общие_данные!O:O),0)')
        ws_operators.cell(row=row, column=4, value=f'=IFERROR(AVERAGEIF(Общие_данные!C:C,"{operator}",Общие_данные!P:P),0)')
        ws_operators.cell(row=row, column=5, value=f'=SUMIF(Общие_данные!C:C,"{operator}",Общие_данные!I:I)')
        ws_operators.cell(row=row, column=6, value=f'=SUMIF(Общие_данные!C:C,"{operator}",Общие_данные!J:J)')
        ws_operators.cell(row=row, column=7, value=f'=IFERROR(INDEX(Общие_данные!B:B,MATCH(MAXIFS(Общие_данные!O:O,Общие_данные!C:C,"{operator}"),Общие_данные!O:O,0)),"N/A")')
        ws_operators.cell(row=row, column=8, value=f'=IF(D{row}>=85,"Высокая",IF(D{row}>=75,"Средняя","Низкая"))')
        ws_operators.cell(row=row, column=9, value=f'=IF(D{row}>=90,"🏆 Топ",IF(D{row}>=80,"⭐ Хорошо",IF(D{row}>=70,"✅ Средне","⚠️ Нужно улучшить")))')
        ws_operators.cell(row=row, column=10, value=f'=IF(D{row}>=85,"🟢",IF(D{row}>=75,"🟡","🔴"))')
        
        # Форматирование процентов
        for col in [3, 4]:
            ws_operators.cell(row=row, column=col).number_format = '0.0"%"'
    
    # Ширина колонок
    for col in range(1, 11):
        ws_operators.column_dimensions[chr(64 + col)].width = 16
    
    # ============ СОВРЕМЕННЫЙ DASHBOARD ============
    print("🎛️ Создаем современный дашборд...")
    ws_dash = wb.create_sheet("Dashboard")
    
    # Главный заголовок
    ws_dash['A1'] = "🏭 ПАНЕЛЬ УПРАВЛЕНИЯ ПРОИЗВОДСТВОМ"
    ws_dash['A1'].font = Font(bold=True, size=20, color="FFFFFF")
    ws_dash['A1'].fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type="solid")
    ws_dash.merge_cells('A1:H1')
    ws_dash['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[1].height = 35
    
    # Время обновления
    ws_dash['A3'] = "🕐 Последнее обновление:"
    ws_dash['A3'].font = Font(bold=True, size=12)
    ws_dash['D3'] = "=NOW()"
    ws_dash['D3'].number_format = 'DD.MM.YYYY HH:MM:SS'
    ws_dash['D3'].font = Font(bold=True, color=colors['primary'])
    
    # Ключевые показатели
    dash_metrics = [
        # Основные KPI
        ("A5", "📊 ОБЩИЙ OEE:", "C5", "=IFERROR(AVERAGE(Общие_данные!O:O),0)", "0.0\"%\""),
        ("A6", "📈 ОБЩИЙ KPI:", "C6", "=IFERROR(AVERAGE(Общие_данные!P:P),0)", "0.0\"%\""),
        ("A7", "🏭 Активные станки:", "C7", f"={len(stations)}", "0"),
        ("A8", "👥 Операторы в смене:", "C8", "=COUNTA(Общие_данные!C3:C50)-COUNTBLANK(Общие_данные!C3:C50)", "0"),
        
        # Производство
        ("E5", "🔧 Всего произведено:", "G5", "=SUM(Общие_данные!I:I)", "0"),
        ("E6", "❌ Общий брак:", "G6", "=SUM(Общие_данные!J:J)", "0"),
        ("E7", "📦 Остаток производства:", "G7", "=SUM(Общие_данные!F:F)", "0"),
        ("E8", "⚡ Средняя эффективность:", "G8", "=IFERROR((SUM(Общие_данные!I:I)-SUM(Общие_данные!J:J))/SUM(Общие_данные!I:I)*100,0)", "0.0\"%\""),
        
        # Топы
        ("A11", "🏆 Лучший по OEE:", "C11", "=INDEX(Общие_данные!C:C,MATCH(MAX(Общие_данные!O:O),Общие_данные!O:O,0))", ""),
        ("A12", "⭐ Лучший по KPI:", "C12", "=INDEX(Общие_данные!C:C,MATCH(MAX(Общие_данные!P:P),Общие_данные!P:P,0))", ""),
        ("A13", "🚀 Топ станок:", "C13", "=INDEX(Общие_данные!B:B,MATCH(MAX(Общие_данные!O:O),Общие_данные!O:O,0))", ""),
        
        ("E11", "⚠️ Требует внимания:", "G11", "=INDEX(Общие_данные!B:B,MATCH(MIN(Общие_данные!O:O),Общие_данные!O:O,0))", ""),
        ("E12", "🔧 Высокий брак:", "G12", "=INDEX(Общие_данные!C:C,MATCH(MAX(Общие_данные!J:J),Общие_данные!J:J,0))", ""),
        ("E13", "📉 Низкая эффективность:", "G13", "=INDEX(Общие_данные!B:B,MATCH(MIN(Общие_данные!P:P),Общие_данные!P:P,0))", ""),
    ]
    
    for item in dash_metrics:
        # Метка
        ws_dash[item[0]] = item[1]
        ws_dash[item[0]].font = Font(bold=True, size=11)
        
        # Значение
        ws_dash[item[2]] = item[3]
        ws_dash[item[2]].font = Font(bold=True, size=11, color=colors['primary'])
        if len(item) > 4 and item[4]:
            ws_dash[item[2]].number_format = item[4]
    
    # Секционные заголовки
    sections = [
        ("A4", "📊 ОСНОВНЫЕ ПОКАЗАТЕЛИ", colors['success']),
        ("A10", "🏆 ЛИДЕРЫ И ОТСТАЮЩИЕ", colors['warning'])
    ]
    
    for section in sections:
        ws_dash[section[0]] = section[1]
        ws_dash[section[0]].font = Font(bold=True, color="FFFFFF", size=12)
        ws_dash[section[0]].fill = PatternFill(start_color=section[2], end_color=section[2], fill_type="solid")
        ws_dash.merge_cells(f'{section[0]}:{chr(ord(section[0][0])+7)}{section[0][1:]}')
        ws_dash[section[0]].alignment = Alignment(horizontal="center")
    
    # Ширина колонок дашборда
    for col in range(1, 9):
        ws_dash.column_dimensions[chr(64 + col)].width = 18
    
    # ============ ГРАФИК ГАНТА ============
    print("📅 Создаем график Ганта...")
    ws_gantt = wb.create_sheet("График_Ганта")
    
    # Заголовок
    ws_gantt['A1'] = "📅 ГРАФИК ГАНТА - ПРОИЗВОДСТВЕННОЕ ПЛАНИРОВАНИЕ"
    ws_gantt['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws_gantt['A1'].fill = PatternFill(start_color=colors['accent'], end_color=colors['accent'], fill_type="solid")
    ws_gantt.merge_cells('A1:L1')
    ws_gantt['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_gantt.row_dimensions[1].height = 30
    
    # Заголовки Ганта
    gantt_headers = [
        "Станок", "Оператор", "Статус", "Начало", "Окончание", "Прогресс_%", 
        "Остаток_ч", "Приоритет", "Заказ", "Сложность", "Примечания", "Индикатор"
    ]
    
    for col, header in enumerate(gantt_headers, 1):
        cell = ws_gantt.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=colors['primary'], end_color=colors['primary'], fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Данные для Ганта
    gantt_data = [
        ["Doosan_Yashana", "Andrey", "🟢 Активно", "08:00", "16:00", 75, 2.5, "Высокий", "#001", "Сложная", "Фрезеровка корпуса", "🟢"],
        ["Doosan_Hadasha", "Denis", "✅ Завершено", "08:00", "14:30", 100, 0, "Средний", "#002", "Стандарт", "Токарная обработка", "✅"],
        ["Doosan_3", "Daniel", "🟡 Задержка", "09:00", "17:00", 45, 4.5, "Высокий", "#003", "Высокая", "Наладка остановлена", "🟡"],
        ["Pinnacle_Gdola", "Kirill", "🟢 Активно", "08:30", "16:30", 85, 1.5, "Средний", "#004", "Средняя", "Серийное производство", "🟢"],
        ["Mitsubishi", "Slava", "🔵 Ожидание", "10:00", "18:00", 15, 7.0, "Низкий", "#005", "Простая", "Ожидание материала", "🔵"],
        ["JohnFord", "Arkady", "🟠 Настройка", "08:00", "15:00", 30, 5.0, "Высокий", "#006", "Сложная", "Первичная наладка", "🟠"],
        ["Okuma", "Andrey", "⏸️ Пауза", "14:00", "22:00", 60, 3.0, "Средний", "#007", "Средняя", "Обеденный перерыв", "⏸️"]
    ]
    
    for row, data in enumerate(gantt_data, 3):
        for col, value in enumerate(data, 1):
            cell = ws_gantt.cell(row=row, column=col, value=value)
            if col == 6 and isinstance(value, (int, float)):  # Прогресс
                cell.number_format = '0"%"'
            
            # Цветовое кодирование статуса
            if col == 3:  # Статус
                if "🟢" in str(value):
                    cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type="solid")
                elif "🟡" in str(value):
                    cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type="solid")
                elif "🔴" in str(value):
                    cell.fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type="solid")
    
    # Ширина колонок Ганта
    for col in range(1, 13):
        ws_gantt.column_dimensions[chr(64 + col)].width = 15
    
    # ============ СВОДКА ПО ПРОИЗВОДСТВУ ============
    print("📈 Создаем сводку по производству...")
    ws_summary = wb.create_sheet("Сводка_производства")
    
    # Заголовок
    ws_summary['A1'] = "📈 ОБЩАЯ СВОДКА ПРОИЗВОДСТВА"
    ws_summary['A1'].font = Font(bold=True, size=18, color="FFFFFF")
    ws_summary['A1'].fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type="solid")
    ws_summary.merge_cells('A1:H1')
    ws_summary['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 30
    
    # Сводка по станкам
    summary_headers = ["Станок", "Ср_OEE_%", "Ср_KPI_%", "Всего_произведено", "Общий_брак", "Эффективность", "Рейтинг", "Статус"]
    
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=colors['primary'], end_color=colors['primary'], fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Данные по станкам
    for row, station in enumerate(stations, 4):
        ws_summary.cell(row=row, column=1, value=station.replace('_', ' '))
        ws_summary.cell(row=row, column=2, value=f'=IFERROR(AVERAGEIF(Общие_данные!B:B,"{station}",Общие_данные!O:O),0)')
        ws_summary.cell(row=row, column=3, value=f'=IFERROR(AVERAGEIF(Общие_данные!B:B,"{station}",Общие_данные!P:P),0)')
        ws_summary.cell(row=row, column=4, value=f'=SUMIF(Общие_данные!B:B,"{station}",Общие_данные!I:I)')
        ws_summary.cell(row=row, column=5, value=f'=SUMIF(Общие_данные!B:B,"{station}",Общие_данные!J:J)')
        ws_summary.cell(row=row, column=6, value=f'=IFERROR((D{row}-E{row})/D{row}*100,0)')
        ws_summary.cell(row=row, column=7, value=f'=IF(B{row}>=85,"🏆",IF(B{row}>=75,"⭐",IF(B{row}>=65,"✅","⚠️")))')
        ws_summary.cell(row=row, column=8, value=f'=IF(B{row}>=85,"🟢 Отлично",IF(B{row}>=75,"🟡 Хорошо","🔴 Требует внимания"))')
        
        # Форматирование
        for col in [2, 3, 6]:
            ws_summary.cell(row=row, column=col).number_format = '0.0"%"'
    
    # Общие итоги
    ws_summary['A12'] = "ОБЩИЕ ИТОГИ:"
    ws_summary['A12'].font = Font(bold=True, size=14, color=colors['header'])
    ws_summary['A12'].fill = PatternFill(start_color=colors['light'], end_color=colors['light'], fill_type="solid")
    
    totals = [
        ("A13", "📊 Общий OEE производства:", "D13", "=AVERAGE(B4:B10)"),
        ("A14", "📈 Общий KPI производства:", "D14", "=AVERAGE(C4:C10)"),
        ("A15", "🔧 Всего произведено:", "D15", "=SUM(D4:D10)"),
        ("A16", "❌ Общий брак:", "D16", "=SUM(E4:E10)"),
        ("A17", "⚡ Общая эффективность:", "D17", "=AVERAGE(F4:F10)")
    ]
    
    for total in totals:
        ws_summary[total[0]] = total[1]
        ws_summary[total[0]].font = Font(bold=True)
        ws_summary[total[2]] = total[3]
        ws_summary[total[2]].font = Font(bold=True, color=colors['primary'])
        
        if "OEE" in total[1] or "KPI" in total[1] or "эффективность" in total[1]:
            ws_summary[total[2]].number_format = '0.0"%"'
    
    # Ширина колонок
    for col in range(1, 9):
        ws_summary.column_dimensions[chr(64 + col)].width = 18
    
    # ============ СОХРАНЕНИЕ ============
    filename = f"Advanced_OEE_KPI_v10_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(os.getcwd(), filename)
    
    try:
        wb.save(filepath)
        print(f"✅ Улучшенный Excel файл создан: {filename}")
        print(f"📁 Размер файла: {os.path.getsize(filepath) / 1024:.1f} KB")
        
        print("\n🚀 НОВЫЕ ВОЗМОЖНОСТИ v10.0:")
        print("   📊 Отдельные листы данных для каждого станка")
        print("   👥 Детальный KPI по каждому оператору")
        print("   🔄 Остаток производства с компенсацией брака")
        print("   📝 Выпадающие списки для операторов")
        print("   📌 Закрепленная шапка таблицы")
        print("   📈 Общий OEE и KPI по всему производству")
        print("   📅 Современный график Ганта")
        print("   🎛️ Улучшенный дашборд с современным дизайном")
        print("   🎨 Цветовое кодирование и условное форматирование")
        print("   📊 Автоматические сводки и итоги")
        
        print("\n📋 СТРУКТУРА ФАЙЛА:")
        print("   🔸 Общие_данные - основная таблица с закрепленной шапкой")
        print(f"   🔸 {len(stations)} листов станков - индивидуальные данные")
        print("   🔸 KPI_Операторов - детальная аналитика по сотрудникам")
        print("   🔸 Dashboard - современная панель управления")
        print("   🔸 График_Ганта - производственное планирование")
        print("   🔸 Сводка_производства - общие итоги и рейтинги")
        
        return filepath
        
    except Exception as e:
        print(f"❌ Ошибка при создании: {e}")
        return None

def main():
    """Основная функция"""
    print("=" * 80)
    print("🏭 УЛУЧШЕННАЯ СИСТЕМА OEE/KPI v10.0")
    print("ОТДЕЛЬНЫЕ ЛИСТЫ СТАНКОВ | KPI ОПЕРАТОРОВ | СОВРЕМЕННЫЙ ДИЗАЙН")
    print("=" * 80)
    
    try:
        import openpyxl
        print("✅ Библиотеки загружены")
    except ImportError:
        print("❌ Установите: pip install openpyxl")
        return
    
    filepath = create_advanced_oee_kpi()
    
    if filepath:
        print(f"\n🎉 УСПЕШНО! Улучшенная система создана!")
        print(f"📂 Файл: {os.path.basename(filepath)}")
        print(f"\n⭐ Все ваши требования выполнены:")
        print(f"   ✅ Отдельные листы для каждого станка")
        print(f"   ✅ KPI по каждому оператору")
        print(f"   ✅ Остаток производства вместо фиксированного")
        print(f"   ✅ Выпадающие списки операторов")
        print(f"   ✅ Закрепленная шапка")
        print(f"   ✅ Общий OEE/KPI по производству")
        print(f"   ✅ График Ганта")
        print(f"   ✅ Современный дашборд")
        print(f"\n🚀 Готово к использованию!")
    else:
        print("\n❌ Не удалось создать файл")

if __name__ == "__main__":
    main()
