import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

def create_three_sheets_test():
    """
    Тест: создаем 'Данные' + 'Ввод_данных' + 'Сводка'.
    """
    try:
        print("--- Тест: создание ТРЕХ листов ---")
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
        
        # === ЧАСТЬ 1: Лист "Данные" (работает) ===
        print("1. Создаю лист 'Данные'...")
        ws_data = wb.create_sheet("Данные", 0)
        data_headers = [ "Дата", "Станок", "Оператор", "Смена", "Наладка", "Производство", "Простои", "План_шт", "Факт_шт", "Брак", "Годные", "Доступность", "Производительность", "Качество", "OEE", "Доля_наладки", "Качество_наладки", "KPI" ]
        ws_data.append(data_headers)
        header_font = Font(bold=True, size=11); header_fill = PatternFill(start_color="FFDDEBF7", fill_type="solid"); header_align = Alignment(horizontal="center", vertical="center")
        for cell in ws_data["1:1"]: cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align
        formulas = { 'K2': '=I2-J2', 'L2': '=IF(D2>0,(D2-G2)/D2*100,0)', 'M2': '=IF(H2>0,I2/H2*100,0)', 'N2': '=IF(I2>0,K2/I2*100,0)', 'O2': '=IFERROR(L2*M2*N2/10000,0)', 'P2': '=IF(D2>0,E2/D2*100,0)', 'Q2': '=IF(I2>0,(I2-J2)/I2*100,0)', 'R2': '=IFERROR(O2*0.5+(100-P2)*0.2+Q2*0.15+90*0.15,0)' }
        for cell, formula in formulas.items(): ws_data[cell] = formula
        print("   ... 'Данные' готов.")

        # === ЧАСТЬ 2: Лист "Ввод_данных" (работает) ===
        print("2. Создаю лист 'Ввод_данных'...")
        ws_input = wb.create_sheet("Ввод_данных", 1)
        ws_input['A1'] = "ФОРМА ВВОДА ДАННЫХ ЗА СМЕНУ"
        input_labels = {'A3':"Дата:",'A4':"Станок:",'A5':"Оператор:",'A7':"ВРЕМЕННЫЕ ДАННЫЕ (в минутах):",'A8':"Продолжительность смены:",'A9':"Время наладки:",'A10':"Время производства:",'A11':"Время простоев:",'A13':"ПРОИЗВОДСТВЕННЫЕ ДАННЫЕ:",'A14':"Запланировано деталей:",'A15':"Фактически произведено:",'A16':"Из них брак:",'A18':"КНОПКА 'ДОБАВИТЬ ЗАПИСЬ'"}
        for cell,text in input_labels.items(): ws_input[cell]=text
        ws_input['C8']=480
        dv_stanok=DataValidation(type="list",formula1='"CNC-1,CNC-2,CNC-3,CNC-4,CNC-5,CNC-6,CNC-7"',allow_blank=True)
        ws_input.add_data_validation(dv_stanok); dv_stanok.add('C4')
        dv_operator=DataValidation(type="list",formula1='"Профи-фрезеровщик,Средний специалист,Слабый фрезеровщик,Помощник,Новый токарь"',allow_blank=True)
        ws_input.add_data_validation(dv_operator); dv_operator.add('C5')
        print("   ... 'Ввод_данных' готов.")

        # === ЧАСТЬ 3: Добавляем лист "Сводка" ===
        print("3. Создаю лист 'Сводка'...")
        ws_summary = wb.create_sheet("Сводка", 2)
        summary_headers = ["Станок", "Ср_OEE", "Ср_Доступность", "Ср_Производительность", "Ср_Качество", "Ср_Доля_наладки", "Общий_KPI", "Статус"]
        ws_summary.append(summary_headers)
        for cell in ws_summary["1:1"]: cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align
        stanki = [f"CNC-{i}" for i in range(1, 8)]
        for i, stanok in enumerate(stanki, start=2):
            ws_summary[f'A{i}'] = stanok
            ws_summary[f'B{i}'] = f'=IFERROR(AVERAGEIF(Данные!B:B,A{i},Данные!O:O),"")'
            ws_summary[f'C{i}'] = f'=IFERROR(AVERAGEIF(Данные!B:B,A{i},Данные!L:L),"")'
            ws_summary[f'D{i}'] = f'=IFERROR(AVERAGEIF(Данные!B:B,A{i},Данные!M:M),"")'
            ws_summary[f'E{i}'] = f'=IFERROR(AVERAGEIF(Данные!B:B,A{i},Данные!N:N),"")'
            ws_summary[f'F{i}'] = f'=IFERROR(AVERAGEIF(Данные!B:B,A{i},Данные!P:P),"")'
            ws_summary[f'G{i}'] = f'=IFERROR(AVERAGEIF(Данные!B:B,A{i},Данные!R:R),"")'
            ws_summary[f'H{i}'] = f'=IF(G{i}<>"",IF(G{i}>=85,"🏆 Отлично",IF(G{i}>=75,"✅ Хорошо",IF(G{i}>=65,"⚠️ Средне","❌ Плохо"))),"")'
        print("   ... 'Сводка' готов.")
        
        # === ЧАСТЬ 4: Сохранение ===
        filename = "test_three_sheets.xlsx"
        print(f"\nСохраняю файл как '{filename}'...")
        wb.save(filename)
        
        print("\n" + "="*50)
        print(f"✅ УСПЕХ! Тест завершен.")
        print(f"Файл '{filename}' должен был создаться с ТРЕМЯ листами.")
        print("="*50)

    except Exception as e:
        import traceback
        print("\n" + "!"*50)
        print(f"❌ Произошла ошибка: {e}")
        traceback.print_exc()
        print("!"*50)

# --- Запуск ---
create_three_sheets_test()
print("\nНажмите Enter для выхода.")
input()