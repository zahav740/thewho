import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def create_just_data_sheet_test():
    """
    Тест: создаем ТОЛЬКО лист 'Данные'.
    Это поможет найти точное место сбоя.
    """
    try:
        print("--- Тест: создание ОДНОГО листа 'Данные' ---")

        wb = openpyxl.Workbook()
        
        # Удаляем лист по умолчанию, чтобы он не мешал
        default_sheet = wb.active
        wb.remove(default_sheet)
        print("Стандартный лист удален.")
        
        # Создаем только наш целевой лист
        ws_data = wb.create_sheet("Данные", 0)
        print("Лист 'Данные' создан.")

        # --- Заполнение листа ---
        data_headers = [
            "Дата", "Станок", "Оператор", "Смена", "Наладка", "Производство",
            "Простои", "План_шт", "Факт_шт", "Брак", "Годные", "Доступность",
            "Производительность", "Качество", "OEE", "Доля_наладки", "Качество_наладки", "KPI"
        ]
        ws_data.append(data_headers)
        print("Заголовки добавлены.")
        
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="FFDDEBF7", end_color="FFDDEBF7", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        for cell in ws_data["1:1"]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        print("Заголовки отформатированы.")

        formulas = {
            'K2': '=I2-J2', 'L2': '=IF(D2>0,(D2-G2)/D2*100,0)',
            'M2': '=IF(H2>0,I2/H2*100,0)', 'N2': '=IF(I2>0,K2/I2*100,0)',
            'O2': '=IFERROR(L2*M2*N2/10000,0)', 'P2': '=IF(D2>0,E2/D2*100,0)',
            'Q2': '=IF(I2>0,(I2-J2)/I2*100,0)', 'R2': '=IFERROR(O2*0.5+(100-P2)*0.2+Q2*0.15+90*0.15,0)'
        }
        for cell, formula in formulas.items():
            ws_data[cell] = formula
        print("Формулы добавлены.")

        filename = "test_data_sheet_only.xlsx"
        print(f"Сохраняю файл как '{filename}'...")
        wb.save(filename)
        
        print("\n" + "="*50)
        print(f"✅ УСПЕХ! Тест завершен.")
        print(f"Файл '{filename}' должен был создаться.")
        print("="*50)

    except Exception as e:
        import traceback
        print("\n" + "!"*50)
        print(f"❌ Произошла ошибка: {e}")
        traceback.print_exc()
        print("!"*50)

# --- Запуск ---
create_just_data_sheet_test()
print("\nНажмите Enter для выхода.")
input()