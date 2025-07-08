#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Тестовая версия генератора Excel
"""

print("🔥 Запуск генератора Excel OEE/KPI...")

try:
    import openpyxl
    print("✅ openpyxl загружен успешно")
except ImportError as e:
    print(f"❌ Ошибка импорта openpyxl: {e}")
    print("💡 Установите: pip install openpyxl")
    exit(1)

try:
    from datetime import datetime, date
    import os
    print("✅ Стандартные библиотеки загружены")
except ImportError as e:
    print(f"❌ Ошибка стандартных библиотек: {e}")
    exit(1)

print("🚀 Создание Excel файла...")

# Создаем простую книгу Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test"

# Добавляем тестовые данные
ws['A1'] = "ТЕСТ ГЕНЕРАТОРА OEE/KPI"
ws['A2'] = f"Создано: {datetime.now()}"
ws['A3'] = "Это тестовая версия для проверки работоспособности"

# Сохраняем файл
filename = f"Test_OEE_KPI_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
try:
    wb.save(filename)
    print(f"✅ Тестовый файл создан: {filename}")
    
    # Проверяем размер файла
    if os.path.exists(filename):
        size = os.path.getsize(filename)
        print(f"📁 Размер файла: {size} байт")
        print("🎉 Тест прошел успешно!")
    else:
        print("❌ Файл не найден после создания")
        
except Exception as e:
    print(f"❌ Ошибка при сохранении: {e}")

print("✅ Тест завершен")
